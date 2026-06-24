import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st
from pymongo import MongoClient
import datetime
from matplotlib.animation import FuncAnimation
import time
import datetime
from datetime import datetime
from streamlit_autorefresh import st_autorefresh
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np
from st_aggrid import AgGrid
from st_aggrid import GridOptionsBuilder, GridUpdateMode, DataReturnMode, JsCode
import asyncio
import motor.motor_asyncio
from motor.motor_asyncio import AsyncIOMotorClient
import subprocess
import os
import signal
import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import range_boundaries
import commentjson as json
from io import BytesIO
# from st_select_engine import load_config
# from st_main import pages
# from pages.st_select_engine import load_config
import pytz
import datetime
import scipy
from scipy.interpolate import interp1d
import sys
from collections import OrderedDict
# from collections import OrderedDict

# sys.path.insert(0, r"C:\Users\KieranCalder\Code\AerotecTestCell")

# st.set_page_config(layout="wide", page_title="Dashboard", page_icon=st.session_state.page_icon, initial_sidebar_state="collapsed")
# Clear the session state at the start of the app
# if 'clear_state' not in st.session_state:
#     st.session_state.clear()
#     st.session_state['clear_state'] = True  # Prevent further clearing in the same session

st.set_page_config(layout="wide", page_title="Engine Dashboard", page_icon=st.session_state.page_icon, initial_sidebar_state="collapsed")

# Ensure the user has selected an engine
if 'selected_engine' not in st.session_state: #or st.session_state.selected_engine is None:
    st.warning("No engine selected. Redirecting to select engine page...")
    time.sleep(1)
    st.switch_page(r"C:\Users\KieranCalder\Code\AerotecTestCell\pages\st_select_engine.py")  # Redirect back to engine selection

elif 'page_icon' not in st.session_state: #or st.session_state.selected_engine is None:
    st.warning("Initialization incomplete. Redirecting to main page...")
    time.sleep(1)
    st.switch_page(r"C:\Users\KieranCalder\Code\AerotecTestCell\st_main.py")  # Redirect back to engine selection

elif 'logo' not in st.session_state: 
    st.warning("Initialization incomplete. Redirecting to main page...")
    time.sleep(1)
    st.switch_page(r"C:\Users\KieranCalder\Code\AerotecTestCell\st_main.py")  # Redirect back to engine selection

# Add Aerotec Logo
st.logo(st.session_state.logo)

# Use the full page instead of a narrow central column
# st.set_page_config(layout="wide", page_title=f"{st.session_state.selected_engine} Dashboard", page_icon=st.session_state.page_icon, initial_sidebar_state="collapsed")


global engine_type, dataOnly
dataOnly = "Data Only"
engine_type = st.session_state.selected_engine
# prop_type = st.session_state.selected_prop
prop_type = st.session_state.get("selected_prop")

def load_config():
    """Load JSON configuration file."""
    with open(r"C:\Users\KieranCalder\Code\AerotecTestCell\Dashboard Config.json", "r") as f:
        return json.load(f, object_pairs_hook=OrderedDict) # Use OrderedDict to maintain order

engine_config = load_config()

global cylinderCount, turboStatus
cylinderCount = st.session_state.get('cylinders')
turboStatus = st.session_state.get('turbo')

# Define Global Excel Doc Thresholds and Sort by Columns
global thresholds 
thresholds = {
    2:(st.session_state.get('minMP'), st.session_state.get('nominalMP'), st.session_state.get('highMP'), st.session_state.get('maxMP')), 
    4:(st.session_state.get('minOilTemp'), st.session_state.get('nominalOilTemp'), st.session_state.get('highOilTemp'), st.session_state.get('maxOilTemp')), 
    5:(st.session_state.get('minFrontOilPressure'), st.session_state.get('nominalFrontOilPressure'), st.session_state.get('highFrontOilPressure'), st.session_state.get('maxFrontOilPressure')),
    6:(st.session_state.get('minRearOilPressure'), st.session_state.get('nominalRearOilPressure'), st.session_state.get('highRearOilPressure'), st.session_state.get('maxRearOilPressure')), 
    7:(st.session_state.get('minMeteredFuelPressure'), st.session_state.get('nominalMeteredFuelPressure'), st.session_state.get('highMeteredFuelPressure'), st.session_state.get('maxMeteredFuelPressure')),
    8:(st.session_state.get('minUnmeteredFuelPressure'), st.session_state.get('nominalUnmeteredFuelPressure'), st.session_state.get('highUnmeteredFuelPressure'), st.session_state.get('maxUnmeteredFuelPressure')),
    9:(st.session_state.get('minEGT'), st.session_state.get('nominalEGT'), st.session_state.get('highEGT'), st.session_state.get('maxEGT')),
    10:(st.session_state.get('minEGT'), st.session_state.get('nominalEGT'), st.session_state.get('highEGT'), st.session_state.get('maxEGT')),
    11:(st.session_state.get('minEGT'), st.session_state.get('nominalEGT'), st.session_state.get('highEGT'), st.session_state.get('maxEGT')),
    12:(st.session_state.get('minEGT'), st.session_state.get('nominalEGT'), st.session_state.get('highEGT'), st.session_state.get('maxEGT')),
    15:(st.session_state.get('minCHT'), st.session_state.get('nominalCHT'), st.session_state.get('highCHT'), st.session_state.get('maxCHT')),
    16:(st.session_state.get('minCHT'), st.session_state.get('nominalCHT'), st.session_state.get('highCHT'), st.session_state.get('maxCHT')),
    17:(st.session_state.get('minCHT'), st.session_state.get('nominalCHT'), st.session_state.get('highCHT'), st.session_state.get('maxCHT')),
    18:(st.session_state.get('minCHT'), st.session_state.get('nominalCHT'), st.session_state.get('highCHT'), st.session_state.get('maxCHT'))
}

refresh_time = 1
st_autorefresh(refresh_time*1000, key="data_refresh")  # Refresh at a specified interval
# # MongoDB credentials and connection string
# username = urllib.parse.quote_plus('Aerotec')  # Add your username here
# password = urllib.parse.quote_plus('1@QWASZX')  # Add your password here

# # MongoDB Atlas Cloud Database Connection
# uri = "mongodb+srv://{}:{}@atlascluster.qrjrg5e.mongodb.net/".format(username, password)
# client = motor.motor_asyncio.AsyncIOMotorClient(uri)
# db = client['RTDO']  # Replace 'MVP50' with your database name

# MongoDB Local Database Connection
uri = "mongodb://localhost:27017/"
client = motor.motor_asyncio.AsyncIOMotorClient(uri)
# db = client['RTDO']

# Initialize db_name in session state if not already done
if 'db_name' not in st.session_state:
    st.session_state.db_name = None

db_str = str(st.session_state.db_name)
db = client[db_str]


# Single collection for all data points
# collection = db['Data']

# Initialize collection_name in session state if not already done:
if 'collection_name' not in st.session_state:
    st.session_state.collection_name = None
    
collection_str = str(st.session_state.collection_name)
collection = db[collection_str]

# --- HELPER: ANCHOR MAPPING ---
def get_row_map(df):
    """
    Creates a dictionary mapping the content of Column A (Labels) to their Row Index.
    Example: {'Spec Oil Pressure': 19, '2200': 15, 'Idle Oil Press': 21}
    """
    # Assuming Column 0 ('Unnamed: 0') holds the labels like "750", "Spec Oil Pressure", etc.
    # We strip whitespace and convert to string to ensure matches found
    mapping = {}
    for idx, value in df['Unnamed: 0'].items():
        if pd.notna(value):
            clean_label = str(value).strip()
            mapping[clean_label] = idx
    return mapping

# --- HELPER: RPM SEQUENCE ---
def get_rpm_sequence():
    """Returns the RPM list in descending order based on prop type"""
    prop_type = st.session_state.get("selected_prop") 
    
    # Common RPMs
    base_rpms = [2000, 1800, 1500, 1300, 1000, 750]
    
    # if "2 Blade" in prop_type:
    if prop_type == "2 Blade (Yellow Tipped)":
        # 2-Blade starts higher
        high_rpms = [2700, 2400, 2200]
    else:
        # 4-Blade starts at 2200
        high_rpms = [2200]
        
    # Combine and sort descending (High -> Low)
    full_sequence = sorted(high_rpms + base_rpms, reverse=True)
    return [str(x) for x in full_sequence] # Return as strings to match Excel labels

@st.cache_data
def test_sheet_df(prop_type):
    # test_sheet = pd.read_csv(r"C:\Users\theca\Documents\Victor\Aerotec\MVP 50\Code\Engine Test Sheet Template.csv")
    # test_sheet = pd.read_excel(r"C:\Users\KieranCalder\Code\AerotecTestCell\Engine Test Sheet Template.xlsx")

    if prop_type == "2 Blade (Yellow Tipped)":
        path = r"C:\Users\KieranCalder\Code\AerotecTestCell\Engine Test Sheet Template (2 Blade).xlsx"
    else:
        path = r"C:\Users\KieranCalder\Code\AerotecTestCell\Engine Test Sheet Template (4 Blade).xlsx"

    test_sheet = pd.read_excel(path)

    return(test_sheet)

# Initialize session state if not already done
if 'test_sheet' not in st.session_state:
    st.session_state.test_sheet = test_sheet_df(prop_type)
    st.toast("Cache Empty. Loading test sheet template...")

# Function to get or initialize the row index for snapshots
# def get_snapshot_index():
#     if 'snapshot_index' not in st.session_state:
#         if prop_type == "2 Blade (Yellow Tipped)":
#             st.session_state.snapshot_index = 17
#         else:
#             st.session_state.snapshot_index = 15
#         # remind user to confirm full throttle RPM
#         st.toast("Confirm full throttle RPM")

#     return st.session_state.snapshot_index

def get_snapshot_index():
    if 'snapshot_index' not in st.session_state:
        # DYNAMICALLY find the start row
        row_map = get_row_map(st.session_state.test_sheet)
        rpm_sequence = get_rpm_sequence()
        start_rpm = rpm_sequence[0] # e.g. "2700" or "2200"
        
        if start_rpm in row_map:
            st.session_state.snapshot_index = row_map[start_rpm]
            st.toast(f"Ready. Starting at {start_rpm} RPM (Row {row_map[start_rpm]})")
        else:
            st.error(f"Could not find start RPM '{start_rpm}' in the Excel sheet labels.")
            st.session_state.snapshot_index = 0

    return st.session_state.snapshot_index




async def fetch_data():
    num_documents = 5
    
    pipeline = [
        {"$sort": {"timestamp": -1}},  # Sort by timestamp, latest first
        {"$limit": num_documents},  # Limit the number of documents
        {"$project": {
            "_id": 0,  # Exclude the _id field
            "timestamp": 1,
            "volts": "$fields.volts",
            "manifold_pressure": "$fields.manifold_pressure",
            "rpm": "$fields.rpm",
            "hp": "$fields.hp",
            "fuel_flow": "$fields.fuel_flow",
            "metered_fuel_pressure": "$fields.metered_fuel_pressure",
            "unmetered_fuel_pressure": "$fields.unmetered_fuel_pressure",
            "front_oil_p": "$fields.front_oil_p",
            "rear_oil_p": "$fields.rear_oil_p", 
            "oil_temperature": "$fields.oil_temperature",
            "amps": "$fields.amps",
            "cdt": "$fields.cdt",
            "tit": "$fields.tit",
            "iat_carb": "$fields.iat_carb",
            "oat": "$fields.oat",
            "mag_drop":"$fields.mag_drop",
            "egt_1": "$fields.egt_1",
            "egt_2": "$fields.egt_2",
            "egt_3": "$fields.egt_3",
            "egt_4": "$fields.egt_4",
            "egt_5": "$fields.egt_5",
            "egt_6": "$fields.egt_6",
            "cht_1": "$fields.cht_1",
            "cht_2": "$fields.cht_2",
            "cht_3": "$fields.cht_3",
            "cht_4": "$fields.cht_4",
            "cht_5": "$fields.cht_5",
            "cht_6": "$fields.cht_6",
        }}
    ]
    
    cursor = collection.aggregate(pipeline)
    data_list = []
    async for document in cursor:
        data_list.append(document)
    
    return data_list


def get_dataframe():
    start_time = time.time()  # Start time for performance measurement

    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        data_list = loop.run_until_complete(fetch_data())
        loop.close()
        
        if not data_list:
            st.write("No data fetched from MongoDB.")
            return pd.DataFrame()
        
        df = pd.DataFrame(data_list)
        
        if df.empty:
            st.write("DataFrame is empty.")
            return df
        
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        
        numeric_fields = [
            'volts', 'manifold_pressure', 'rpm', 'hp', 'fuel_flow',
            'metered_fuel_pressure', 'unmetered_fuel_pressure', 'front_oil_p', 'rear_oil_p' 'oil_temperature',
            'amps', 'cdt', 'iat_carb', 'oat', 'tit',
            'egt_1', 'egt_2', 'egt_3', 'egt_4', 'egt_5', 'egt_6',
            'cht_1', 'cht_2', 'cht_3', 'cht_4', 'cht_5', 'cht_6' 
        ]
        
        # # This commented section is for adding random variables during presentation
        # # Convert numeric fields to floats
        # for field in numeric_fields:
        #     if field in df.columns:
        #         df[field] = pd.to_numeric(df[field], errors='coerce')
        
        # for i in range(len(df)):
        #     for field in numeric_fields:
        #         if field in df.columns and not pd.isnull(df.at[i, field]):
        #             # Simulate changing data by adding a small random variation
        #             df.at[i, field] = round(df.at[i, field] + random.uniform(-5, 5), 2)

        for field in numeric_fields:
            if field in df.columns:
                df[field] = pd.to_numeric(df[field], errors='coerce') #+ random.uniform(-5,5)
            else:
                df[field] = None  # Handle missing fields


        # df['oil_temperature'] = ((df['oil_temperature']) -32 ) * (5/9)
        df['oat'] = ((df['oat']) -32 ) * (5/9)
        df['iat_carb'] = ((df['iat_carb']) -32 ) * (5/9)

        
        # # Step 1: Create a calibration DataFrame
        # calibration_data = {
        #     'Actual_Temp': [500, 450, 400, 350, 300, 250, 200, 150],
        #     'Probe_Reading': [546, 485, 393, 367, 310, 253, 200, 150]
        # }
        # calibration_df = pd.DataFrame(calibration_data)

        # st.write(calibration_df)

        # # Create the interpolation function
        # interp_func = interp1d(calibration_df['Probe_Reading'], calibration_df['Actual_Temp'], fill_value="extrapolate")

        # # Interpolation function with direct match for values <= 200°F
        # def interpolate_cht(cht_value):
        #     if cht_value <= 200:
        #         return cht_value  # Direct match for values below or equal to 200°F
        #     return interp_func(cht_value)  # Use scipy interpolation for other values

        # # Step 3: Apply interpolation to each CHT channel
        # for i in range(1, 7):  # For cht_1 to cht_6
        #     field_name = f'cht_{i}'
        #     if field_name in df.columns:
        #         cht_value = int(df[field_name].iloc[0])
        #         # st.write(cht_value)
        #         # df[f'cht_{i}_calibrated'] = df[field_name].apply(interpolate_cht(cht_value))
        #         df[f'cht_{i}_calibrated'] = int(interpolate_cht(cht_value))


        df['elapsed_time'] = (df['timestamp'] - df['timestamp'].max()).dt.total_seconds()
        
        fetch_time = time.time() - start_time
        st.session_state['data_fetch_time'] = fetch_time
        
        return df
    
    except Exception as e:
        st.write(f"An error occurred: {e}")
        return pd.DataFrame()
    
    
def get_color(field, value):
    """
    Determine the color for the bar based on the value.
    
    Args:
    - value (float): The value to determine the color for.
    
    Returns:
    - str: The color based on the value range.
    """
    if field == 'CHT':
        if value >= st.session_state.get('minCHT') and value <= st.session_state.get('nominalCHT'):
            return 'green'
        elif value >= st.session_state.get('nominalCHT') and value <= st.session_state.get('highCHT'):
            return 'green'
        elif value >= st.session_state.get('highCHT') and value <= (0.9 * st.session_state.get('maxCHT')):
            return 'yellow'
        else:
            return 'red'

    if field == 'EGT':
        if value >= st.session_state.get('minEGT') and value <= st.session_state.get('nominalEGT'):
            return 'blue'
        elif value >= st.session_state.get('nominalEGT') and value <= st.session_state.get('highEGT'):
            return 'blue'
        elif value >= st.session_state.get('highEGT') and value <= (0.9 * st.session_state.get('maxEGT')):
            return 'yellow'
        else:
            return 'red'


# def display_metrics(df, engine_type):
#     """
#     Display metrics as text with labels and units.
    
#     Args:
#     - df (pd.DataFrame): DataFrame containing the data.
#     """
#     # Create rows of columns
#     row1 = st.columns(5)
#     row2 = st.columns(5)
#     row3 = st.columns(5)

#     # Combine all columns into a single list
#     all_columns = row1 + row2 + row3

#     # CSS for custom styling
#     css = """
#     <style>
#     .metric-tile {
#         border-radius: 8px;  /* Rounded corners */
#         padding: 10px;  /* Padding inside the tile */
#         margin: 5px;  /* Margin around the tile */
#         text-align: center;  /* Center-align text */
#     }
#     .metric-label {
#         font-size: 14px;  /* Adjust label font size */
#         text-align: left;  /* Center-align text */
#         color: #ffffff;  /* Darker color for label */
#     }
#     .metric-value {
#         font-size: 18px;  /* Adjust value font size */
#         color: #ffffff;  /* Black color for value */
#     .metric-delta {
#         font-size: 14px;  /* Adjust delta font size */
#         color: #007bff;  /* Blue color for delta */
#     }
#     </style>
#     """

#     st.markdown(css, unsafe_allow_html=True)

#     if not df.empty:
#         if engine_type == dataOnly:
#             metrics = {
#             "OAT": ("oat", "°F"),
#             "IAT Carb": ("iat_carb", "°F"),
#             "Front Oil P": ("front_oil_p", "PSI"),
#             "Rear Oil P": ("rear_oil_p", "PSI"),
#             # Add Upper Deck Pressure
#             "TIT": ("tit", "°F"),
#             "Oil Temp": ("oil_temperature", "°C"),
#             "Fuel Flow": ("fuel_flow", "GPH"),
#             "Metered Fuel P": ("metered_fuel_pressure", "PSI"),
#             "Unmetered Fuel P": ("unmetered_fuel_pressure", "PSI"),
#             "Volts": ("volts", "V"),
#             "CDT": ("cdt", "°F"), 
#             "Mag Drop": ("mag_drop", "")
#             }

#             # if turboStatus == True:
#             #     metrics["TIT"] = ("tit", "°F")
#             # else:
#             #     return

#             # Iterate over metrics and place them in the grid
#             for i, (label, (field, unit)) in enumerate(metrics.items()):
#                 col_index = i % len(all_columns)  # Determine the column index
#                 with all_columns[col_index].container(height=110):  # Create a container for each tile
#                     if field in df.columns and len(df) > 1:
#                         # Get the latest and previous values
#                         latest_value = float(df[field].iloc[0])  # Latest value (last row)
#                         previous_value = float(df[field].iloc[-1])  # Previous value (second last row)
#                         delta = latest_value - previous_value  # Calculate the delta

#                         text_color = "#ffffff"
#                         display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
        
#                         # Display delta value
#                         if delta > 0:
#                             delta_display = f"<span style='color:green; font-size:18px;'>↑ {delta:+.1f}</span>"
#                         elif delta < 0:
#                             delta_display = f"<span style='color:red; font-size:18px;'>↓ {delta:+.1f}</span>"
#                         else:
#                             delta_display = ""

#                         # Display each metric using st.markdown to include HTML
#                         st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
#                         st.markdown(f"<div style='text-align:left;'>{display_text}</div>", unsafe_allow_html=True)
#                         st.markdown(f"<div style='text-align:left;'>{delta_display}</div>", unsafe_allow_html=True)

#                     else:
#                         display_text = "<span style='color:#ffffff; font-size:24px;'>Data not available</span>"
#                         delta_display = "<span style='color:#ffffff;'>N/A</span>"

#                         st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
#                         st.markdown(f"<div style='text-align:left;'>{display_text}</div>", unsafe_allow_html=True)
#                         st.markdown(f"<div style='text-align:left;'>{delta_display}</div>", unsafe_allow_html=True)
#                     # Display each metric using st.metric with delta
#                     # st.metric(label, display_text, delta=delta_display)
                    
#         else:

#             metrics = {
#                 "Front Oil P": ("front_oil_p", "PSI", st.session_state.get('minFrontOilPressure'), st.session_state.get('nominalFrontOilPressure'), st.session_state.get('highFrontOilPressure'), st.session_state.get('maxFrontOilPressure')),
#                 "Rear Oil P": ("rear_oil_p", "PSI", st.session_state.get('minRearOilPressure'), st.session_state.get('nominalRearOilPressure'), st.session_state.get('highRearOilPressure'), st.session_state.get('maxRearOilPressure')),
#                 "OAT": ("oat", "°F", -500.0, 200.0, 200.0, 200.0),
#                 "IAT Carb": ("iat_carb", "°F", st.session_state.get('minIAT'), st.session_state.get('nominalIAT'), st.session_state.get('highIAT'), st.session_state.get('maxIAT')),
#                 "Oil Temp": ("oil_temperature", "°C", st.session_state.get('minOilTemp'), st.session_state.get('nominalOilTemp'), st.session_state.get('highOilTemp'), st.session_state.get('maxOilTemp')),
#                 "Fuel Flow": ("fuel_flow", "GPH", st.session_state.get('minFuelFlow'), st.session_state.get('nominalFuelFlow'), st.session_state.get('highFuelFlow'), st.session_state.get('maxFuelFlow')),
#                 "Metered Fuel P": ("metered_fuel_pressure", "PSI", st.session_state.get('minMeteredFuelPressure'), st.session_state.get('nominalMeteredFuelPressure'), st.session_state.get('highMeteredFuelPressure'), st.session_state.get('maxMeteredFuelPressure')),
#                 "Unmetered Fuel P": ("unmetered_fuel_pressure", "PSI", st.session_state.get('minUnmeteredFuelPressure'), st.session_state.get('nominalUnmeteredFuelPressure'), st.session_state.get('highUnmeteredFuelPressure'), st.session_state.get('maxUnmeteredFuelPressure')),
#                 "Volts": ("volts", "V", st.session_state.get('minFuelFlow'), st.session_state.get('nominalFuelFlow'), st.session_state.get('highFuelFlow'), st.session_state.get('maxFuelFlow')),
#                 # "CDT": ("cdt", "°F", st.session_state.get('minCDT'), st.session_state.get('minCDT'), st.session_state.get('minCDT'), st.session_state.get('minCDT')), 
#                 "Mag Drop": ("mag_drop", "", st.session_state.get('minMagDrop'), st.session_state.get('nominalMagDrop'), st.session_state.get('highMagDrop'), st.session_state.get('maxMagDrop'))
#             }

#             # Add thresholds for TIT
#             # if turboStatus == "True" or turboStatus == "true":
#             if st.session_state.get('turbo') == True:
#                 metrics["TIT"] = ("tit", "°F", st.session_state.get('minTIT'), st.session_state.get('nominalTIT'), st.session_state.get('highTIT'), st.session_state.get('maxTIT'))
#                 metrics["CDT"] = ("cdt", "°F", st.session_state.get('minCDT'), st.session_state.get('minCDT'), st.session_state.get('minCDT'), st.session_state.get('minCDT'))
#             else:
#                 # st.write(f"Turbo Status: {turboStatus}, Datatype: {type(turboStatus)}")
#                 pass

#             # Iterate over metrics and place them in the grid
#             for i, (label, (field, unit, minThreshold, nominalThreshold, highThreshold, maxThreshold)) in enumerate(metrics.items()):
#                 col_index = i % len(all_columns)  # Determine the column index
#                 with all_columns[col_index].container(height=110):  # Create a container for each tile
#                     if field in df.columns and len(df) > 1:
#                         # Get the latest and previous values
#                         latest_value = float(df[field].iloc[0])  # Latest value (last row)
#                         # st.write(latest_value)
#                         previous_value = float(df[field].iloc[-1])  # Previous value (second last row)
#                         delta = latest_value - previous_value  # Calculate the delta
#                         # st.write(f"Field: {field}, Previous: {previous_value}, Latest: {latest_value}, Min: {minThreshold}, Nominal: {nominalThreshold}, High: {highThreshold}, Max: {maxThreshold}")
#                         # st.write(previous_value)
#                         # st.write(latest_value)
                        
                        
#                         # Determine text color based on thresholds
#                         if latest_value > maxThreshold:
#                             text_color = "red"
#                             # st.markdown(f"<div class='flash'><span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span></div>", unsafe_allow_html=True)
#                             # display_text = f"<div class='flash'><span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span></div>"#, unsafe_allow_html=True
#                             display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span></div>"#, unsafe_allow_html=True
#                         elif highThreshold < latest_value <= 0.9 * maxThreshold:
#                             text_color = "yellow"
#                             # st.markdown(f"<span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span>", unsafe_allow_html=True)
#                             display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
#                         elif nominalThreshold < latest_value <= highThreshold:
#                             text_color = "green"
#                             # st.markdown(f"<span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span>", unsafe_allow_html=True)
#                             display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
#                         elif minThreshold < latest_value <= nominalThreshold:
#                             text_color = "green"
#                             # st.markdown(f"<span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span>", unsafe_allow_html=True)
#                             display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
#                         else:
#                             text_color = "white"
#                             # st.markdown(f"<span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span>", unsafe_allow_html=True)
#                             display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
                        
#                         # Display delta value
#                         if delta > 0:
#                             delta_display = f"<span style='color:green; font-size:18px;'>↑ {delta:+.1f}</span>"
#                         elif delta < 0:
#                             delta_display = f"<span style='color:red; font-size:18px;'>↓ {delta:+.1f}</span>"
#                         else:
#                             delta_display = ""

#                         # Display each metric using st.markdown to include HTML
#                         st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
#                         st.markdown(f"<div style='text-align:left;'>{display_text}</div>", unsafe_allow_html=True)
#                         st.markdown(f"<div style='text-align:left;'>{delta_display}</div>", unsafe_allow_html=True)
                    
#                     else:
#                         display_text = "<span style='color:#ffffff; font-size:24px;'>Data not available</span>"
#                         delta_display = "<span style='color:#ffffff;'>N/A</span>"

#                         st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
#                         st.markdown(f"<div style='text-align:left;'>{display_text}</div>", unsafe_allow_html=True)
#                         st.markdown(f"<div style='text-align:left;'>{delta_display}</div>", unsafe_allow_html=True)




def display_metrics(df, engine_type):
    """
    Display metrics as text with labels and units.
    
    Args:
    - df (pd.DataFrame): DataFrame containing the data.
    """
    # Create rows of columns
    row1 = st.columns(5)
    row2 = st.columns(5)
    row3 = st.columns(5)

    # Combine all columns into a single list
    all_columns = row1 + row2 + row3

    # CSS for custom styling
    css = """
    <style>
    .metric-tile {
        border-radius: 8px;  /* Rounded corners */
        padding: 10px;  /* Padding inside the tile */
        margin: 5px;  /* Margin around the tile */
        text-align: center;  /* Center-align text */
    }
    .metric-label {
        font-size: 14px;  /* Adjust label font size */
        text-align: left;  /* Center-align text */
        color: #ffffff;  /* Darker color for label */
    }
    .metric-value {
        font-size: 18px;  /* Adjust value font size */
        color: #ffffff;  /* Black color for value */
    .metric-delta {
        font-size: 14px;  /* Adjust delta font size */
        color: #007bff;  /* Blue color for delta */
    }
    </style>
    """

    st.markdown(css, unsafe_allow_html=True)

    if not df.empty:
        if engine_type == dataOnly:
            metrics = {
            "OAT": ("oat", "°C"),
            "IAT Carb": ("iat_carb", "°C"),
            "Front Oil P": ("front_oil_p", "PSI"),
            "Rear Oil P": ("rear_oil_p", "PSI"),
            # Add Upper Deck Pressure
            "TIT": ("tit", "°F"),
            "Oil Temp": ("oil_temperature", "°C"),
            "Fuel Flow": ("fuel_flow", "GPH"),
            "Metered Fuel P": ("metered_fuel_pressure", "PSI"),
            "Unmetered Fuel P": ("unmetered_fuel_pressure", "PSI"),
            "Volts": ("volts", "V"),
            "CDT": ("cdt", "°F"), 
            "Mag Drop": ("mag_drop", "")
            }

            # if turboStatus == True:
            #     metrics["TIT"] = ("tit", "°F")
            # else:
            #     return

            # Iterate over metrics and place them in the grid
            for i, (label, (field, unit)) in enumerate(metrics.items()):
                col_index = i % len(all_columns)  # Determine the column index
                with all_columns[col_index].container(height=110):  # Create a container for each tile
                    if field in df.columns and len(df) > 1:
                        # Get the latest and previous values
                        latest_value = float(df[field].iloc[0])  # Latest value (last row)
                        previous_value = float(df[field].iloc[-1])  # Previous value (second last row)
                        delta = latest_value - previous_value  # Calculate the delta

                        text_color = "#ffffff"
                        display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
        
                        # Display delta value
                        if delta > 0:
                            delta_display = f"<span style='color:green; font-size:18px;'>↑ {delta:+.1f}</span>"
                        elif delta < 0:
                            delta_display = f"<span style='color:red; font-size:18px;'>↓ {delta:+.1f}</span>"
                        else:
                            delta_display = ""

                        # Display each metric using st.markdown to include HTML
                        st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
                        st.markdown(f"<div style='text-align:left;'>{display_text}</div>", unsafe_allow_html=True)
                        st.markdown(f"<div style='text-align:left;'>{delta_display}</div>", unsafe_allow_html=True)

                    else:
                        display_text = "<span style='color:#ffffff; font-size:24px;'>Data not available</span>"
                        delta_display = "<span style='color:#ffffff;'>N/A</span>"

                        st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
                        st.markdown(f"<div style='text-align:left;'>{display_text}</div>", unsafe_allow_html=True)
                        st.markdown(f"<div style='text-align:left;'>{delta_display}</div>", unsafe_allow_html=True)
                    # Display each metric using st.metric with delta
                    # st.metric(label, display_text, delta=delta_display)
                    
        else:

            metrics = {
                "Oil Temp": ("oil_temperature", "°C", st.session_state.get('minOilTemp'), st.session_state.get('nominalOilTemp'), st.session_state.get('highOilTemp'), st.session_state.get('maxOilTemp')),
                "Front Oil P": ("front_oil_p", "PSI", st.session_state.get('minFrontOilPressure'), st.session_state.get('nominalFrontOilPressure'), st.session_state.get('highFrontOilPressure'), st.session_state.get('maxFrontOilPressure')),
                "Rear Oil P": ("rear_oil_p", "PSI", st.session_state.get('minRearOilPressure'), st.session_state.get('nominalRearOilPressure'), st.session_state.get('highRearOilPressure'), st.session_state.get('maxRearOilPressure')), 
                "Metered Fuel P": ("metered_fuel_pressure", "PSI", st.session_state.get('minMeteredFuelPressure'), st.session_state.get('nominalMeteredFuelPressure'), st.session_state.get('highMeteredFuelPressure'), st.session_state.get('maxMeteredFuelPressure')),
                "Unmetered Fuel P": ("unmetered_fuel_pressure", "PSI", st.session_state.get('minUnmeteredFuelPressure'), st.session_state.get('nominalUnmeteredFuelPressure'), st.session_state.get('highUnmeteredFuelPressure'), st.session_state.get('maxUnmeteredFuelPressure')),
                "Fuel Flow": ("fuel_flow", "GPH", st.session_state.get('minFuelFlow'), st.session_state.get('nominalFuelFlow'), st.session_state.get('highFuelFlow'), st.session_state.get('maxFuelFlow')),
                # "Volts": ("volts", "V", st.session_state.get('minFuelFlow'), st.session_state.get('nominalFuelFlow'), st.session_state.get('highFuelFlow'), st.session_state.get('maxFuelFlow')),
                # "CDT": ("cdt", "°F", st.session_state.get('minCDT'), st.session_state.get('minCDT'), st.session_state.get('minCDT'), st.session_state.get('minCDT')), 
                "OAT": ("oat", "°C", -500.0, 200.0, 200.0, 200.0),
                "IAT Carb": ("iat_carb", "°C", st.session_state.get('minIAT'), st.session_state.get('nominalIAT'), st.session_state.get('highIAT'), st.session_state.get('maxIAT')),
                "Mag Drop": ("mag_drop", "", st.session_state.get('minMagDrop'), st.session_state.get('nominalMagDrop'), st.session_state.get('highMagDrop'), st.session_state.get('maxMagDrop'))
            }

            # Add thresholds for TIT
            # if turboStatus == "True" or turboStatus == "true":
            if st.session_state.get('turbo') == True:
                metrics["TIT"] = ("tit", "°F", st.session_state.get('minTIT'), st.session_state.get('nominalTIT'), st.session_state.get('highTIT'), st.session_state.get('maxTIT'))
                metrics["CDT"] = ("cdt", "°F", st.session_state.get('minCDT'), st.session_state.get('minCDT'), st.session_state.get('minCDT'), st.session_state.get('minCDT'))
            else:
                # st.write(f"Turbo Status: {turboStatus}, Datatype: {type(turboStatus)}")
                pass

            # Iterate over metrics and place them in the grid
            for i, (label, (field, unit, minThreshold, nominalThreshold, highThreshold, maxThreshold)) in enumerate(metrics.items()):
                col_index = i % len(all_columns)  # Determine the column index
                with all_columns[col_index].container(height=110):  # Create a container for each tile
                    if field in df.columns and len(df) > 1:
                        # Get the latest and previous values
                        latest_value = float(df[field].iloc[0])  # Latest value (last row)
                        # st.write(latest_value)
                        previous_value = float(df[field].iloc[-1])  # Previous value (second last row)
                        delta = latest_value - previous_value  # Calculate the delta
                        # st.write(f"Field: {field}, Previous: {previous_value}, Latest: {latest_value}, Min: {minThreshold}, Nominal: {nominalThreshold}, High: {highThreshold}, Max: {maxThreshold}")
                        # st.write(previous_value)
                        # st.write(latest_value)
                        
                        if "oil_p" in field.lower():
                            # Expected semantic meaning:
                            # minThreshold → minimum acceptable pressure
                            # nominalThreshold → typical/idle pressure
                            # highThreshold → upper normal operating limit
                            # maxThreshold → redline limit

                            # Add derived regions based on those threshold names
                            warning_low = (nominalThreshold + highThreshold) / 2   # midpoint between nominal & high (~45 PSI region)
                            caution_high = (highThreshold + maxThreshold) / 2      # midpoint between high & max (~90–105 PSI region)
                            red_high = maxThreshold * 1.1667                       # defines upper red beyond max (~105 PSI for 90 max)

                            if latest_value < nominalThreshold:                    # Below nominal → RED
                                text_color = "red"
                                display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
                            elif latest_value < warning_low:                       # Between nominal & midpoint → YELLOW
                                text_color = "yellow"
                                display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
                            elif latest_value < highThreshold:                     # Between ~45 and 60 → WHITE
                                text_color = "green"
                                display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
                            elif latest_value <= maxThreshold:                     # Normal range (60–90) → GREEN
                                text_color = "green"
                                display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
                            elif latest_value <= caution_high:                     # Slightly above normal → YELLOW
                                text_color = "yellow"
                                display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
                            elif latest_value > red_high:                          # Far above max → RED
                                text_color = "red"
                                display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
                            else:
                                text_color = "yellow"
                                display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
                            # Display delta value
                            if delta > 0:
                                delta_display = f"<span style='color:green; font-size:18px;'>↑ {delta:+.1f}</span>"
                            elif delta < 0:
                                delta_display = f"<span style='color:red; font-size:18px;'>↓ {delta:+.1f}</span>"
                            else:
                                delta_display = ""

                            # Display each metric using st.markdown to include HTML
                            st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
                            st.markdown(f"<div style='text-align:left;'>{display_text}</div>", unsafe_allow_html=True)
                            st.markdown(f"<div style='text-align:left;'>{delta_display}</div>", unsafe_allow_html=True)
                        # === Normal color logic for all other metrics ===
                        # Determine text color based on thresholds
                        else: 
                            if latest_value > maxThreshold:
                                text_color = "red"
                                # st.markdown(f"<div class='flash'><span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span></div>", unsafe_allow_html=True)
                                # display_text = f"<div class='flash'><span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span></div>"#, unsafe_allow_html=True
                                display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span></div>"#, unsafe_allow_html=True
                            elif highThreshold < latest_value <= 0.9 * maxThreshold:
                                text_color = "yellow"
                                # st.markdown(f"<span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span>", unsafe_allow_html=True)
                                display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
                            elif nominalThreshold < latest_value <= highThreshold:
                                text_color = "green"
                                # st.markdown(f"<span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span>", unsafe_allow_html=True)
                                display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
                            elif minThreshold < latest_value <= nominalThreshold:
                                text_color = "green"
                                # st.markdown(f"<span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span>", unsafe_allow_html=True)
                                display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
                            else:
                                text_color = "white"
                                # st.markdown(f"<span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span>", unsafe_allow_html=True)
                                display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"#, unsafe_allow_html=True
                            
                            # Display delta value
                            if delta > 0:
                                delta_display = f"<span style='color:green; font-size:18px;'>↑ {delta:+.1f}</span>"
                            elif delta < 0:
                                delta_display = f"<span style='color:red; font-size:18px;'>↓ {delta:+.1f}</span>"
                            else:
                                delta_display = ""

                            # Display each metric using st.markdown to include HTML
                            st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
                            st.markdown(f"<div style='text-align:left;'>{display_text}</div>", unsafe_allow_html=True)
                            st.markdown(f"<div style='text-align:left;'>{delta_display}</div>", unsafe_allow_html=True)
                    
                    else:
                        display_text = "<span style='color:#ffffff; font-size:24px;'>Data not available</span>"
                        delta_display = "<span style='color:#ffffff;'>N/A</span>"

                        st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
                        st.markdown(f"<div style='text-align:left;'>{display_text}</div>", unsafe_allow_html=True)
                        st.markdown(f"<div style='text-align:left;'>{delta_display}</div>", unsafe_allow_html=True)



# def plot_egt_cht(df,engine_type):
#     """
#     Plot EGT and CHT values in a bar graph.
    
#     Args:
#     - df (pd.DataFrame): DataFrame containing the data.
#     """
#     temps_start_time = time.time()  # Start time for performance measurement
    
#     fig, ax = plt.subplots(figsize=(12, 6))
    
#     # Identify EGT and CHT columns based on engine type
#     if st.session_state.get('cylinders') == 4:
#         egt_columns = sorted([col for col in df.columns if col.startswith('egt_')])[:4]
#         cht_columns = sorted([col for col in df.columns if col.startswith('cht_')])[:4]
#     elif st.session_state.get('cylinders') == 6:
#         egt_columns = sorted([col for col in df.columns if col.startswith('egt_')])[:6]
#         cht_columns = sorted([col for col in df.columns if col.startswith('cht_')])[:6]
#     else:
#         # raise ValueError("Invalid engine type. Must be '4-cylinder' or '6-cylinder'.")
#         egt_columns = sorted([col for col in df.columns if col.startswith('egt_')])[:6]
#         cht_columns = sorted([col for col in df.columns if col.startswith('cht_')])[:6]

    

#     # Calculate the highest EGT and CHT values
#     egt_values = [df[col].iloc[0] for col in egt_columns]
#     cht_values = [df[col].iloc[0] for col in cht_columns]

#     max_egt_value = max(egt_values)
#     max_cht_value = max(cht_values)

#     # st.write(f"EGT Columns: {egt_values}")
#     # st.write(f"CHT Columns: {cht_values}")

#     # Interleave EGT and CHT labels and values
#     labels = []
#     values = []
#     text = []
#     colors = []

#     # Create traces
#     if engine_type == dataOnly:
#         for egt_label, cht_label in zip(egt_columns, cht_columns):
#             labels.append(egt_label)
#             values.append(df[egt_label].iloc[0])
#             egt_value = float(df[egt_label].iloc[0])
#             # text.append(f'{egt_value} °F')
#             text.append(f'{egt_value} °F' + (' (H)' if egt_value == max_egt_value else ''))
#             labels.append(cht_label)
#             values.append(df[cht_label].iloc[0]) 
#             cht_value = float(df[cht_label].iloc[0])
#             # text.append(f'{cht_value} °F')
#             text.append(f'{cht_value} °F'+ (' (H)' if cht_value == max_cht_value else ''))

#         traces = [
#         go.Bar(
#             x=labels,
#             y=values,
#             name='Temperature',
#             text=text,
#             textposition='auto',
#             textfont=dict(size=16),  # ← Increase text size here
#             marker_color=['green' if 'egt_' in label else 'blue' for label in labels]  # Color based on label
#         )
#     ]

#     else:
#         for egt_label, cht_label in zip(egt_columns, cht_columns):
#             labels.append(egt_label)
#             values.append(df[egt_label].iloc[0])
#             egt_value = float(df[egt_label].iloc[0])
#             # text.append(f'{egt_value} °F')
#             text.append(f'{egt_value} °F' + (' (H)' if egt_value == max_egt_value else ''))
#             colors.append(get_color('EGT', egt_value))  # Default color for EGT

#             labels.append(cht_label)
#             values.append(df[cht_label].iloc[0])
#             cht_value = float(df[cht_label].iloc[0])
#             # text.append(f'{cht_value} °F')
#             text.append(f'{cht_value} °F'+ (' (H)' if cht_value == max_cht_value else ''))
#             colors.append(get_color('CHT', cht_value))  # Default color for EGT

#         traces = [
#             go.Bar(
#                 x=labels,
#                 y=values,
#                 name='Temperature',
#                 marker_color=colors,  # Color based on label
#                 text=text,  # Show values on hover
#                 textfont=dict(size=16),  # ← Increase text size here
#                 textposition='auto'  # Automatically position text on the bars
#             )
#         ]
    
#     # Create figure
#     fig = go.Figure(data=traces)
    
#     # Update layout
#     fig.update_layout(
#         width=800,  # Set the desired width
#         height=350,  # Set the desired height
#         margin=dict(l=20, r=20, t=40, b=20),  # Adjust margins if necessary
#         barmode='group',  # Group bars together
#         xaxis_tickangle=-45  # Rotate x-axis labels for better readability
#     )
    
#     temps_plot_time = time.time() - temps_start_time  # Measure time taken to render plot
#     st.session_state['temps_plot_render_time'] = temps_plot_time  # Store plot render time in session state
    
#     return fig
    
def plot_egt_cht(df, engine_type):
    """
    Plot EGT and CHT values in a bar graph with Diff metrics.
    """
    temps_start_time = time.time()
    
    # --- 1. Identify EGT and CHT columns based on engine type ---
    if st.session_state.get('cylinders') == 4:
        egt_columns = sorted([col for col in df.columns if col.startswith('egt_')])[:4]
        cht_columns = sorted([col for col in df.columns if col.startswith('cht_')])[:4]
    else:
        # Default to 6 cylinders if 6 or unknown
        egt_columns = sorted([col for col in df.columns if col.startswith('egt_')])[:6]
        cht_columns = sorted([col for col in df.columns if col.startswith('cht_')])[:6]

    # --- 2. Calculate Max, Min, and Diff (The new functionality) ---
    egt_values = [df[col].iloc[0] for col in egt_columns]
    cht_values = [df[col].iloc[0] for col in cht_columns]

    max_egt = max(egt_values)
    min_egt = min(egt_values)
    diff_egt = max_egt - min_egt

    max_cht = max(cht_values)
    min_cht = min(cht_values)
    diff_cht = max_cht - min_cht

    # --- 3. Prepare Data for Plotting ---
    labels = []
    values = []
    text = []
    colors = []

    # Helper function to generate labels/colors
    for egt_col, cht_col in zip(egt_columns, cht_columns):
        # Process EGT
        egt_val = int(df[egt_col].iloc[0])
        labels.append(egt_col)
        values.append(egt_val)
        
        # Add (H) for High and (L) for Low to the bar text
        note = ""
        if egt_val == max_egt: note = " (H)"
        # elif egt_val == min_egt: note = " (L)"
        text.append(f'{egt_val} °F{note}')
        
        # Color logic
        if engine_type != 'dataOnly':
             colors.append(get_color('EGT', egt_val)) # Assuming get_color exists in your scope
        
        # Process CHT
        cht_val = int(df[cht_col].iloc[0])
        labels.append(cht_col)
        values.append(cht_val)

        # Add (H) for High and (L) for Low
        note = ""
        if cht_val == max_cht: note = " (H)"
        # elif cht_val == min_cht: note = " (L)"
        text.append(f'{cht_val} °F{note}')

        # Color logic
        if engine_type != 'dataOnly':
             colors.append(get_color('CHT', cht_val))

    # --- 4. Create Traces ---
    if engine_type == 'dataOnly':
        # Simple Green/Blue coloring
        trace = go.Bar(
            x=labels,
            y=values,
            text=text,
            textposition='auto',
            textfont=dict(size=16),
            marker_color=['green' if 'egt_' in label else 'blue' for label in labels]
        )
    else:
        # Custom coloring based on get_color function
        trace = go.Bar(
            x=labels,
            y=values,
            marker_color=colors,
            text=text,
            textfont=dict(size=16),
            textposition='auto'
        )

    fig = go.Figure(data=[trace])

    # --- 5. Update Layout with Diff Metrics in Title ---
    # This formats a title like: "EGT Diff: 100°F  |  CHT Diff: 25°F"
    title_text = (
        f"<b>Temps Diffs</b><br>"
        f"<span style='font-size: 16px; color: gray;'>"
        f"EGT Diff: {diff_egt}°F  |  CHT Diff: {diff_cht}°F"
        f"</span>"
    )

    fig.update_layout(
        title=dict(text=title_text, x=0.5, xanchor='center'), # Center the title
        width=800,
        height=400,
        margin=dict(l=20, r=20, t=60, b=20), # Increased top margin (t) to fit the title
        barmode='group',
        xaxis_tickangle=-45
    )

    # Performance tracking
    temps_plot_time = time.time() - temps_start_time
    st.session_state['temps_plot_render_time'] = temps_plot_time
    
    return fig



def animate_rpm(df, engine_type):
    """
    Create a gauge chart for RPM values.
    
    Args:
    - df (pd.DataFrame): DataFrame containing the data.
    """
    rpm_start_time = time.time()  # Start time for performance measurement
    
    rpm = df['rpm'].iloc[0] if not df.empty else 0

    if engine_type == dataOnly:
        gaugeSetup = {
            'axis': {'range': [None, 3000]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, 750], 'color': "lightblue"},
                {'range': [750, 1500], 'color': "skyblue"},
                {'range': [1500, 2250], 'color': "deepskyblue"},
                {'range': [2250, 3000], 'color': "dodgerblue"}
            ]
        }

    else:
        gaugeSetup = {
            'axis': {'range': [st.session_state.get('minRPM'), int(st.session_state.get('maxRPM'))+500]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, st.session_state.get('maxRPM')/3], 'color': "lightblue"},  
                {'range': [st.session_state.get('maxRPM')/3, 2*(st.session_state.get('maxRPM')/3)], 'color': "skyblue"}, 
                {'range': [2*(st.session_state.get('maxRPM')/3), st.session_state.get('maxRPM')], 'color': "deepskyblue"},
                {'range': [st.session_state.get('maxRPM'), int(st.session_state.get('maxRPM'))+500], 'color': "dodgerblue"}
                ],
            'threshold' : {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': st.session_state.get('maxRPM')}
        }


    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=rpm,
        domain={'x': [0,0.55], 'y': [0,1]},
        title={'text': "RPM"},
        gauge=gaugeSetup
    ))

    # Update layout to adjust size
    fig.update_layout(
        width=300,  # Set the desired width
        height=120,  # Set the desired height
        margin=dict(l=20, r=20, t=45, b=20)  # Adjust margins if necessary
    )

    rpm_plot_time = time.time() - rpm_start_time  # Measure time taken to render plot
    st.session_state['rpm_plot_render_time'] = rpm_plot_time  # Store plot render time in session state

    return fig


def animate_manifold_pressure(df, engine_type):
    """
    Create a gauge chart for Manifold Pressure values.
    
    Args:
    - df (pd.DataFrame): DataFrame containing the data.
    """
    mp_start_time = time.time()  # Start time for performance measurement
    manifold_pressure = df['manifold_pressure'].iloc[0] if not df.empty else 0

    if engine_type == dataOnly:
        gaugeSetup ={
            'axis': {'range': [None, 40]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, 10], 'color': "lightblue"},
                {'range': [10, 20], 'color': "skyblue"},
                {'range': [20, 30], 'color': "deepskyblue"}, 
                {'range': [30, 40], 'color': "dodgerblue"}
            ]
        }
    else:
        gaugeSetup = {
            'axis': {'range': [st.session_state.get('minMP'), int(st.session_state.get('maxMP'))+5]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, st.session_state.get('maxMP')/3], 'color': "lightblue"},
                {'range': [st.session_state.get('maxMP')/3, 2*(st.session_state.get('maxMP')/3)], 'color': "skyblue"},
                {'range': [2*(st.session_state.get('maxMP')/3), st.session_state.get('maxMP')], 'color': "deepskyblue"},
                {'range': [st.session_state.get('maxMP'), int(st.session_state.get('maxMP'))+5], 'color': "dodgerblue"}
            ],
            'threshold' : {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': st.session_state.get('maxMP')}
        }

    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=manifold_pressure,
        number={'suffix': " hg"},  # Add the units here
        domain={'x': [0,0.55], 'y': [0,1]},
        title={'text': "MP"},
        gauge=gaugeSetup
    ))

    # Update layout to adjust size
    fig.update_layout(
        width=300,  # Set the desired width
        height=120,  # Set the desired height
        margin=dict(l=20, r=20, t=45, b=20)  # Adjust margins if necessary
    )

    mp_plot_time = time.time() - mp_start_time  # Measure time taken to render plot
    st.session_state['mp_plot_render_time'] = mp_plot_time  # Store plot render time in session state

    return fig

def animate_hp(df, engine_type):
    """
    Create a gauge chart for horsepower values.
    
    Args:
    - df (pd.DataFrame): DataFrame containing the data.
    """
    hp_start_time = time.time()  # Start time for performance measurement
    horsepower = df['hp'].iloc[0] if not df.empty else 0

    if engine_type == dataOnly:
        gaugeSetup = {
            'axis': {'range': [None, 200]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, 50], 'color': "lightblue"},
                {'range': [50, 100], 'color': "skyblue"},
                {'range': [100, 150], 'color': "deepskyblue"}, 
                {'range': [150, 200], 'color': "dodgerblue"}
            ]
        }

    else:
        gaugeSetup = {
            'axis': {'range': [st.session_state.get('minHP'), int(st.session_state.get('maxHP'))+30]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, st.session_state.get('maxHP')/3], 'color': "lightblue"},
                {'range': [st.session_state.get('maxHP')/3, 2*(st.session_state.get('maxHP')/3)], 'color': "skyblue"},
                {'range': [2*(st.session_state.get('maxHP')/3), st.session_state.get('maxHP')], 'color': "deepskyblue"},
                {'range': [st.session_state.get('maxHP'), int(st.session_state.get('maxHP'))+30], 'color': "dodgerblue"}
            ],
            'threshold' : {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': st.session_state.get('maxHP')}
        }


    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=horsepower,
        domain={'x': [0,0.55], 'y': [0,1]},
        title={'text': "HP"},
        gauge=gaugeSetup
    ))

    # Update layout to adjust size
    fig.update_layout(
        width=300,  # Set the desired width
        height=120,  # Set the desired height
        margin=dict(l=20, r=20, t=45, b=20)  # Adjust margins if necessary
    )

    mp_plot_time = time.time() - hp_start_time  # Measure time taken to render plot
    st.session_state['hp_plot_render_time'] = mp_plot_time  # Store plot render time in session state

    return fig


def plot_performance_metrics():
    """
    Create gauge charts for Data Fetch Time and Plot Render Time.
    """

    global refresh_time
    data_fetch_time = st.session_state.get('data_fetch_time', 0)
    temps_plot_render_time = st.session_state.get('temps_plot_render_time', 0)
    rpm_plot_render_time = st.session_state.get('rpm_plot_render_time', 0)
    mp_plot_render_time = st.session_state.get('mp_plot_render_time', 0)
    hp_plot_render_time = st.session_state.get('hp_plot_render_time', 0)
    total_plot_time = temps_plot_render_time + rpm_plot_render_time + mp_plot_render_time + refresh_time + hp_plot_render_time

    # Create figure with two subplots
    #fig = go.Figure()

    # Data Fetch & Plot Render Time Gauge
    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        domain={'x': [0,0.75], 'y': [0,1]},
        value=data_fetch_time + total_plot_time,
        number={'suffix': " s"},  # Add the units here
        title={'text': "Lag"},
        gauge={
            'axis': {'range': [None, 1.5]},  # Adjust the range according to your needs
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, 0.4], 'color': "lightblue"},
                {'range': [0.4, 0.8], 'color': "skyblue"},
                {'range': [0.8, 1.5], 'color': "deepskyblue"},
                # {'range': [1.4, 2.5], 'color': "dodgerblue"}
            ]
        }
    ))

    # Update layout to adjust size
    fig.update_layout(
        width=300,  # Set the desired width
        height=120,  # Set the desired height
        margin=dict(l=20, r=20, t=45, b=20)  # Adjust margins if necessary
    )
    
    return fig


# Function to convert DataFrame to CSV
def convert_df_to_csv(test_sheet_df):
    return test_sheet_df.to_csv(index=False).encode('utf-8')


# Button handlers
# @st.fragment
# def start_button_handler(df):
#     # st.write("HELLLOOOOOOOO")
#     now = datetime.datetime.now()
#     test_date = now.strftime("%Y-%m-%d")
#     # start_time = now.strftime("%H:%M:%S")
#     utc_now = datetime.datetime.now(datetime.UTC)
#     ast_tz = pytz.timezone('America/Halifax')  # Atlantic Standard Time (AST) timezone
#     ast_now = utc_now.replace(tzinfo=pytz.utc).astimezone(ast_tz)
#     start_time = ast_now.isoformat()  # Use AST timestamp
#     start_time_str = ast_now.strftime("%H:%M:%S")
#     specOilPress = st.session_state.get('SpecOilPressure')
#     specIdlePress = st.session_state.get('SpecIdlePressure')
#     specFTFuelPressureMetered = st.session_state.get('SpecF/TFuelPressureMetered')
#     specFTFuelPressureUnMetered = st.session_state.get('SpecF/TFuelPressureUnMetered')
#     specIdleFuelPressureMetered = st.session_state.get('SpecIdleFuelPressureMetered')
#     specIdleFuelPressureUnMetered = st.session_state.get('SpecIdleFuelPressureUnMetered')
#     maxCHT = st.session_state.get("MaxCHT")
#     maxEGTTIT = st.session_state.get("MAXEGT/TIT")
#     maxManifold = st.session_state.get("MaxManifold")
#     maxTurboBoost = st.session_state.get("MaxTurboBoost")

#     split_engine_str = engine_type.split() 

#     engineMake = split_engine_str[0]
#     engineModel = split_engine_str[1]
    
    

#     # if 'oat' in df.columns:
#     #     outside_temp = int(df['oat'].iloc[0])
#     # else:
#     #     st.error("Column 'oat' not found in DataFrame.")
#     #     outside_temp = 0
#         # return  # or handle this case appropriately

#     # outside_temp = int(df['oat'].iloc[0])

#     # Ensure the columns 'Unnamed: 8' and 'Unnamed: 15' exist
#     for col in ['Unnamed: 1','Unnamed: 4','Unnamed: 7','Unnamed: 8','Unnamed: 13','Unnamed: 15','Unnamed: 17',]:
#         if col not in st.session_state.test_sheet.columns:
#             st.session_state.test_sheet[col] = None

#     # Modify specific cells with current date and time
#     st.session_state.test_sheet.loc[2, 'Unnamed: 8'] = test_date
#     # st.session_state.test_sheet.loc[3, 'Unnamed: 8'] = outside_temp
#     st.session_state.test_sheet.loc[1, 'Unnamed: 15'] = prop_type
#     st.session_state.test_sheet.loc[2, 'Unnamed: 15'] = start_time_str
#     st.session_state.test_sheet.loc[19, 'Unnamed: 1'] = specOilPress
#     st.session_state.test_sheet.loc[20, 'Unnamed: 1'] = specIdlePress
#     st.session_state.test_sheet.loc[19, 'Unnamed: 4'] = specFTFuelPressureMetered
#     st.session_state.test_sheet.loc[20, 'Unnamed: 4'] = specIdleFuelPressureMetered
#     st.session_state.test_sheet.loc[19, 'Unnamed: 7'] = specFTFuelPressureUnMetered
#     st.session_state.test_sheet.loc[20, 'Unnamed: 7'] = specIdleFuelPressureUnMetered
#     st.session_state.test_sheet.loc[19, 'Unnamed: 13'] = maxCHT
#     st.session_state.test_sheet.loc[20, 'Unnamed: 13'] = maxEGTTIT
#     st.session_state.test_sheet.loc[19, 'Unnamed: 17'] = maxManifold
#     st.session_state.test_sheet.loc[20, 'Unnamed: 17'] = maxTurboBoost
#     st.session_state.test_sheet.loc[2, 'Unnamed: 1'] = engineMake
#     st.session_state.test_sheet.loc[3, 'Unnamed: 1'] = engineModel

#     # st.write("THEREEEE")

#     # # Check if cells are populated
#     # start_cells = [(2, 'Unnamed: 8'), (3, 'Unnamed: 8'), (2, 'Unnamed: 15'), (19, 'Unnamed: 1')
#     #                , (20, 'Unnamed: 1'), (19, 'Unnamed: 4'), (20, 'Unnamed: 4'), (19, 'Unnamed: 7')
#     #                , (20, 'Unnamed: 7'), (19, 'Unnamed: 13'), (20, 'Unnamed: 13'), (19, 'Unnamed: 17')
#     #                , (20, 'Unnamed: 17')]
    
#     # Check if cells are populated (OAT column removed)
#     start_cells = [(2, 'Unnamed: 8'), (3, 'Unnamed: 8'), (2, 'Unnamed: 15'), (19, 'Unnamed: 1')
#                    , (20, 'Unnamed: 1'), (19, 'Unnamed: 4'), (20, 'Unnamed: 4'), (19, 'Unnamed: 7')
#                    , (20, 'Unnamed: 7'), (19, 'Unnamed: 13'), (20, 'Unnamed: 13'), (19, 'Unnamed: 17')
#                    , (20, 'Unnamed: 17')]

#     for row, col in start_cells:
#         cell_value = st.session_state.test_sheet.loc[row, col]
#         is_empty = pd.isna(cell_value) or cell_value == ''
#         if is_empty:
#             st.session_state.start_btn_toast = "Oops!! Sorry something went wrong, try again"
#             # time.sleep(1)
#         else:
#             st.session_state.start_btn_toast = "Date & start time populated"
#             # time.sleep(1)

#     # st.write("!!!!!!!!!!!!!!!!")

#     st.toast(st.session_state.start_btn_toast)
#     st.session_state.test_date = test_date
#     st.session_state.start_time = start_time
#     st.session_state.start_time_str = start_time_str


def start_button_handler(df):
    row_map = get_row_map(st.session_state.test_sheet) # Generate the map
    
    now = datetime.datetime.now()
    test_date = now.strftime("%Y-%m-%d")
    utc_now = datetime.datetime.now(datetime.UTC)
    ast_tz = pytz.timezone('America/Halifax')
    ast_now = utc_now.replace(tzinfo=pytz.utc).astimezone(ast_tz)
    start_time_str = ast_now.strftime("%H:%M:%S")

    # Get values from state
    # ... (Your existing variable retrieval code here) ...
    specOilPress = st.session_state.get('SpecOilPressure')
    specIdlePress = st.session_state.get('SpecIdlePressure')
    specFTFuelPressureMetered = st.session_state.get('SpecF/TFuelPressureMetered')
    specFTFuelPressureUnMetered = st.session_state.get('SpecF/TFuelPressureUnMetered')
    specIdleFuelPressureMetered = st.session_state.get('SpecIdleFuelPressureMetered')
    specIdleFuelPressureUnMetered = st.session_state.get('SpecIdleFuelPressureUnMetered')
    maxCHT = st.session_state.get("MAXCHT")
    maxEGTTIT = st.session_state.get("MAXEGT/TIT")
    maxManifold = st.session_state.get("MaxManifold")
    maxTurboBoost = st.session_state.get("MaxTurboBoost")
    
    prop_type = st.session_state.get("selected_prop")
    engine_type = st.session_state.get("selected_engine") # Safety default

    split_engine_str = engine_type.split() 
    engineMake = split_engine_str[0] if len(split_engine_str) > 0 else ""
    engineModel = split_engine_str[1] if len(split_engine_str) > 1 else ""

    # Ensure columns exist (Keep your existing loop)
    # for col in ['Unnamed: 1','Unnamed: 4','Unnamed: 7','Unnamed: 8','Unnamed: 13','Unnamed: 15','Unnamed: 17',]:
    for col in ['Unnamed: 2','Unnamed: 3','Unnamed: 8','Unnamed: 9','Unnamed: 14','Unnamed: 16','Unnamed: 18',]:
        if col not in st.session_state.test_sheet.columns:
            st.session_state.test_sheet[col] = None

    # --- ANCHOR BASED UPDATES ---
    # We use row_map.get('Label') to find the row index dynamically
    
    # Header Info (Usually fixed at top, but safer to anchor if possible, or keep hardcoded if header never changes)
    # st.session_state.test_sheet.loc[2, 'Unnamed: 8'] = test_date
    # st.session_state.test_sheet.loc[1, 'Unnamed: 15'] = prop_type
    # st.session_state.test_sheet.loc[2, 'Unnamed: 15'] = start_time_str
    # st.session_state.test_sheet.loc[2, 'Unnamed: 1'] = engineMake
    # st.session_state.test_sheet.loc[3, 'Unnamed: 1'] = engineModel
    st.session_state.test_sheet.loc[2, 'Unnamed: 9'] = test_date
    st.session_state.test_sheet.loc[1, 'Unnamed: 16'] = prop_type
    st.session_state.test_sheet.loc[2, 'Unnamed: 16'] = start_time_str
    st.session_state.test_sheet.loc[2, 'Unnamed: 2'] = engineMake
    st.session_state.test_sheet.loc[3, 'Unnamed: 2'] = engineModel
    
    # Specs - Dynamic Rows
    try:
        # Find rows based on labels
        r_spec_oil = row_map['Spec Oil Pressure']
        r_spec_idle = row_map['Spec Idle Pressure'] # Ensure this matches your excel label exactly!
        
        # Populate based on found rows
        # st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 1'] = specOilPress
        # st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 1'] = specIdlePress
        st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 2'] = specOilPress
        st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 2'] = specIdlePress
        
        # st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 4'] = specFTFuelPressureMetered
        # st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 4'] = specIdleFuelPressureMetered
        st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 5'] = specFTFuelPressureMetered
        st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 5'] = specIdleFuelPressureMetered
        
        # st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 7'] = specFTFuelPressureUnMetered
        # st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 7'] = specIdleFuelPressureUnMetered
        st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 8'] = specFTFuelPressureUnMetered
        st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 8'] = specIdleFuelPressureUnMetered
        
        # st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 13'] = maxCHT
        # st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 13'] = maxEGTTIT
        st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 14'] = maxCHT
        st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 14'] = maxEGTTIT
        
        # st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 17'] = maxManifold
        # st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 17'] = maxTurboBoost
        st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 18'] = maxManifold
        st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 18'] = maxTurboBoost
        st.session_state.start_btn_toast = "Date, Time & Specs populated"
        st.toast(st.session_state.start_btn_toast)
        
    except KeyError as e:
        st.error(f"Error: Could not find row label {e} in the Excel sheet.")

    st.session_state.test_date = test_date
    st.session_state.start_time_str = start_time_str

# @st.fragment
# def snap_button_handler(df):
#     current_index = get_snapshot_index()
    
#     # This if statement will skip row 13 when populating fields in descending RPM
    
#     if current_index == 14:
#         st.session_state.snapshot_index -= 1
#     # This if statement will change the row being populated
#     if current_index != 8:
#         st.session_state.snapshot_index -= 1
#     # When row is reached, it will be populated and the index moved to row 7 to avoid rewriting row 8    
#     else: 
#         st.session_state.snapshot_index -= 1
#         # remind user to input final RPM
#         st.toast("Confirm Idle RPM")
#     # This if statement set's the boundary for the snap button, doing nothing once row 8 is populated and remaining at row 7
#     if current_index < 8:
#         st.session_state.snapshot_index = 7
#         st.toast(f"All RPM fields are populated. Current index: {current_index}")
#         return

#     manifold_pressure = int(df['manifold_pressure'].iloc[0])
#     front_oil_p = int(df['front_oil_p'].iloc[0])
#     rear_oil_p = int(df['rear_oil_p'].iloc[0])
#     oil_temp = int(df['oil_temperature'].iloc[0])
#     metered_fuel_pressure = int(df['metered_fuel_pressure'].iloc[0])
#     unmetered_fuel_pressure = int(df['unmetered_fuel_pressure'].iloc[0])
#     egt_1 = int(df['egt_1'].iloc[0])
#     egt_2 = int(df['egt_2'].iloc[0])
#     egt_3 = int(df['egt_3'].iloc[0])
#     egt_4 = int(df['egt_4'].iloc[0])
#     egt_5 = int(df['egt_5'].iloc[0])
#     egt_6 = int(df['egt_5'].iloc[0])
#     cht_1 = int(df['cht_1'].iloc[0])
#     cht_2 = int(df['cht_2'].iloc[0])
#     cht_3 = int(df['cht_3'].iloc[0])
#     cht_4 = int(df['cht_4'].iloc[0])
#     cht_5 = int(df['cht_5'].iloc[0])
#     cht_6 = int(df['cht_6'].iloc[0])
#     row_rpm = st.session_state.test_sheet.loc[current_index, 'Unnamed: 0']

#     # Ensure the columns exist
#     for col in ['Unnamed: 1', 'Unnamed: 2','Unnamed: 4','Unnamed: 6','Unnamed: 8'
#                ,'Unnamed: 9','Unnamed: 10','Unnamed: 11', 'Unnamed: 12', 'Unnamed: 13'
#                ,'Unnamed: 14','Unnamed: 15','Unnamed: 16','Unnamed: 17', 'Unnamed: 18', 'Unnamed: 19']:
#         if col not in st.session_state.test_sheet.columns:
#             st.session_state.test_sheet[col] = None

#     # Modify specific cells with current date and time
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 1'] = manifold_pressure
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 3'] = oil_temp
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 4'] = front_oil_p
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 5'] = rear_oil_p
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 6'] = metered_fuel_pressure
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 7'] = unmetered_fuel_pressure
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 8'] = egt_1
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 9'] = egt_2
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 10'] = egt_3
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 11'] = egt_4
    
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 14'] = cht_1
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 15'] = cht_2
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 16'] = cht_3
#     st.session_state.test_sheet.loc[current_index, 'Unnamed: 17'] = cht_4
    

#     if cylinderCount == 6:
#         st.session_state.test_sheet.loc[current_index, 'Unnamed: 12'] = egt_5
#         st.session_state.test_sheet.loc[current_index, 'Unnamed: 13'] = egt_6
#         st.session_state.test_sheet.loc[current_index, 'Unnamed: 18'] = cht_5
#         st.session_state.test_sheet.loc[current_index, 'Unnamed: 19'] = cht_6

#     else:
#         st.session_state.test_sheet.loc[current_index, 'Unnamed: 12'] = 0
#         st.session_state.test_sheet.loc[current_index, 'Unnamed: 13'] = 0
#         st.session_state.test_sheet.loc[current_index, 'Unnamed: 18'] = 0
#         st.session_state.test_sheet.loc[current_index, 'Unnamed: 19'] = 0

#     # Check if cells are populated
#     snap_cells = [(current_index, 'Unnamed: 1'), (current_index, 'Unnamed: 3'), (current_index, 'Unnamed: 4')
#                   ,(current_index, 'Unnamed: 6'), (current_index, 'Unnamed: 6'), (current_index, 'Unnamed: 8')
#                   ,(current_index, 'Unnamed: 9'), (current_index, 'Unnamed: 10'), (current_index, 'Unnamed: 11')
#                   , (current_index, 'Unnamed: 12'), (current_index, 'Unnamed: 13'), (current_index, 'Unnamed: 14')
#                   , (current_index, 'Unnamed: 15'), (current_index, 'Unnamed: 16'), (current_index, 'Unnamed: 17')
#                   , (current_index, 'Unnamed: 18'), (current_index, 'Unnamed: 19')]

#     for row, col in snap_cells:
#         cell_value = st.session_state.test_sheet.loc[row, col]
#         is_empty = pd.isna(cell_value) or cell_value == ''
#         if is_empty:
#             st.session_state.snap_btn_toast = "Oops!! Sorry something went wrong, try again"
            
#         else:
#             st.session_state.snap_btn_toast = f"Snapshot taken. RPM: {row_rpm}."
            
#     st.toast(st.session_state.snap_btn_toast)

def snap_button_handler(df):
    current_index = get_snapshot_index()
    row_map = get_row_map(st.session_state.test_sheet)
    rpm_sequence = get_rpm_sequence() # e.g. ['2700', '2400', ... '750']
    
    # 1. Identify where we are in the sequence
    # Get the RPM string of the current row
    current_row_label = str(st.session_state.test_sheet.loc[current_index, 'Unnamed: 0']).strip()
    
    # Verify we are on a valid RPM row
    if current_row_label not in rpm_sequence:
        st.warning(f"Current row '{current_row_label}' is not a recognized RPM step.")
        # Try to recover? For now, just proceed with data capture
    
    # ... (Your existing Data Capture logic variables here) ...
    # manifold_pressure = ...
    # front_oil_p = ...
    # ...
    manifold_pressure = int(df['manifold_pressure'].iloc[0])
    front_oil_p = int(df['front_oil_p'].iloc[0])
    rear_oil_p = int(df['rear_oil_p'].iloc[0])
    oil_temp = int(df['oil_temperature'].iloc[0])
    metered_fuel_pressure = int(df['metered_fuel_pressure'].iloc[0])
    unmetered_fuel_pressure = int(df['unmetered_fuel_pressure'].iloc[0])
    egt_1 = int(df['egt_1'].iloc[0])
    egt_2 = int(df['egt_2'].iloc[0])
    egt_3 = int(df['egt_3'].iloc[0])
    egt_4 = int(df['egt_4'].iloc[0])
    egt_5 = int(df['egt_5'].iloc[0])
    egt_6 = int(df['egt_5'].iloc[0])
    cht_1 = int(df['cht_1'].iloc[0])
    cht_2 = int(df['cht_2'].iloc[0])
    cht_3 = int(df['cht_3'].iloc[0])
    cht_4 = int(df['cht_4'].iloc[0])
    cht_5 = int(df['cht_5'].iloc[0])
    cht_6 = int(df['cht_6'].iloc[0])
    row_rpm = st.session_state.test_sheet.loc[current_index, 'Unnamed: 0']
    
    # ... (Your existing st.session_state.test_sheet.loc assignments here) ...
    # This part writes to 'current_index' which is correct.
    # Ensure the columns exist
    for col in ['Unnamed: 1', 'Unnamed: 2','Unnamed: 4','Unnamed: 6','Unnamed: 8'
               ,'Unnamed: 9','Unnamed: 10','Unnamed: 11', 'Unnamed: 12', 'Unnamed: 13'
               ,'Unnamed: 14','Unnamed: 15','Unnamed: 16','Unnamed: 17', 'Unnamed: 18', 'Unnamed: 19']:
        if col not in st.session_state.test_sheet.columns:
            st.session_state.test_sheet[col] = None

    # Modify specific cells with current date and time
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 1'] = manifold_pressure
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 3'] = oil_temp
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 4'] = front_oil_p
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 5'] = rear_oil_p
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 6'] = metered_fuel_pressure
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 7'] = unmetered_fuel_pressure
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 8'] = egt_1
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 9'] = egt_2
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 10'] = egt_3
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 11'] = egt_4
    
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 14'] = cht_1
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 15'] = cht_2
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 16'] = cht_3
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 17'] = cht_4
    

    if cylinderCount == 6:
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 12'] = egt_5
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 13'] = egt_6
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 18'] = cht_5
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 19'] = cht_6

    else:
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 12'] = 0
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 13'] = 0
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 18'] = 0
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 19'] = 0

    # Check if cells are populated
    snap_cells = [(current_index, 'Unnamed: 1'), (current_index, 'Unnamed: 3'), (current_index, 'Unnamed: 4')
                  ,(current_index, 'Unnamed: 6'), (current_index, 'Unnamed: 6'), (current_index, 'Unnamed: 8')
                  ,(current_index, 'Unnamed: 9'), (current_index, 'Unnamed: 10'), (current_index, 'Unnamed: 11')
                  , (current_index, 'Unnamed: 12'), (current_index, 'Unnamed: 13'), (current_index, 'Unnamed: 14')
                  , (current_index, 'Unnamed: 15'), (current_index, 'Unnamed: 16'), (current_index, 'Unnamed: 17')
                  , (current_index, 'Unnamed: 18'), (current_index, 'Unnamed: 19')]

    # for row, col in snap_cells:
    #     cell_value = st.session_state.test_sheet.loc[row, col]
    #     is_empty = pd.isna(cell_value) or cell_value == ''
    #     if is_empty:
    #         st.session_state.snap_btn_toast = "Oops!! Sorry something went wrong, try again"
            
    #     else:
    #         st.session_state.snap_btn_toast = f"Snapshot taken. RPM: {row_rpm}."
            
    # st.toast(st.session_state.snap_btn_toast)
    
    # 2. MOVE TO NEXT RPM
    try:
        # Find index in the list ['2700', '2400'...]
        seq_idx = rpm_sequence.index(current_row_label)
        
        # Check if there is a next step
        if seq_idx < len(rpm_sequence) - 1:
            next_rpm = rpm_sequence[seq_idx + 1] # Get next lower RPM, e.g. "2400"
            next_row_index = row_map.get(next_rpm)
            
            if next_row_index:
                st.session_state.snapshot_index = next_row_index
                st.toast(f"Snapshot taken. Next: {next_rpm} RPM")
            else:
                st.error(f"Next RPM {next_rpm} not found in sheet.")
        else:
            # We are at the end (750 RPM)
            st.toast("All RPMs populated. Confirm Idle.")
            
    except ValueError:
        # current_row_label was not in our list (maybe user manually clicked elsewhere?)
        st.error("Lost track of RPM sequence.")
    


# def auto_snapshot_handler(df):
#     """
#     Automatically populates the test sheet when current RPM matches 
#     a target value (±15 RPM). Each target RPM row is filled only once 
#     per test run, tracked via st.session_state['snapshot_filled_flags'].
#     """
#     # Target RPMs (sheet order)
#     target_rpms = [2700, 2400, 2200, 2000, 1800, 1500, 1300, 1000, 700]
#     rpm_tolerance = 30

#     # Ensure flag dictionary exists
#     if 'snapshot_filled_flags' not in st.session_state:
#         st.session_state['snapshot_filled_flags'] = {rpm: False for rpm in target_rpms}

#     flags = st.session_state['snapshot_filled_flags']

#     # Basic checks
#     if df is None or df.empty:
#         return

#     if 'rpm' not in df.columns:
#         st.toast("RPM field missing from data.")
#         return

#     try:
#         current_rpm = float(df['rpm'].iloc[0])
#     except Exception:
#         st.toast("Unable to parse RPM value.")
#         return

#     # Find matching target (first match within tolerance)
#     matched_target = None
#     for target in target_rpms:
#         if abs(current_rpm - target) <= rpm_tolerance:
#             matched_target = target
#             # st.toast(f"Current RPM: {current_rpm} matched")
#             break

#     if matched_target is None:
#         # not inside any target window -> nothing to do
#         return

#     # Skip if this RPM was already filled
#     if flags.get(matched_target, False):
#         return  # silently skip

#     # Locate test sheet
#     sheet = st.session_state.get('test_sheet')
#     if sheet is None:
#         st.toast("Test sheet not loaded in session_state.")
#         return

#     # Find the row corresponding to the matched RPM
#     try:
#         if 'Unnamed: 0' in sheet.columns:
#             matches = sheet.index[sheet['Unnamed: 0'] == matched_target].tolist()
#             if matches:
#                 row_idx = matches[0]
#             else:
#                 pos = target_rpms.index(matched_target)
#                 if pos < len(sheet):
#                     row_idx = sheet.index[pos]
#                 else:
#                     st.toast("Could not locate the matching row in the test sheet.")
#                     return
#         else:
#             pos = target_rpms.index(matched_target)
#             if pos < len(sheet):
#                 row_idx = sheet.index[pos]
#             else:
#                 st.toast("Could not locate the matching row in the test sheet.")
#                 return
#     except Exception as e:
#         st.toast(f"Error locating row: {e}")
#         return

#     # Ensure columns exist
#     expected_columns = (
#         ['Unnamed: 1','Unnamed: 3','Unnamed: 4','Unnamed: 5','Unnamed: 6','Unnamed: 7']
#         + [f'Unnamed: {i}' for i in range(8, 14)]
#         + [f'Unnamed: {i}' for i in range(14, 20)]
#     )
#     for col in expected_columns:
#         if col not in sheet.columns:
#             st.session_state.test_sheet[col] = None

#     # Helper to safely read df fields
#     def safe_int_field(field_name, default=0):
#         try:
#             if field_name in df.columns:
#                 v = df[field_name].iloc[0]
#                 if pd.isna(v):
#                     return default
#                 return int(round(float(v)))
#         except Exception:
#             pass
#         return default

#     # Extract sensor data
#     manifold_pressure = safe_int_field('manifold_pressure')
#     front_oil_p = safe_int_field('front_oil_p')
#     rear_oil_p = safe_int_field('rear_oil_p')
#     oil_temp = safe_int_field('oil_temperature')
#     metered_fuel_pressure = safe_int_field('metered_fuel_pressure')
#     unmetered_fuel_pressure = safe_int_field('unmetered_fuel_pressure')
#     egt_vals = [safe_int_field(f'egt_{i}', 0) for i in range(1, 7)]
#     cht_vals = [safe_int_field(f'cht_{i}', 0) for i in range(1, 7)]

#     # Write values (first and only time)
#     st.session_state.test_sheet.loc[row_idx, 'Unnamed: 1'] = manifold_pressure
#     st.session_state.test_sheet.loc[row_idx, 'Unnamed: 3'] = oil_temp
#     st.session_state.test_sheet.loc[row_idx, 'Unnamed: 4'] = front_oil_p
#     st.session_state.test_sheet.loc[row_idx, 'Unnamed: 5'] = rear_oil_p
#     st.session_state.test_sheet.loc[row_idx, 'Unnamed: 6'] = metered_fuel_pressure
#     st.session_state.test_sheet.loc[row_idx, 'Unnamed: 7'] = unmetered_fuel_pressure

#     for j in range(6):
#         st.session_state.test_sheet.loc[row_idx, f'Unnamed: {8 + j}'] = egt_vals[j]
#         st.session_state.test_sheet.loc[row_idx, f'Unnamed: {14 + j}'] = cht_vals[j]

#     cylinder_count = st.session_state.get('cylinderCount', 4)
#     if cylinder_count == 4:
#         st.session_state.test_sheet.loc[row_idx, 'Unnamed: 12'] = 0
#         st.session_state.test_sheet.loc[row_idx, 'Unnamed: 13'] = 0
#         st.session_state.test_sheet.loc[row_idx, 'Unnamed: 18'] = 0
#         st.session_state.test_sheet.loc[row_idx, 'Unnamed: 19'] = 0

#     # Mark this RPM as filled
#     flags[matched_target] = True
#     st.session_state['snapshot_filled_flags'] = flags  # store back

#     # Feedback toast
#     st.toast(f"Snapshot taken — {matched_target} RPM recorded (row {int(row_idx)}).")

#     # Optionally record for reference
#     st.session_state['last_snapshot_rpm'] = matched_target
#     st.session_state['last_snapshot_row'] = int(row_idx)



# def clear_button_handler(df):
#     # Same RPMs as used in auto_snapshot_handler
#     target_rpms = [2700, 2400, 2200, 2000, 1800, 1500, 1300, 1000, 700]

#     # Columns that the auto_snapshot_handler populates
#     cols_to_clear = [
#         'Unnamed: 1', 'Unnamed: 3', 'Unnamed: 4', 'Unnamed: 5', 'Unnamed: 6',
#         'Unnamed: 7', 'Unnamed: 8', 'Unnamed: 9', 'Unnamed: 10', 'Unnamed: 11',
#         'Unnamed: 12', 'Unnamed: 13', 'Unnamed: 14', 'Unnamed: 15', 'Unnamed: 16',
#         'Unnamed: 17', 'Unnamed: 18', 'Unnamed: 19'
#     ]

#     # Ensure the columns exist
#     for col in cols_to_clear:
#         if col not in st.session_state.test_sheet.columns:
#             st.session_state.test_sheet[col] = None

#     # Iterate through each target RPM and clear only that row’s data
#     for rpm in target_rpms:
#         # Find the corresponding row by matching RPM value in the first column
#         row_index = st.session_state.test_sheet[
#             st.session_state.test_sheet['Unnamed: 0'] == rpm
#         ].index

#         # If found, clear its relevant columns
#         if not row_index.empty:
#             st.session_state.test_sheet.loc[row_index, cols_to_clear] = ""

#     # Reset auto handler state
#     st.session_state.snapshot_index = 0
#     if "filled_rpms" in st.session_state:
#         st.session_state.filled_rpms.clear()
#     st.session_state.auto_mode = False

#     # User feedback
#     st.toast("Auto snapshot rows cleared and state reset.")



# @st.fragment
def undo_button_handler(df):
    current_index = get_snapshot_index() + 1
    st.session_state.snapshot_index = current_index
    # If the snap boundary has been reached i.e the last row has been populated; undo the last row
    if current_index == 7:
        st.session_state.snapshot_index = 8
        return

    if current_index == 12:
            st.session_state.snapshot_index +=1 #current_index

    if current_index == 13:
        # current_index = current_index - 1 
        st.session_state.snapshot_index -=2
        return

    if current_index > 17:
        st.session_state.snapshot_index = 17
        st.toast(f"Limits exceeded. Current index: {current_index}")
        
        return
    
    # Read the rpm at current index from test sheet dataframe
    row_rpm = st.session_state.test_sheet.loc[current_index, 'Unnamed: 0']
    
    # Ensure the columns exist
    for col in ['Unnamed: 1', 'Unnamed: 2','Unnamed: 4','Unnamed: 5','Unnamed: 6'
               ,'Unnamed: 7','Unnamed: 8','Unnamed: 9','Unnamed: 10','Unnamed: 11'
               ,'Unnamed: 12', 'Unnamed: 13','Unnamed: 14','Unnamed: 15','Unnamed: 16'
               ,'Unnamed: 17','Unnamed: 18','Unnamed: 19']:
        
        if col not in st.session_state.test_sheet.columns:
            st.session_state.test_sheet[col] = None

    # Modify specific cells with current date and time
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 1'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 3'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 4'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 5'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 6'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 7'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 8'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 9'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 10'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 11'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 12'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 13'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 14'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 15'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 16'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 17'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 18'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 19'] = ''
    

    # Check if cells are populated
    snap_cells = [(current_index, 'Unnamed: 1'), (current_index, 'Unnamed: 3'), (current_index, 'Unnamed: 4')
                  , (current_index, 'Unnamed: 5'), (current_index, 'Unnamed: 6'), (current_index, 'Unnamed: 7')
                  , (current_index, 'Unnamed: 8'), (current_index, 'Unnamed: 9'), (current_index, 'Unnamed: 10')
                  , (current_index, 'Unnamed: 11'), (current_index, 'Unnamed: 12'), (current_index, 'Unnamed: 13')
                  , (current_index, 'Unnamed: 14'), (current_index, 'Unnamed: 15'), (current_index, 'Unnamed: 16')
                  , (current_index, 'Unnamed: 17'), (current_index, 'Unnamed: 18'), (current_index, 'Unnamed: 19')]

    for row, col in snap_cells:
        cell_value = st.session_state.test_sheet.loc[row, col]
        is_empty = pd.isna(cell_value) or cell_value == ''
        if is_empty:
            st.session_state.snap_btn_toast = f"RPM: {row_rpm} cleared"
            
        else:
            st.session_state.snap_btn_toast = "Oops!! Sorry something went wrong, try again"
            
    st.toast(st.session_state.snap_btn_toast)


# @st.fragment
# def end_button_handler(df):
#     # Get current UTC time and convert to AST
#     utc_now = datetime.datetime.now(datetime.UTC)
#     ast_tz = pytz.timezone('America/Halifax')  # Atlantic Standard Time (AST) timezone
#     ast_now = utc_now.replace(tzinfo=pytz.utc).astimezone(ast_tz)
#     end_time = ast_now.isoformat()  # Use AST timestamp
#     end_time_str = ast_now.strftime("%H:%M:%S")
#     outside_temp = int(df['oat'].iloc[0])
#     # now = datetime.now()
#     # end_time = now.strftime("%H:%M:%S")
#     st.session_state.end_time = end_time
 
#     # Ensure the columns 'Unnamed: 15' exist
#     if 'Unnamed: 15' not in st.session_state.test_sheet.columns:
#         st.session_state.test_sheet['Unnamed: 15'] = None

#     # Calculate the total time of the test
#     start_time_hold = st.session_state.start_time_str
#     end_time_hold = end_time_str

#     # Define a time format
#     time_format = "%H:%M:%S"

#     # Convert the time strings to datetime objects (using today's date)
#     now = datetime.datetime.now()
#     start_time_obj = datetime.datetime.strptime(start_time_hold, time_format).replace(year=now.year, month=now.month, day=now.day)
#     end_time_obj = datetime.datetime.strptime(end_time_hold, time_format).replace(year=now.year, month=now.month, day=now.day)

#     # start_time_obj = datetime.strptime(st.session_state.start_time, '%H:%M:%S')
#     # end_time_obj = datetime.strptime(end_time, '%H:%M:%S')
#     total_time = end_time_obj - start_time_obj

#     # Add outside air temperature
#     outside_temp = int(df['oat'].iloc[0])
#     st.session_state.test_sheet.loc[3, 'Unnamed: 8'] = outside_temp

#     # Modify specific cells with current date and time
#     st.session_state.test_sheet.loc[3, 'Unnamed: 15'] = end_time_str
#     st.session_state.test_sheet.loc[4, 'Unnamed: 15'] = str(total_time)
    
#     # Add Idle Fuel Pressure
#     st.session_state.test_sheet.loc[21, 'Unnamed: 1'] = st.session_state.test_sheet.loc[8, 'Unnamed: 6']
#     st.session_state.test_sheet.loc[21, 'Unnamed: 4'] = st.session_state.test_sheet.loc[8, 'Unnamed: 7']
    
#     # Add Idle RPM
#     st.session_state.test_sheet.loc[23, 'Unnamed: 1'] = st.session_state.test_sheet.loc[8, 'Unnamed: 0']
#     st.session_state.test_sheet.loc[23, 'Unnamed: 4'] = st.session_state.test_sheet.loc[8, 'Unnamed: 1']

#     # Add Idle Oil Pressure
#     st.session_state.test_sheet.loc[21, 'Unnamed: 8'] = st.session_state.test_sheet.loc[8, 'Unnamed: 5']
#     st.session_state.test_sheet.loc[21, 'Unnamed: 12'] = st.session_state.test_sheet.loc[8, 'Unnamed: 3']

#     # Add Full Throttle Fuel Pressure
#     st.session_state.test_sheet.loc[22, 'Unnamed: 1'] = st.session_state.test_sheet.loc[17, 'Unnamed: 6']
#     st.session_state.test_sheet.loc[22, 'Unnamed: 4'] = st.session_state.test_sheet.loc[17, 'Unnamed: 7']

#     # Add Full Throttle RPM
#     st.session_state.test_sheet.loc[24, 'Unnamed: 1'] = st.session_state.test_sheet.loc[17, 'Unnamed: 0']
#     st.session_state.test_sheet.loc[24, 'Unnamed: 4'] = st.session_state.test_sheet.loc[17, 'Unnamed: 1']

#     # Add Full Oil Pressure
#     st.session_state.test_sheet.loc[23, 'Unnamed: 8'] = st.session_state.test_sheet.loc[17, 'Unnamed: 5']
#     st.session_state.test_sheet.loc[23, 'Unnamed: 12'] = st.session_state.test_sheet.loc[17, 'Unnamed: 3']

#     # Check if cells are populated
#     end_cells = [(3, 'Unnamed: 15'), (4, 'Unnamed: 15'), (21, 'Unnamed: 1'), (21, 'Unnamed: 4')
#                  , (23, 'Unnamed: 1'), (23, 'Unnamed: 4'), (21, 'Unnamed: 8'), (21, 'Unnamed: 12')
#                  , (22, 'Unnamed: 1'), (22, 'Unnamed: 4'), (24, 'Unnamed: 1'), (24, 'Unnamed: 4')
#                  , (23, 'Unnamed: 8'), (23, 'Unnamed: 12')]

#     for row, col in end_cells:
#         cell_value = st.session_state.test_sheet.loc[row, col]
#         is_empty = pd.isna(cell_value) or cell_value == ''
#         if is_empty:
#             st.session_state.end_btn_toast = "Oops!! Sorry something went wrong, try again"
#             # time.sleep(1)
#         else:
#             st.session_state.end_btn_toast = "Finish time & total time fields populated"
#             # time.sleep(1)

#     st.toast(st.session_state.end_btn_toast)

def end_button_handler(df):
    row_map = get_row_map(st.session_state.test_sheet)
    
    # ... (Time calculation logic remains the same) ...
    # Get current UTC time and convert to AST
    utc_now = datetime.datetime.now(datetime.UTC)
    ast_tz = pytz.timezone('America/Halifax')  # Atlantic Standard Time (AST) timezone
    ast_now = utc_now.replace(tzinfo=pytz.utc).astimezone(ast_tz)
    end_time = ast_now.isoformat()  # Use AST timestamp
    end_time_str = ast_now.strftime("%H:%M:%S")
    outside_temp = int(df['oat'].iloc[0])
    # now = datetime.now()
    # end_time = now.strftime("%H:%M:%S")
    st.session_state.end_time = end_time

    # Ensure the columns 'Unnamed: 15' exist
    # if 'Unnamed: 15' not in st.session_state.test_sheet.columns:
    #     st.session_state.test_sheet['Unnamed: 16'] = None
    if 'Unnamed: 16' not in st.session_state.test_sheet.columns:
        st.session_state.test_sheet['Unnamed: 16'] = None

    # Calculate the total time of the test
    start_time_hold = st.session_state.start_time_str
    end_time_hold = end_time_str

    # Define a time format
    time_format = "%H:%M:%S"

    # Convert the time strings to datetime objects (using today's date)
    now = datetime.datetime.now()
    start_time_obj = datetime.datetime.strptime(start_time_hold, time_format).replace(year=now.year, month=now.month, day=now.day)
    end_time_obj = datetime.datetime.strptime(end_time_hold, time_format).replace(year=now.year, month=now.month, day=now.day)

    # start_time_obj = datetime.strptime(st.session_state.start_time, '%H:%M:%S')
    # end_time_obj = datetime.strptime(end_time, '%H:%M:%S')
    total_time = end_time_obj - start_time_obj

    # Add outside air temperature
    outside_temp = int(df['oat'].iloc[0])
    # st.session_state.test_sheet.loc[3, 'Unnamed: 8'] = outside_temp
    st.session_state.test_sheet.loc[3, 'Unnamed: 9'] = outside_temp

    # Modify specific cells with current date and time
    # st.session_state.test_sheet.loc[3, 'Unnamed: 15'] = end_time_str
    # st.session_state.test_sheet.loc[4, 'Unnamed: 15'] = str(total_time)
    st.session_state.test_sheet.loc[3, 'Unnamed: 16'] = end_time_str
    st.session_state.test_sheet.loc[4, 'Unnamed: 16'] = str(total_time)
    
    # --- ANCHOR BASED FOOTER POPULATION ---
    try:
        # Find Footer Rows
        r_idle_fuel = row_map.get('Idle Fuel Pressure') # Or whatever the label is in the footer
        r_ft_fuel = row_map.get('Fuel Throttle Fuel Press') # Check exact label in Excel
        r_idle_rpm = row_map.get('Idle RPM')
        r_ft_rpm = row_map.get('Full Throttle RPM')
        
        # Find Source Data Rows (The actual RPM test rows)
        r_src_idle = row_map.get('750') # Idle RPM row
        
        # Determine Full Throttle RPM Source (2700 for 2-blade, 2200 for 4-blade)
        prop_type = st.session_state.get("selected_prop")
        # ft_rpm_label = "2700" if "2 Blade" in prop_type else "2200"
        if prop_type == "2 Blade (Yellow Tipped)":
            ft_rpm_label = "2700"
        else:
            ft_rpm_label = "2200"
        r_src_ft = row_map.get(ft_rpm_label)

        if not r_src_idle or not r_src_ft:
            st.error("Could not find source data rows (750 or Max RPM)")
            return

        # 1. Populate Idle Section
        if r_idle_fuel:
            # Idle Fuel Press (Metered/Unmetered)
            # st.session_state.test_sheet.loc[r_idle_fuel, 'Unnamed: 1'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 6']
            # st.session_state.test_sheet.loc[r_idle_fuel, 'Unnamed: 4'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 7']
            st.session_state.test_sheet.loc[r_idle_fuel, 'Unnamed: 2'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 6']
            st.session_state.test_sheet.loc[r_idle_fuel, 'Unnamed: 5'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 7']
            # Idle Oil Press
            # st.session_state.test_sheet.loc[r_idle_fuel, 'Unnamed: 8'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 5']
            # st.session_state.test_sheet.loc[r_idle_fuel, 'Unnamed: 12'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 3']
            st.session_state.test_sheet.loc[r_idle_fuel, 'Unnamed: 9'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 5']
            st.session_state.test_sheet.loc[r_idle_fuel, 'Unnamed: 13'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 3']

        # 2. Populate Full Throttle Section
        if r_ft_fuel:
            # F/T Fuel Press (Metered/Unmetered)
            # st.session_state.test_sheet.loc[r_ft_fuel, 'Unnamed: 1'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 6']
            # st.session_state.test_sheet.loc[r_ft_fuel, 'Unnamed: 4'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 7']
            st.session_state.test_sheet.loc[r_ft_fuel, 'Unnamed: 2'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 6']
            st.session_state.test_sheet.loc[r_ft_fuel, 'Unnamed: 5'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 7']
            
        if r_idle_rpm:
            # Idle RPM @ Manifold Pressure
            # st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 1'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 0']
            # st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 4'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 1']
            st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 2'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 0']
            st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 5'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 1']
            # F/T Oil Press
            # st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 8'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 5']
            # st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 12'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 3']
            st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 9'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 5']
            st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 13'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 3']

        if r_ft_rpm:
            # F/T RPM @ Manifold Pressure
            # st.session_state.test_sheet.loc[r_ft_rpm, 'Unnamed: 1'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 0']
            # st.session_state.test_sheet.loc[r_ft_rpm, 'Unnamed: 4'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 1']
            st.session_state.test_sheet.loc[r_ft_rpm, 'Unnamed: 2'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 0']
            st.session_state.test_sheet.loc[r_ft_rpm, 'Unnamed: 5'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 1']

        st.toast("Footer populated successfully")
        
    except Exception as e:
        st.error(f"Error populating footer: {str(e)}")



# Function to start the subprocess and store its information in session_state
def start_process():
    # Get the current date and time from session state
    now = datetime.datetime.now()
    # current_date = st.session_state.test_date
    current_date = now.strftime("%Y-%m-%d")
    utc_now = datetime.datetime.now(datetime.UTC)
    ast_tz = pytz.timezone('America/Halifax')
    ast_now = utc_now.replace(tzinfo=pytz.utc).astimezone(ast_tz)
    current_time = ast_now.strftime("%H:%M:%S")
    # current_time = st.session_state.start_time_str

    # Check id db_name and collection_name are in session state
    if 'db_name' not in st.session_state:
        st.session_state.db_name = None
    elif 'collection_name' not in st.session_state:
        st.session_state.collection_name = None
    
    #  Define the database and collection names
    # db_name = current_date
    # collection_name = f"{current_date}_{current_time}"

    split_engine_str = engine_type.split() 
    engineMake = split_engine_str[0]
    engineModel = split_engine_str[1]
    db_name = engineMake + "-" + engineModel + "-" + current_date
    collection_name = current_time


    # Store the db_name & collection name in session state
    st.session_state.db_name = db_name
    st.session_state.collection_name = collection_name

    if 'subprocess' not in st.session_state:
        process = subprocess.Popen(["python", r"C:\Users\KieranCalder\Code\AerotecTestCell\Parse_data.py", db_name, collection_name])
        st.session_state.subprocess = process
        st.toast(f"Process started. Writing to {db_name}_{collection_name}")

# Function to stop the subprocess if it's running
def stop_process():
    if 'subprocess' in st.session_state:
        process = st.session_state.subprocess
        process.terminate()  # This sends SIGTERM to the process
        st.toast("Process terminated")
        del st.session_state.subprocess  # Remove it from session_state

# Define a function to display the grid and capture changes
def display_ag_grid(df):
    # Build grid options
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(editable=True)
    grid_options = gb.build()

    # Display grid and return data based on DataReturnMode.AS_INPUT
    grid_response = AgGrid(
        df,
        gridOptions=grid_options,
        data_return_mode=DataReturnMode.AS_INPUT,
        update_mode=GridUpdateMode.VALUE_CHANGED,
        fit_columns_on_grid_load=True
    )

    # Extract the updated DataFrame
    # return pd.DataFrame(grid_response['data'])
    return grid_response

# def copy_to_excel(df, template_path, thresholds, engine_type):       #"""What does df and thresholds do??"""
#    wb = load_workbook(template_path)
#    ws = wb.active

   

# #    edited_cells = st.session_state.test_sheet.loc[2, 'Unnamed: 8']
# #    ws['I4'] = edited_cells
#    row_offset = 1  # Offset by 1 to account for the extra row at the top of the Excel sheet

#    # Define fill styles for different ranges
#    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Below minimum
#    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")   # Nominal and high ranges
#    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Between high and 0.9 * max
#    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")     # Above maximum

#    for r_idx, row in enumerate(dataframe_to_rows(st.session_state.test_sheet, index=False, header=False), start=1):
#         actual_row = r_idx + row_offset  # Adjusted row considering the offset
        
#         # Skip row 15
#         if actual_row == 15:
#             continue

#         for c_idx, value in enumerate(row, start=1):
#             cell = ws.cell(row=actual_row, column=c_idx)

#             # Check if the cell is part of a merged cell range
#             if cell.coordinate in ws.merged_cells:
#                 # Get the top-left cell of the merged range
#                 for merged_range in ws.merged_cells.ranges:
#                     min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
#                     if min_row <= actual_row <= max_row and min_col <= c_idx <= max_col:
#                         # Write to the top-left cell of the merged range
#                         if engine_type == dataOnly:
#                             top_left_cell = ws.cell(row=min_row, column=min_col, value=value)
#                         else:
#                             top_left_cell = ws.cell(row=min_row, column=min_col, value=value)
#                             if 10 <= min_row <= 19 and isinstance(value, (int, float)):
#                                 threshold = thresholds.get(c_idx)
#                                 if threshold is not None:
#                                     min_val, nominal_val, high_val, max_val = threshold
#                                     # Apply formatting based on thresholds
#                                     if value < min_val:
#                                         ws.cell(row=min_row, column=min_col).fill = green_fill
#                                     elif min_val <= value <= nominal_val:
#                                         ws.cell(row=min_row, column=min_col).fill = green_fill
#                                     elif nominal_val < value <= high_val:
#                                         ws.cell(row=min_row, column=min_col).fill = green_fill
#                                     elif high_val < value <  0.9 * max_val:
#                                         ws.cell(row=min_row, column=min_col).fill = yellow_fill
#                                     else:
#                                         ws.cell(row=min_row, column=min_col).fill = red_fill
#                             break
#             else:
#                 if engine_type == dataOnly:
#                     cell.value = value
#                 else:
#                     cell.value = value
#                     # Apply formatting if within the specified row range and if the value is numeric
#                     if 10 <= actual_row <= 19 and isinstance(value, (int, float)):
#                         threshold = thresholds.get(c_idx)
                        
#                         if threshold is not None:
#                             min_val, nominal_val, high_val, max_val = threshold
#                             # Apply formatting based on thresholds
#                             # st.write(cell)
#                             # st.write(f"Value: {value}, Min: {min_val}, Nominal: {nominal_val}, High: {high_val}, Max: {max_val}")
#                             if value < min_val:
#                                 cell.fill = green_fill
#                             elif min_val <= value <= nominal_val:
#                                 cell.fill = green_fill
#                             elif nominal_val <= value <= high_val:
#                                 cell.fill = green_fill
#                             elif high_val < value <  0.9 * max_val:
#                                 cell.fill = yellow_fill
#                             else:
#                                 cell.fill = red_fill
#    buffer = BytesIO()
#    wb.save(buffer)
#    buffer.seek(0)

#    return buffer

def copy_to_excel(df, template_path, thresholds, engine_type):
    wb = load_workbook(template_path)
    ws = wb.active

    # 1. Define the Crosshatch Style (Diagonal Stripes)
    crosshatch_fill = PatternFill(
        patternType='darkTrellis', # This creates the diagonal stripes
        fgColor="000000",        # Pattern color (Black)
        bgColor="FFFFFF"         # Background (Transparent/White)
    )

    # 2. Get Engine Specs
    is_turbo = st.session_state.get('turbo', False)
    num_cylinders = st.session_state.get('cylinders', 6)
    prop_type = st.session_state.get('prop_type', '')
    
    # 3. Create Destination Map { "Label": Row_Index }
    dest_map = {}
    for i, row in enumerate(ws.iter_rows(min_col=1, max_col=1, values_only=True), start=1):
        if row[0]:
            # dest_map[str(row[0]).strip()] = i
            label = str(row[0]).strip()
            if label:
                dest_map[label] = i
            
    # --- RULE: Prop Type Masking ---
    if prop_type != "2 Blade (Yellow Tipped)":
        for label in ["2700", "2400"]:
            if label in dest_map:
                r = dest_map[label]
                for c in range(2, 21): 
                    ws.cell(row=r, column=c).fill = crosshatch_fill

    # 4. Iterate and Fill Data
    # for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False)):
    for row_data in dataframe_to_rows(df, index=False, header=False):
        source_label = str(row_data[0]).strip() if row_data[0] else None
        
        if source_label in dest_map:
            target_row = dest_map[source_label]
        # elif r_idx < 10:
        #     target_row = r_idx + 1
        # else:
        #     continue

            for c_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=target_row, column=c_idx)
                
                # --- MERGED CELL SAFE WRITING ---
                # If the cell is merged, we must write to the top-left cell of the range
                write_cell = cell
                if cell.coordinate in ws.merged_cells:
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # Find the top-left cell of the range
                            # min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                            min_col, min_row, _, _ = range_boundaries(str(merged_range))
                            write_cell = ws.cell(row=min_row, column=min_col)
                            break
                
                # Write the value if it exists
                # if not pd.isna(value) and value != "":
                if c_idx > 1 and not pd.isna(value) and value != "":
                    write_cell.value = value

                # --- FORMATTING RULES ---
                if engine_type != 'dataOnly':
                    # if source_label and (source_label.isdigit() or source_label == "750"):
                    # Only apply to RPM rows (numbers or "750")
                    if source_label.isdigit() or source_label == "750":
                        # Cylinder Rule
                        if num_cylinders == 4 and c_idx in [13, 14, 19, 20]:
                            cell.fill = crosshatch_fill
                        # Turbo Rule
                        if not is_turbo and c_idx == 3: # Unnamed: 2
                            cell.fill = crosshatch_fill

        else:
            # If the row doesn't have a label in Column A, we skip it entirely.
            # This automatically ignores "AIR PRESSURES", "Engine RPM", etc.
            continue

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer



# UPDATE THIS PATH: The specific folder where you want files to appear
testsheet_folder = r"K:\New Test Cell\Test Sheets"

# --- 2. CALLBACK FUNCTION ---
def save_excel_callback(buffer, file_name):
    """Callback to save the excel buffer to a specific local directory."""
    try:
        if not file_name or file_name.strip() == ".xlsx":
            st.error("⚠️ Please enter a valid file name.")
            return

        # Ensure the directory exists
        if not os.path.exists(testsheet_folder):
            os.makedirs(testsheet_folder)

        # Construct full path
        full_save_path = os.path.join(testsheet_folder, file_name)

        # Write buffer to disk
        with open(full_save_path, "wb") as f:
            f.write(buffer.getbuffer())
        
        st.toast(f"✅ File successfully saved to: {full_save_path}")

    except Exception as e:
        st.error(f"❌ Error saving file: {str(e)}")
   

def main():
    """
    Main function to run the Streamlit app.
    """

    df = get_dataframe()

    st.markdown("""
    <style>
    
           /* Remove blank space at top and bottom */ 
           .block-container {
               padding-top: 15px;
               padding-bottom: 15px;
               padding-left: 25px;
               padding-right: 25px; 
            }
           
           
    </style>
    """, unsafe_allow_html=True)

    # st.write(df)

    # st.write(f"Database Name: {st.session_state.db_name}")
    # st.write(f"Collection Name: {st.session_state.collection_name}")

    # --- Maintain Auto state ---
    # if "auto_mode" not in st.session_state:
    #     st.session_state.auto_mode = False
    
    # Using session state to track components
    if 'components_created' not in st.session_state:
        st.session_state['components_created'] = False

    if not st.session_state['components_created']:

        col1, col2, col3, col4, col5 = st.columns([1, 1, 1.5, 1, 1], gap="small", vertical_alignment='center')  # Create three columns side by side
        with col1:
            with st.container(height=130):
                cola, colb, colc = st.columns([1.5, 6, 0.2])
                with cola:
                    st.write("")
                with colb:
                    performance_fig = plot_performance_metrics()
                    st.plotly_chart(performance_fig, use_container_width=False)
                with colc:
                    st.write("")
                

        with col2:
            with st.container(height=130):
                cola, colb, colc = st.columns([1.5, 6, 0.2])
                with cola:
                    st.write("")
                with colb:
                    rpm_fig = animate_rpm(df, engine_type)
                    st.plotly_chart(rpm_fig, use_container_width=False)
                with colc:
                    st.write("")

        with col3:
            with st.container(height=130):
                cola, colb, colc, cold, cole, colf = st.columns([0.01, 1.3, 1.3, 1.3, 1.3, 0.01], vertical_alignment='top')
                with colb:
                    st.write("")
                    st.write("")
                    # start_btn = st.button("START")
                    # FIX: Use on_click callback
                    # We define a lambda or wrapper if we need to call multiple functions
                    def handle_start():
                        start_process()
                        start_button_handler(df)

                    st.button("START", on_click=handle_start)
                with colc:
                    st.write("")
                    st.write("")
                    # FIX: Pass the dataframe via args
                    st.button("SNAP", on_click=snap_button_handler, args=(df,))
                    # snap_btn = st.button("SNAP")
                    # auto_btn = st.button("AUTO")
                    # st.toast("AUTO mode ON" if st.session_state.auto_mode else "AUTO mode OFF")
                with cold:
                    st.write("")
                    st.write("")
                    # FIX: Undo callback
                    st.button("UNDO", on_click=undo_button_handler, args=(df,))
                    # undo_btn = st.button("UNDO")
                    # clear_btn = st.button("CLEAR")
                with cole:
                    st.write("")
                    st.write("")
                    # FIX: End callback
                    def handle_end():
                        end_button_handler(df)
                        stop_process()
                        
                    st.button("END", on_click=handle_end)
                    # end_btn = st.button("END")

        with col4:
            with st.container(height=130):
                cola, colb, colc = st.columns([1.5, 6, 0.2])
                with cola:
                    st.write("")
                with colb:
                    manifold_pressure_fig = animate_manifold_pressure(df, engine_type)
                    st.plotly_chart(manifold_pressure_fig, use_container_width=False)
                with colc:
                    st.write("")

        with col5:
            with st.container(height=130):
                cola, colb, colc = st.columns([1.5, 6, 0.2])
                with cola:
                    st.write("")
                with colb:
                    horsepower_fig = animate_hp(df, engine_type)
                    st.plotly_chart(horsepower_fig, use_container_width=False)
                with colc:
                    st.write("")

        with st.container(height=400):
            col6, col7 = st.columns(2, vertical_alignment='center')
            with col6:
                if df.empty:
                    st.write("Turn on MASTER, and kindly click START to populate the plot")
                else:
                    temps_bar_plot = plot_egt_cht(df, engine_type)
                    st.plotly_chart(temps_bar_plot, use_container_width=False)

            with col7:
                display_metrics(df, engine_type)

        # if start_btn:
        #     start_process()
        #     # st.toast("HELLLLOOO!!!!")
        #     start_button_handler(df)
        #     # st.toast("THERE!!!!!!")
        #     # if df.empty:
        #     #      time.sleep(15)
        #     # else: 
        #     #     st.write("HELLLLOOO!!!!")
        #     #     start_button_handler(df)
        #     #     st.write("THERE!!!!!!")
        #     #     # st.write(st.session_state.start_btn_toast)
        #     time.sleep(refresh_time)
        # if snap_btn:
        # # if auto_btn:
        #     snap_button_handler(df)
        #     time.sleep(refresh_time)
        #     st.session_state.auto_mode = not st.session_state.auto_mode
        # if undo_btn:
        # # if clear_btn:
        #     undo_button_handler(df)
        #     # clear_button_handler(df)
        #     time.sleep(refresh_time)
        # if end_btn:
        #     end_button_handler(df) 
        #     stop_process()
        #     time.sleep(refresh_time)

        # --- AUTO Mode Behavior ---
        # if st.session_state.auto_mode:
        #     auto_snapshot_handler(df)
        #     st.toast("AUTO snapshot running...")


        # Display the engine test sheet dataframe using streamlit aggrid
        # And update session state
        grid_response = display_ag_grid(st.session_state.test_sheet)
        if not grid_response['data'].empty:
            updated_data = pd.DataFrame(grid_response['data'])
            for i in range(len(updated_data)):
                for col in updated_data.columns:
                    st.session_state.test_sheet.iloc[i, st.session_state.test_sheet.columns.get_loc(col)] = updated_data.iloc[i][col]

        excel = r"C:\Users\KieranCalder\Code\AerotecTestCell\Engine Test Sheet Template.xlsx"

        # if prop_type == "2 Blade (Yellow Tipped)":
        #     excel = r"C:\Users\KieranCalder\Code\AerotecTestCell\Engine Test Sheet Template (2 Blade).xlsx"
        # else:
        #     excel = r"C:\Users\KieranCalder\Code\AerotecTestCell\Engine Test Sheet Template (4 Blade).xlsx"

        col8, col9, col10 = st.columns([1.2, 1.2, 1], vertical_alignment='bottom')
        with col8:
            pass
        with col9:

            # Prompt the user for a file name
            file_name_input = st.text_input("Enter the desired file name (without extension):")
            output_path = f"{file_name_input}.xlsx"

            colf, colg, colh = st.columns([1.2, 1.2, 1], vertical_alignment='bottom')
            with colf:
                pass
            with colg:
                # Add a download button for the CSV file
                # if st.button("Download as XLSX"):
                buffer = copy_to_excel(st.session_state.test_sheet, excel, thresholds, engine_type)
                
                # st.download_button(
                #     label="Download file",
                #     data=buffer,
                #     file_name=output_path,
                #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                # )

                # Replaced download_button with standard button + callback
                st.button(
                    label="Save to Folder",
                    on_click=save_excel_callback,
                    args=(buffer, output_path)  # Pass the buffer and name to the callback
                )

                #"K:\New Test Cell\Test Sheets"

                
            with colh:
                pass
        with col10:
            pass

if __name__ == "__main__":
    main()
    # subprocess.Popen(["python", r"C:\Users\theca\Documents\Victor\Aerotec\MVP 50\Code\AerotecTestCell\Parse_data.py"])

