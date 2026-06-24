import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st
from pymongo import MongoClient
import datetime
from matplotlib.animation import FuncAnimation
import time
import datetime
from datetime import datetime
from streamlit_autorefresh import st_autorefresh
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np
from st_aggrid import AgGrid
from st_aggrid import GridOptionsBuilder, GridUpdateMode, DataReturnMode, JsCode
import asyncio
import motor.motor_asyncio
from motor.motor_asyncio import AsyncIOMotorClient
import subprocess
import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import range_boundaries
import commentjson as json
from io import BytesIO
from pages.st_select_engine import load_config
from st_main import pages
# from collections import OrderedDict

# st.set_page_config(layout="wide", page_title="Dashboard", page_icon=st.session_state.page_icon, initial_sidebar_state="collapsed")

# Ensure the user has selected an engine
if 'selected_engine' not in st.session_state: #or st.session_state.selected_engine is None:
    st.warning("No engine selected. Redirecting to select engine page...")
    time.sleep(3)
    st.switch_page(pages["Select Engine"])  # Redirect back to engine selection
else:
    # st.title(f"Dashboard for {st.session_state.selected_engine}")
    # Display the dashboard content for the selected engine
    st.write(f"Showing data for {st.session_state.selected_engine}")
    

# Add Aerotec Logo
# st.session_state.logo = r"C:\Users\theca\Documents\Victor\Aerotec\MVP 50\Code\AerotecTestCell\ae-logo_svg.svg"
st.logo(st.session_state.logo)

# Use the full page instead of a narrow central column
# st.session_state.page_icon = r"C:\Users\theca\Documents\Victor\Aerotec\MVP 50\Code\AerotecTestCell\ae-logo_jpg.jpeg"


engine_type = st.session_state.selected_engine


engine_config = load_config()

# st.write(engine_config)

## Function to initialize engine parameters in st.session_state
def initialize_engine_parameters(engine_type, config):
    """
    Initialize engine parameters in Streamlit's session state.
    
    Args:
    - engine_type (str): The type of engine.
    - config (dict): The loaded configuration dictionary.
    """
    # Ensure that the engine_type exists in the config
    if engine_type not in config:
        st.error(f"Engine type '{engine_type}' not found in configuration.")
        return

    # Loop through the parameters of the specified engine type
    engine_parameters = config[engine_type]
    for key, value in engine_parameters.items():
        # Set the parameters in session_state
        if key not in st.session_state:
            st.session_state[key] = value

initialize_engine_parameters(engine_type, engine_config)

# st.write(st.session_state.get('minMP'))

# config = engine_config[engine_type]

# Define Global Excel Doc Thresholds and Sort by Columns
global thresholds 
thresholds = {
    2:(st.session_state.get('minMP'), st.session_state.get('nominalMP'), st.session_state.get('highMP'), st.session_state.get('maxMP')), 
    4:(st.session_state.get('minOilTemp'), st.session_state.get('nominalOilTemp'), st.session_state.get('highOilTemp'), st.session_state.get('maxOilTemp')), 
    5:(st.session_state.get('minFrontOilPressure'), st.session_state.get('nominalFrontOilPressure'), st.session_state.get('highFrontOilPressure'), st.session_state.get('maxOilFrontOilPressure')),
    6:(st.session_state.get('minRearOilPressure'), st.session_state.get('nominalRearOilPressure'), st.session_state.get('highRearOilPressure'), st.session_state.get('maxOilRearOilPressure')), 
    7:(st.session_state.get('minMeteredFuelPressure'), st.session_state.get('nominalMeteredFuelPressure'), st.session_state.get('highMeteredFuelPressure'), st.session_state.get('maxMeteredFuelPressure')),
    8:(st.session_state.get('minUnmeteredFuelPressure'), st.session_state.get('nominalUnmeteredFuelPressure'), st.session_state.get('highUnmeteredFuelPressure'), st.session_state.get('maxUnmeteredFuelPressure')),
    9:(st.session_state.get('minEGT'), st.session_state.get('nominalEGT'), st.session_state.get('highEGT'), st.session_state.get('maxEGT')),
    10:(st.session_state.get('minEGT'), st.session_state.get('nominalEGT'), st.session_state.get('highEGT'), st.session_state.get('maxEGT')),
    11:(st.session_state.get('minEGT'), st.session_state.get('nominalEGT'), st.session_state.get('highEGT'), st.session_state.get('maxEGT')),
    12:(st.session_state.get('minEGT'), st.session_state.get('nominalEGT'), st.session_state.get('highEGT'), st.session_state.get('maxEGT')),
    15:(st.session_state.get('minCHT'), st.session_state.get('nominalCHT'), st.session_state.get('highCHT'), st.session_state.get('maxCHT')),
    16:(st.session_state.get('minCHT'), st.session_state.get('nominalCHT'), st.session_state.get('highCHT'), st.session_state.get('maxCHT')),
    17:(st.session_state.get('minCHT'), st.session_state.get('nominalCHT'), st.session_state.get('highCHT'), st.session_state.get('maxCHT')),
    18:(st.session_state.get('minCHT'), st.session_state.get('nominalCHT'), st.session_state.get('highCHT'), st.session_state.get('maxCHT'))
}

refresh_time = 1

# # MongoDB credentials and connection string
# username = urllib.parse.quote_plus('Aerotec')  # Add your username here
# password = urllib.parse.quote_plus('1@QWASZX')  # Add your password here

# # MongoDB Atlas Cloud Database Connection
# uri = "mongodb+srv://{}:{}@atlascluster.qrjrg5e.mongodb.net/".format(username, password)
# client = motor.motor_asyncio.AsyncIOMotorClient(uri)
# db = client['RTDO']  # Replace 'MVP50' with your database name

# MongoDB Local Database Connection
uri = "mongodb://localhost:27017/"
client = motor.motor_asyncio.AsyncIOMotorClient(uri)
db = client['RTDO']

# Single collection for all data points
collection = db['4_Cyl']


@st.cache_data
def test_sheet_df():
    # test_sheet = pd.read_csv(r"C:\Users\theca\Documents\Victor\Aerotec\MVP 50\Code\Engine Test Sheet Template.csv")
    test_sheet = pd.read_excel(r"C:\Users\theca\Documents\Victor\Aerotec\MVP 50\Code\AerotecTestCell\Engine Test Sheet Template.xlsx")

    return(test_sheet)

# Initialize session state if not already done
if 'test_sheet' not in st.session_state:
    st.session_state.test_sheet = test_sheet_df()
    st.toast("Cache Empty. Loading test sheet template...")

# Function to get or initialize the row index for snapshots
def get_snapshot_index():
    if 'snapshot_index' not in st.session_state:
        st.session_state.snapshot_index = 17
        # remind user to confirm full throttle RPM
        st.toast("Confirm full throttle RPM")

    return st.session_state.snapshot_index

st_autorefresh(refresh_time*1000, key="data_refresh")  # Refresh at a specified interval

async def fetch_data():
    num_documents = 5
    
    pipeline = [
        {"$sort": {"timestamp": -1}},  # Sort by timestamp, latest first
        {"$limit": num_documents},  # Limit the number of documents
        {"$project": {
            "timestamp": 1,
            "volts": "$fields.volts",
            "manifold_pressure": "$fields.manifold_pressure",
            "rpm": "$fields.rpm",
            "hp": "$fields.hp",
            "fuel_flow": "$fields.fuel_flow",
            "fuel_pressure": "$fields.fuel_pressure",
            "oil_pressure": "$fields.oil_pressure",
            "oil_pressure_2": "$fields.oil_pressure_2", 
            "oil_temperature": "$fields.oil_temperature",
            "amps": "$fields.amps",
            "cdt": "$fields.cdt",
            "iat_carb": "$fields.iat_carb",
            "oat": "$fields.oat",
            "mag_drop":"$fields.mag_drop",
            "egt_1": "$fields.egt_1",
            "egt_2": "$fields.egt_2",
            "egt_3": "$fields.egt_3",
            "egt_4": "$fields.egt_4",
            "cht_1": "$fields.cht_1",
            "cht_2": "$fields.cht_2",
            "cht_3": "$fields.cht_3",
            "cht_4": "$fields.cht_4",
        }}
    ]
    
    cursor = collection.aggregate(pipeline)
    data_list = []
    async for document in cursor:
        data_list.append(document)
    
    return data_list


def get_dataframe():
    start_time = time.time()  # Start time for performance measurement

    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        data_list = loop.run_until_complete(fetch_data())
        loop.close()
        
        if not data_list:
            st.write("No data fetched from MongoDB.")
            return pd.DataFrame()
        
        df = pd.DataFrame(data_list)
        
        if df.empty:
            st.write("DataFrame is empty.")
            return df
        
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        
        numeric_fields = [
            'volts', 'manifold_pressure', 'rpm', 'hp', 'fuel_flow',
            'fuel_pressure', 'oil_pressure', 'oil_pressure_2' 'oil_temperature',
            'amps', 'cdt', 'iat_carb', 'oat',
            'egt_1', 'egt_2', 'egt_3', 'egt_4',
            'cht_1', 'cht_2', 'cht_3', 'cht_4'
        ]
        
        # This commented section is for adding random variables during presentation
        # Convert numeric fields to floats
        for field in numeric_fields:
            if field in df.columns:
                df[field] = pd.to_numeric(df[field], errors='coerce')
        
        for i in range(len(df)):
            for field in numeric_fields:
                if field in df.columns and not pd.isnull(df.at[i, field]):
                    # Simulate changing data by adding a small random variation
                    df.at[i, field] = round(df.at[i, field] + random.uniform(-5, 5), 2)

        # for field in numeric_fields:
        #     if field in df.columns:
        #         df[field] = pd.to_numeric(df[field], errors='coerce') #+ random.uniform(-5,5)
        #     else:
        #         df[field] = None  # Handle missing fields
        
        df['elapsed_time'] = (df['timestamp'] - df['timestamp'].max()).dt.total_seconds()
        
        fetch_time = time.time() - start_time
        st.session_state['data_fetch_time'] = fetch_time
        
        return df
    
    except Exception as e:
        st.write(f"An error occurred: {e}")
        return pd.DataFrame()
    
    
def get_color(field, value):
    """
    Determine the color for the bar based on the value.
    
    Args:
    - value (float): The value to determine the color for.
    
    Returns:
    - str: The color based on the value range.
    """
    if field == 'CHT':
        if value >= st.session_state.get('minCHT') and value <= st.session_state.get('nominalCHT'):
            return 'blue'
        elif value >= st.session_state.get('nominalCHT') and value <= st.session_state.get('highCHT'):
            return 'green'
        elif value >= st.session_state.get('highCHT') and value <= (0.9 * st.session_state.get('maxCHT')):
            return 'orange'
        else:
            return 'red'

    if field == 'EGT':
        if value >= st.session_state.get('minEGT') and value <= st.session_state.get('nominalEGT'):
            return 'blue'
        elif value >= st.session_state.get('nominalEGT') and value <= st.session_state.get('highEGT'):
            return 'green'
        elif value >= st.session_state.get('highEGT') and value <= (0.9 * st.session_state.get('maxEGT')):
            return 'orange'
        else:
            return 'red'


def display_metrics(df):
    """
    Display metrics as text with labels and units.
    
    Args:
    - df (pd.DataFrame): DataFrame containing the data.
    """
    
    if not df.empty:
        metrics = {
            "Front Oil Pressure": ("oil_pressure", "PSI", st.session_state.get('minFrontOilPressure'), st.session_state.get('nominalFrontOilPressure'), st.session_state.get('highFrontOilPressure'), st.session_state.get('maxOilFrontOilPressure')),
            "Back Oil Pressure": ("oil_pressure_2", "PSI", st.session_state.get('minRearOilPressure'), st.session_state.get('nominalRearOilPressure'), st.session_state.get('highRearOilPressure'), st.session_state.get('maxOilRearOilPressure')),
            "OAT": ("oat", "°F"),
            "IAT Carb": ("iat_carb", "°F", st.session_state.get('minIAT'), st.session_state.get('nominalIAT'), st.session_state.get('highIAT'), st.session_state.get('maxIAT')),
            "Oil Temp": ("oil_temperature", "°F", st.session_state.get('minOilTemp'), st.session_state.get('nominalOilTemp'), st.session_state.get('highOilTemp'), st.session_state.get('maxOilTemp')),
            "Fuel Flow": ("fuel_flow", "GPH", st.session_state.get('minFuelFlow'), st.session_state.get('nominalFuelFlow'), st.session_state.get('highFuelFlow'), st.session_state.get('maxFuelFlow')),
            "Metered Fuel Pressure": ("fuel_pressure", "PSI", st.session_state.get('minMeteredFuelPressure'), st.session_state.get('nominalMeteredFuelPressure'), st.session_state.get('highMeteredFuelPressure'), st.session_state.get('maxMeteredFuelPressure')),
            "Unmetered Fuel Pressure": ("fuel_pressure", "PSI", st.session_state.get('minUnmeteredFuelPressure'), st.session_state.get('nominalUnmeteredFuelPressure'), st.session_state.get('highUnmeteredFuelPressure'), st.session_state.get('maxUnmeteredFuelPressure')),
            "Volts": ("volts", "V", st.session_state.get('minFuelFlow'), st.session_state.get('nominalFuelFlow'), st.session_state.get('highFuelFlow'), st.session_state.get('maxFuelFlow')),
            "CDT": ("cdt", "°F", st.session_state.get('minCDT')), 
            "Mag Drop": ("mag_drop", "", st.session_state.get('minMagDrop'), st.session_state.get('nominalMagDrop'), st.session_state.get('highMagDrop'), st.session_state.get('maxMagDrop'))
        }
        
        # Create rows of columns
        row1 = st.columns(5)
        row2 = st.columns(5)
        row3 = st.columns(5)

        # Combine all columns into a single list
        all_columns = row1 + row2 + row3

        # CSS for custom styling
        css = """
        <style>
        .metric-tile {
            border-radius: 8px;  /* Rounded corners */
            padding: 10px;  /* Padding inside the tile */
            margin: 5px;  /* Margin around the tile */
            text-align: center;  /* Center-align text */
        }
        .metric-label {
            font-size: 14px;  /* Adjust label font size */
            text-align: left;  /* Center-align text */
            color: #ffffff;  /* Darker color for label */
        }
        .metric-value {
            font-size: 18px;  /* Adjust value font size */
            color: #ffffff;  /* Black color for value */
        .metric-delta {
            font-size: 14px;  /* Adjust delta font size */
            color: #007bff;  /* Blue color for delta */
        }
        </style>
        """

        st.markdown(css, unsafe_allow_html=True)

        # Iterate over metrics and place them in the grid
        for i, (label, (field, unit, minThreshold, nominalThreshold, highThreshold, maxThreshold)) in enumerate(metrics.items()):
            col_index = i % len(all_columns)  # Determine the column index
            with all_columns[col_index].container(height=120):  # Create a container for each tile
                if field in df.columns and len(df) > 1:
                    # Get the latest and previous values
                    latest_value = float(df[field].iloc[0])  # Latest value (last row)
                    st.write(latest_value)
                    previous_value = float(df[field].iloc[-1])  # Previous value (second last row)
                    delta = latest_value - previous_value  # Calculate the delta
                    
                    
                    # Determine text color based on thresholds
                    if latest_value > maxThreshold:
                        text_color = "red"
                        st.markdown(f"<div class='flash'><span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span></div>", unsafe_allow_html=True)
                    elif highThreshold < latest_value <= 0.9 * maxThreshold:
                        text_color = "orange"
                        st.markdown(f"<span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span>", unsafe_allow_html=True)
                    elif nominalThreshold < latest_value <= highThreshold:
                        text_color = "green"
                        st.markdown(f"<span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span>", unsafe_allow_html=True)
                    elif minThreshold < latest_value <= nominalThreshold:
                        text_color = "blue"
                        st.markdown(f"<span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span>", unsafe_allow_html=True)
                    else:
                        text_color = "#ffffff"
                        st.markdown(f"<span style='color:{text_color}; font-size:28px;'>{latest_value:.1f} {unit}</span>", unsafe_allow_html=True)
                    
                    # Display delta value
                    if delta > 0:
                        delta_display = f"<span style='color:green; font-size:18px;'>↑ {delta:+.1f}</span>"
                    elif delta < 0:
                        delta_display = f"<span style='color:red; font-size:18px;'>↓ {delta:+.1f}</span>"
                    else:
                        delta_display = ""

                    # Display each metric using st.markdown to include HTML
                    st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
                    st.markdown(f"<div style='text-align:left;'>{display_text}</div>", unsafe_allow_html=True)
                    st.markdown(f"<div style='text-align:left;'>{delta_display}</div>", unsafe_allow_html=True)
                
                else:
                    display_text = "<span style='color:#ffffff; font-size:24px;'>Data not available</span>"
                    delta_display = "<span style='color:#ffffff;'>N/A</span>"

                    st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
                    st.markdown(f"<div style='text-align:left;'>{display_text}</div>", unsafe_allow_html=True)
                    st.markdown(f"<div style='text-align:left;'>{delta_display}</div>", unsafe_allow_html=True)


def plot_egt_cht(df):
    """
    Plot EGT and CHT values in a bar graph.
    
    Args:
    - df (pd.DataFrame): DataFrame containing the data.
    """
    temps_start_time = time.time()  # Start time for performance measurement
    
    fig, ax = plt.subplots(figsize=(12, 6))
    
    # Identify EGT and CHT columns
    egt_columns = sorted([col for col in df.columns if col.startswith('egt_')])
    cht_columns = sorted([col for col in df.columns if col.startswith('cht_')])

    # # Create traces for EGT and CHT
    # egt_values = [df[col].iloc[0] for col in egt_columns]  # Get EGT values
    # cht_values = [df[col].iloc[0] for col in cht_columns]  # Get CHT values

    # Interleave EGT and CHT labels and values
    labels = []
    values = []
    text = []
    colors = []
    for egt_label, cht_label in zip(egt_columns, cht_columns):
        labels.append(egt_label)
        values.append(df[egt_label].iloc[0])
        egt_value = float(df[egt_label].iloc[0])
        text.append(f'{egt_value} °F')
        colors.append(get_color('EGT', egt_value))  # Default color for EGT

        labels.append(cht_label)
        values.append(df[cht_label].iloc[0])
        cht_value = float(df[cht_label].iloc[0])
        text.append(f'{cht_value} °F')
        colors.append(get_color('CHT', cht_value))  # Default color for EGT
        # labels.append(cht_label)
        # values.append(df[cht_label].iloc[0])
        # text.append(f'{df[cht_label].iloc[0]:.1f} °F')
        # colors.append(get_bar_color(cht_values))  # Color based on CHT value

    # Create traces
    traces = [
        go.Bar(
            x=labels,
            y=values,
            name='Temperature',
            marker_color=colors,  # Color based on label
            text=text,  # Show values on hover
            textposition='auto'  # Automatically position text on the bars
        )
    ]
    
    # Create figure
    fig = go.Figure(data=traces)
    
    # Update layout
    fig.update_layout(
        width=800,  # Set the desired width
        height=350,  # Set the desired height
        margin=dict(l=20, r=20, t=40, b=20),  # Adjust margins if necessary
        barmode='group',  # Group bars together
        xaxis_tickangle=-45  # Rotate x-axis labels for better readability
    )
    
    temps_plot_time = time.time() - temps_start_time  # Measure time taken to render plot
    st.session_state['temps_plot_render_time'] = temps_plot_time  # Store plot render time in session state
    
    return fig
    

def animate_rpm(df):
    """
    Create a gauge chart for RPM values.
    
    Args:
    - df (pd.DataFrame): DataFrame containing the data.
    """
    rpm_start_time = time.time()  # Start time for performance measurement
    
    rpm = df['rpm'].iloc[0] if not df.empty else 0

    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=rpm,
        domain={'x': [0,0.75], 'y': [0,1]},
        title={'text': "RPM"},
        gauge={
            'axis': {'range': [st.session_state.get('minRPM'), int(st.session_state.get('maxRPM'))+500]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, st.session_state.get('maxRPM')/3], 'color': "lightblue"},  
                {'range': [st.session_state.get('maxRPM')/3, 2*(st.session_state.get('maxRPM')/3)], 'color': "skyblue"}, 
                {'range': [2*(st.session_state.get('maxRPM')/3), st.session_state.get('maxRPM')], 'color': "deepskyblue"},
                {'range': [st.session_state.get('maxRPM'), int(st.session_state.get('maxRPM'))+500], 'color': "dodgerblue"}
                ],
            'threshold' : {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': st.session_state.get('maxRPM')}
        }
    ))

    # Update layout to adjust size
    fig.update_layout(
        width=300,  # Set the desired width
        height=120,  # Set the desired height
        margin=dict(l=20, r=20, t=45, b=20)  # Adjust margins if necessary
    )

    rpm_plot_time = time.time() - rpm_start_time  # Measure time taken to render plot
    st.session_state['rpm_plot_render_time'] = rpm_plot_time  # Store plot render time in session state

    return fig


def animate_manifold_pressure(df):
    """
    Create a gauge chart for Manifold Pressure values.
    
    Args:
    - df (pd.DataFrame): DataFrame containing the data.
    """
    mp_start_time = time.time()  # Start time for performance measurement
    manifold_pressure = df['manifold_pressure'].iloc[0] if not df.empty else 0

    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=manifold_pressure,
        number={'suffix': " hg"},  # Add the units here
        domain={'x': [0,0.75], 'y': [0,1]},
        title={'text': "MP"},
        gauge={
            'axis': {'range': [st.session_state.get('minMP'), int(st.session_state.get('maxMP'))+5]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, st.session_state.get('maxMP')/3], 'color': "lightblue"},
                {'range': [st.session_state.get('maxMP')/3, 2*(st.session_state.get('maxMP')/3)], 'color': "skyblue"},
                {'range': [2*(st.session_state.get('maxMP')/3), st.session_state.get('maxMP')], 'color': "deepskyblue"},
                {'range': [st.session_state.get('maxMP'), int(st.session_state.get('maxMP'))+5], 'color': "dodgerblue"}
            ],
            'threshold' : {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': st.session_state.get('maxMP')}
        }
    ))

    # Update layout to adjust size
    fig.update_layout(
        width=300,  # Set the desired width
        height=120,  # Set the desired height
        margin=dict(l=20, r=20, t=45, b=20)  # Adjust margins if necessary
    )

    mp_plot_time = time.time() - mp_start_time  # Measure time taken to render plot
    st.session_state['mp_plot_render_time'] = mp_plot_time  # Store plot render time in session state

    return fig

def animate_hp(df):
    """
    Create a gauge chart for horsepower values.
    
    Args:
    - df (pd.DataFrame): DataFrame containing the data.
    """
    hp_start_time = time.time()  # Start time for performance measurement
    horsepower = df['hp'].iloc[0] if not df.empty else 0

    # maxHPRange = st.

    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=horsepower,
        domain={'x': [0,0.75], 'y': [0,1]},
        title={'text': "HP"},
        gauge={
            'axis': {'range': [st.session_state.get('minHP'), int(st.session_state.get('maxHP'))+30]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, st.session_state.get('maxHP')/3], 'color': "lightblue"},
                {'range': [st.session_state.get('maxHP')/3, 2*(st.session_state.get('maxHP')/3)], 'color': "skyblue"},
                {'range': [2*(st.session_state.get('maxHP')/3), st.session_state.get('maxHP')], 'color': "deepskyblue"},
                {'range': [st.session_state.get('maxHP'), int(st.session_state.get('maxHP'))+30], 'color': "dodgerblue"}
            ],
            'threshold' : {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': st.session_state.get('maxHP')}
        }
    ))

    # Update layout to adjust size
    fig.update_layout(
        width=300,  # Set the desired width
        height=120,  # Set the desired height
        margin=dict(l=20, r=20, t=45, b=20)  # Adjust margins if necessary
    )

    mp_plot_time = time.time() - hp_start_time  # Measure time taken to render plot
    st.session_state['hp_plot_render_time'] = mp_plot_time  # Store plot render time in session state

    return fig


def plot_performance_metrics():
    """
    Create gauge charts for Data Fetch Time and Plot Render Time.
    """

    global refresh_time
    data_fetch_time = st.session_state.get('data_fetch_time', 0)
    temps_plot_render_time = st.session_state.get('temps_plot_render_time', 0)
    rpm_plot_render_time = st.session_state.get('rpm_plot_render_time', 0)
    mp_plot_render_time = st.session_state.get('mp_plot_render_time', 0)
    hp_plot_render_time = st.session_state.get('hp_plot_render_time', 0)
    total_plot_time = temps_plot_render_time + rpm_plot_render_time + mp_plot_render_time + refresh_time + hp_plot_render_time

    # Create figure with two subplots
    #fig = go.Figure()

    # Data Fetch & Plot Render Time Gauge
    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        domain={'x': [0,0.75], 'y': [0,1]},
        value=data_fetch_time + total_plot_time,
        number={'suffix': " s"},  # Add the units here
        title={'text': "Lag"},
        gauge={
            'axis': {'range': [None, 1.5]},  # Adjust the range according to your needs
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, 0.4], 'color': "lightblue"},
                {'range': [0.4, 0.8], 'color': "skyblue"},
                {'range': [0.8, 1.5], 'color': "deepskyblue"},
                # {'range': [1.4, 2.5], 'color': "dodgerblue"}
            ]
        }
    ))

    # Update layout to adjust size
    fig.update_layout(
        width=300,  # Set the desired width
        height=120,  # Set the desired height
        margin=dict(l=20, r=20, t=45, b=20)  # Adjust margins if necessary
    )
    

    # Update layout to organize gauges in a single row
    #fig.update_layout(
    #    grid={'rows': 1, 'columns': 2},
    #    title_text="Performance Metrics",
    #    title_x=0.5
    #)

    return fig


# Function to convert DataFrame to CSV
def convert_df_to_csv(test_sheet_df):
    return test_sheet_df.to_csv(index=False).encode('utf-8')


# Button handlers
# @st.fragment
def start_button_handler(df):
    now = datetime.now()
    test_date = now.strftime("%Y-%m-%d")
    start_time = now.strftime("%H:%M:%S")
    outside_temp = int(df['oat'].iloc[0])

    # Ensure the columns 'Unnamed: 8' and 'Unnamed: 15' exist
    for col in ['Unnamed: 8','Unnamed: 15']:
        if col not in st.session_state.test_sheet.columns:
            st.session_state.test_sheet[col] = None

    # Modify specific cells with current date and time
    st.session_state.test_sheet.loc[2, 'Unnamed: 8'] = test_date
    st.session_state.test_sheet.loc[3, 'Unnamed: 8'] = outside_temp
    st.session_state.test_sheet.loc[2, 'Unnamed: 15'] = start_time

    # Check if cells are populated
    start_cells = [(2, 'Unnamed: 8'), (3, 'Unnamed: 8'), (2, 'Unnamed: 15')]

    for row, col in start_cells:
        cell_value = st.session_state.test_sheet.loc[row, col]
        is_empty = pd.isna(cell_value) or cell_value == ''
        if is_empty:
            st.session_state.start_btn_toast = "Oops!! Sorry something went wrong, try again"
            # time.sleep(1)
        else:
            st.session_state.start_btn_toast = "Date, start time & OAT fields populated"
            # time.sleep(1)

    st.toast(st.session_state.start_btn_toast)
    st.session_state.start_time = start_time

# @st.fragment
def snap_button_handler(df):
    current_index = get_snapshot_index()
    
    # This if statement will skip row 13 when populating fields in descending RPM
    
    if current_index == 14:
        st.session_state.snapshot_index -= 1
    # This if statement will change the row being populated
    if current_index != 8:
        st.session_state.snapshot_index -= 1
    # When row is reached, it will be populated and the index moved to row 7 to avoid rewriting row 8    
    else: 
        st.session_state.snapshot_index -= 1
        # remind user to input final RPM
        st.toast("Confirm Idle RPM")
    # This if statement set's the boundary for the snap button, doing nothing once row 8 is populated and remaining at row 7
    if current_index < 8:
        st.session_state.snapshot_index = 7
        st.toast(f"All RPM fields are populated. Current index: {current_index}")
        return

    manifold_pressure = int(df['manifold_pressure'].iloc[0])
    oil_pressure = int(df['oil_pressure'].iloc[0])
    oil_temp = int(df['oil_temperature'].iloc[0])
    fuel_pressure = int(df['fuel_pressure'].iloc[0])
    egt_1 = int(df['egt_1'].iloc[0])
    egt_2 = int(df['egt_2'].iloc[0])
    egt_3 = int(df['egt_3'].iloc[0])
    egt_4 = int(df['egt_4'].iloc[0])
    cht_1 = int(df['cht_1'].iloc[0])
    cht_2 = int(df['cht_2'].iloc[0])
    cht_3 = int(df['cht_3'].iloc[0])
    cht_4 = int(df['cht_4'].iloc[0])
    row_rpm = st.session_state.test_sheet.loc[current_index, 'Unnamed: 0']

    # Ensure the columns exist
    for col in ['Unnamed: 1', 'Unnamed: 2','Unnamed: 4','Unnamed: 6','Unnamed: 8'
               ,'Unnamed: 9','Unnamed: 10','Unnamed: 11', 'Unnamed: 14','Unnamed: 15'
               ,'Unnamed: 16','Unnamed: 17']:
        if col not in st.session_state.test_sheet.columns:
            st.session_state.test_sheet[col] = None

    # Modify specific cells with current date and time
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 1'] = manifold_pressure
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 3'] = oil_temp
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 4'] = oil_pressure
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 6'] = fuel_pressure
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 8'] = egt_1
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 9'] = egt_2
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 10'] = egt_3
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 11'] = egt_4
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 14'] = cht_1
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 15'] = cht_2
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 16'] = cht_3
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 17'] = cht_4

    # Check if cells are populated
    snap_cells = [(current_index, 'Unnamed: 1'), (current_index, 'Unnamed: 3'), (current_index, 'Unnamed: 4')
                  ,(current_index, 'Unnamed: 6'), (current_index, 'Unnamed: 6'), (current_index, 'Unnamed: 8')
                  ,(current_index, 'Unnamed: 9'), (current_index, 'Unnamed: 10'), (current_index, 'Unnamed: 11')
                  ,(current_index, 'Unnamed: 14'), (current_index, 'Unnamed: 15'), (current_index, 'Unnamed: 16')
                  ,(current_index, 'Unnamed: 17')]

    for row, col in snap_cells:
        cell_value = st.session_state.test_sheet.loc[row, col]
        is_empty = pd.isna(cell_value) or cell_value == ''
        if is_empty:
            st.session_state.snap_btn_toast = "Oops!! Sorry something went wrong, try again"
            
        else:
            st.session_state.snap_btn_toast = f"Snapshot taken. RPM: {row_rpm}."
            
    st.toast(st.session_state.snap_btn_toast)
    
# @st.fragment
def undo_button_handler(df):
    current_index = get_snapshot_index() + 1
    st.session_state.snapshot_index = current_index
    # If the snap boundary has been reached i.e the last row has been populated; undo the last row
    if current_index == 7:
        st.session_state.snapshot_index = 8
        return

    if current_index == 12:
            st.session_state.snapshot_index +=1 #current_index

    if current_index == 13:
        # current_index = current_index - 1 
        st.session_state.snapshot_index -=2
        return

    if current_index > 17:
        st.session_state.snapshot_index = 17
        st.toast(f"Limits exceeded. Current index: {current_index}")
        
        return
    
    # Read the rpm at current index from test sheet dataframe
    row_rpm = st.session_state.test_sheet.loc[current_index, 'Unnamed: 0']
    
    # Ensure the columns exist
    for col in ['Unnamed: 1', 'Unnamed: 2','Unnamed: 4','Unnamed: 6','Unnamed: 8'
               ,'Unnamed: 9','Unnamed: 10','Unnamed: 11''Unnamed: 14','Unnamed: 15','Unnamed: 16','Unnamed: 17'
               ,'Unnamed: 18','Unnamed: 19']:
        if col not in st.session_state.test_sheet.columns:
            st.session_state.test_sheet[col] = None

    # Modify specific cells with current date and time
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 1'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 3'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 4'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 6'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 8'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 9'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 10'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 11'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 14'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 15'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 16'] = ''
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 17'] = ''
    

    # Check if cells are populated
    snap_cells = [(current_index, 'Unnamed: 1'), (current_index, 'Unnamed: 3'), (current_index, 'Unnamed: 4')
                  , (current_index, 'Unnamed: 6'), (current_index, 'Unnamed: 6'), (current_index, 'Unnamed: 8')
                  , (current_index, 'Unnamed: 9'), (current_index, 'Unnamed: 10'), (current_index, 'Unnamed: 11')
                  , (current_index, 'Unnamed: 14'), (current_index, 'Unnamed: 15'), (current_index, 'Unnamed: 16')
                  , (current_index, 'Unnamed: 17')]

    for row, col in snap_cells:
        cell_value = st.session_state.test_sheet.loc[row, col]
        is_empty = pd.isna(cell_value) or cell_value == ''
        if is_empty:
            st.session_state.snap_btn_toast = f"RPM: {row_rpm} cleared"
            
        else:
            st.session_state.snap_btn_toast = "Oops!! Sorry something went wrong, try again"
            
    st.toast(st.session_state.snap_btn_toast)


# @st.fragment
def end_button_handler(df):
    now = datetime.now()
    end_time = now.strftime("%H:%M:%S")
 
    # Ensure the columns 'Unnamed: 15' exist
    if 'Unnamed: 15' not in st.session_state.test_sheet.columns:
        st.session_state.test_sheet['Unnamed: 15'] = None

    # Calculate the total time of the test
    start_time_obj = datetime.strptime(st.session_state.start_time, '%H:%M:%S')
    end_time_obj = datetime.strptime(end_time, '%H:%M:%S')
    total_time = end_time_obj - start_time_obj

    # Modify specific cells with current date and time
    st.session_state.test_sheet.loc[3, 'Unnamed: 15'] = end_time
    st.session_state.test_sheet.loc[4, 'Unnamed: 15'] = str(total_time)
    # st.toast(f"End time added!!!!. End time: {end_time}")

    # Check if cells are populated
    end_cells = [(3, 'Unnamed: 15'), (4, 'Unnamed: 15')]

    for row, col in end_cells:
        cell_value = st.session_state.test_sheet.loc[row, col]
        is_empty = pd.isna(cell_value) or cell_value == ''
        if is_empty:
            st.session_state.end_btn_toast = "Oops!! Sorry something went wrong, try again"
            # time.sleep(1)
        else:
            st.session_state.end_btn_toast = "Finish time & total time fields populated"
            # time.sleep(1)

    st.toast(st.session_state.end_btn_toast)

# Define a function to display the grid and capture changes
def display_ag_grid(df):
    # Build grid options
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(editable=True)
    grid_options = gb.build()

    # Display grid and return data based on DataReturnMode.AS_INPUT
    grid_response = AgGrid(
        df,
        gridOptions=grid_options,
        data_return_mode=DataReturnMode.AS_INPUT,
        update_mode=GridUpdateMode.VALUE_CHANGED,
        fit_columns_on_grid_load=True
    )

    # Extract the updated DataFrame
    # return pd.DataFrame(grid_response['data'])
    return grid_response

def copy_to_excel(df, template_path, thresholds):       #"""What does df and thresholds do??"""
   wb = load_workbook(template_path)
   ws = wb.active

#    edited_cells = st.session_state.test_sheet.loc[2, 'Unnamed: 8']
#    ws['I4'] = edited_cells
   row_offset = 1  # Offset by 1 to account for the extra row at the top of the Excel sheet

   # Define fill styles for within and exceeding threshold
   green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light green
   red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")    # Light red

   for r_idx, row in enumerate(dataframe_to_rows(st.session_state.test_sheet, index=False, header=False), start=1):
        actual_row = r_idx + row_offset  # Adjusted row considering the offset
        
        # Skip row 15
        if actual_row == 15:
            continue

        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=actual_row, column=c_idx)

            # Check if the cell is part of a merged cell range
            if cell.coordinate in ws.merged_cells:
                # Get the top-left cell of the merged range
                for merged_range in ws.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                    if min_row <= actual_row <= max_row and min_col <= c_idx <= max_col:
                        # Write to the top-left cell of the merged range
                        top_left_cell = ws.cell(row=min_row, column=min_col, value=value)
                        if 10 <= min_row <= 19 and isinstance(value, (int, float)):
                            threshold = thresholds.get(c_idx)
                            if threshold is not None:
                                if value <= threshold:
                                    top_left_cell.fill = green_fill
                                else:
                                    top_left_cell.fill = red_fill
                        break
            else:
                cell.value = value
                # Apply formatting if within the specified row range and if the value is numeric
                if 10 <= actual_row <= 19 and isinstance(value, (int, float)):
                    threshold = thresholds.get(c_idx)
                    if threshold is not None:
                        if value <= threshold:
                            cell.fill = green_fill
                        else:
                            cell.fill = red_fill

   buffer = BytesIO()
   wb.save(buffer)
   buffer.seek(0)

   return buffer
   

def main():
    """
    Main function to run the Streamlit app.
    """
    # # Ensure the user has selected an engine
    # if 'selected_engine' not in st.session_state: #or st.session_state.selected_engine is None:
    #     st.write("No engine selected. Please go back and select an engine.")
    #     st.query_params["page"] = "select_engine"  # Redirect back to engine selection
    # else:
    #     st.title(f"Dashboard for {st.session_state.selected_engine}")
    #     # Display the dashboard content for the selected engine
    #     st.write(f"Showing data for {st.session_state.selected_engine}")

    df = get_dataframe()

    st.markdown("""
    <style>
    
           /* Remove blank space at top and bottom */ 
           .block-container {
               padding-top: 5px;
               padding-bottom: 15px;
               padding-left: 25px;
               padding-right: 25px; 
            }
           
           
    </style>
    """, unsafe_allow_html=True)
    # Using session state to track components
    if 'components_created' not in st.session_state:
        st.session_state['components_created'] = False

    if not st.session_state['components_created']:

        col1, col2, col3, col4, col5 = st.columns([1, 1, 1.5, 1, 1], gap="small", vertical_alignment='center')  # Create three columns side by side
        with col1:
            with st.container(height=130):
                cola, colb, colc = st.columns([1.5, 6, 0.2])
                with cola:
                    st.write("")
                with colb:
                    performance_fig = plot_performance_metrics()
                    st.plotly_chart(performance_fig, use_container_width=False)
                with colc:
                    st.write("")
                

        with col2:
            with st.container(height=130):
                cola, colb, colc = st.columns([1.5, 6, 0.2])
                with cola:
                    st.write("")
                with colb:
                    rpm_fig = animate_rpm(df)
                    st.plotly_chart(rpm_fig, use_container_width=False)
                with colc:
                    st.write("")

        with col3:
            with st.container(height=130):
                cola, colb, colc, cold, cole, colf = st.columns([0.01, 1.3, 1.3, 1.3, 1.3, 0.01], vertical_alignment='top')
                with colb:
                    st.write("")
                    st.write("")
                    start_btn = st.button("START")
                with colc:
                    st.write("")
                    st.write("")
                    snap_btn = st.button("SNAP")
                with cold:
                    st.write("")
                    st.write("")
                    undo_btn = st.button("UNDO")
                with cole:
                    st.write("")
                    st.write("")
                    end_btn = st.button("END")

        with col4:
            with st.container(height=130):
                cola, colb, colc = st.columns([1.5, 6, 0.2])
                with cola:
                    st.write("")
                with colb:
                    manifold_pressure_fig = animate_manifold_pressure(df)
                    st.plotly_chart(manifold_pressure_fig, use_container_width=False)
                with colc:
                    st.write("")

        with col5:
            with st.container(height=130):
                cola, colb, colc = st.columns([1.5, 6, 0.2])
                with cola:
                    st.write("")
                with colb:
                    horsepower_fig = animate_hp(df)
                    st.plotly_chart(horsepower_fig, use_container_width=False)
                with colc:
                    st.write("")

        with st.container(height=360):
            col6, col7 = st.columns(2, vertical_alignment='center')
            with col6:
                temps_bar_plot = plot_egt_cht(df)
                st.plotly_chart(temps_bar_plot, use_container_width=False)

            with col7:
                display_metrics(df)

        if start_btn:
            start_button_handler(df)
            time.sleep(refresh_time)
        if snap_btn:
            snap_button_handler(df)
            time.sleep(refresh_time)
        if undo_btn:
            undo_button_handler(df)
            time.sleep(refresh_time)
        if end_btn:
            end_button_handler(df) 
            time.sleep(refresh_time)


        # Display the engine test sheet dataframe using streamlit aggrid
        # And update session state
        grid_response = display_ag_grid(st.session_state.test_sheet)
        if not grid_response['data'].empty:
            updated_data = pd.DataFrame(grid_response['data'])
            for i in range(len(updated_data)):
                for col in updated_data.columns:
                    st.session_state.test_sheet.iloc[i, st.session_state.test_sheet.columns.get_loc(col)] = updated_data.iloc[i][col]

        
        # Generate CSV data
        # csv_data = convert_df_to_csv(st.session_state.test_sheet)
        # csv_data = copy_to_excel()
        # wb.save(csv_data)
        # csv_data.seek(0)
        excel = r"C:\Users\theca\Documents\Victor\Aerotec\MVP 50\Code\AerotecTestCell\Engine Test Sheet Template.xlsx"


        col8, col9, col10 = st.columns([1.2, 1.2, 1], vertical_alignment='bottom')
        with col8:
            pass
        with col9:

            # Prompt the user for a file name
            file_name_input = st.text_input("Enter the desired file name (without extension):")
            output_path = f"{file_name_input}.xlsx"

            colf, colg, colh = st.columns([1.2, 1.2, 1], vertical_alignment='bottom')
            with colf:
                pass
            with colg:
                # Add a download button for the CSV file
                # if st.button("Download as XLSX"):
                buffer = copy_to_excel(st.session_state.test_sheet, excel, thresholds)
                
                st.download_button(
                    label="Download file",
                    data=buffer,
                    file_name=output_path,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                
            with colh:
                pass
        with col10:
            pass

if __name__ == "__main__":
    main()
    subprocess.Popen(["python", r"C:\Users\theca\Documents\Victor\Aerotec\MVP 50\Code\AerotecTestCell\Parse_data_4_Cyl.py"])

