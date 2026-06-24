import streamlit as st
import requests
import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st
from pymongo import MongoClient
import datetime
from matplotlib.animation import FuncAnimation
import time
# import datetime
from datetime import datetime, timedelta, date
from streamlit_autorefresh import st_autorefresh
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np
from st_aggrid import AgGrid
from st_aggrid import GridOptionsBuilder, GridUpdateMode, DataReturnMode, JsCode
import asyncio
import motor.motor_asyncio
from motor.motor_asyncio import AsyncIOMotorClient
import subprocess
import os
import signal
import random
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import range_boundaries
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.styles import Alignment, Font
import commentjson as json
from io import BytesIO
import pytz
import datetime
import scipy
from scipy.interpolate import interp1d
import sys
from collections import OrderedDict


# Page Configuration
st.set_page_config(layout="wide", 
                   page_title="Engine Dashboard", 
                   page_icon=st.session_state.get("page_icon", ":small_airplane:"), 
                   initial_sidebar_state="collapsed")

# Add Aerotec Logo
# st.logo(st.session_state.get("logo", ":small_airplane:"))

# Autorefresh Setup
refresh_time = 1
st_autorefresh(refresh_time*1000, key="data_refresh")

# Pages File Paths
main_page_path = r"C:\Users\KieranCalder\Code\AerotecTestCell\st_main.py"
select_engine_path = r"C:\Users\KieranCalder\Code\AerotecTestCell\pages\st_select_engine.py"

# JSON File Path
engine_json_path = r"C:\Users\KieranCalder\Code\AerotecTestCell\Engine Configuration.json"
adjustable_json_path = r"C:\Users\KieranCalder\Code\AerotecTestCell\3 Blade Adjustable Propeller Logs.json"

# Test Sheet Path
test_sheet_path = r"C:\Users\KieranCalder\Code\AerotecTestCell\Dashboard Test Sheet Template.xlsx"

save_xlxs_path = r"C:\Users\KieranCalder\Code\AerotecTestCell\Test Sheets"
xlxs_copy_path = r"C:\Users\KieranCalder\Code\AerotecTestCell\Dashboard Excel Copy Template.xlsx"

# Parse Data File Path
parse_data_path = r"C:\Users\KieranCalder\Code\AerotecTestCell\Parse_data.py"

# Define Global Variables
global engine_type, dataOnly, cylinderCount, turboStatus, time_step
dataOnly = "Data Only"

# Ensure the user has selected an engine
if 'selected_engine' not in st.session_state: #or st.session_state.selected_engine is None:
    st.warning("No engine selected. Redirecting to select engine page...")
    time.sleep(1)
    st.switch_page(r"C:\Users\KieranCalder\Code\AerotecTestCell\pages\st_select_engine.py")  # Redirect back to engine selection

if 'page_icon' not in st.session_state: #or st.session_state.selected_engine is None:
    st.warning("Initialization incomplete. Redirecting to main page...")
    time.sleep(1)
    st.switch_page(r"C:\Users\KieranCalder\Code\AerotecTestCell\st_main.py")  # Redirect back to engine selection

if 'logo' not in st.session_state: 
    st.warning("Initialization incomplete. Redirecting to main page...")
    time.sleep(1)
    st.switch_page(r"C:\Users\KieranCalder\Code\AerotecTestCell\st_main.py")  # Redirect back to engine selection

# Synchronize UI button states with reality on script execution
if 'subprocess' in st.session_state:
    if st.session_state.subprocess.poll() is None:
        st.session_state.process_running = True
    else:
        # Clear if backgorund script closed itself
        st.session_state.process_running = False
        del st.session_state.subprocess
else:
    if 'process_running' not in st.session_state:
        st.session_state.process_running = False

# Fetch Dashboard Variables from session state
customer_name = st.session_state.customer_name
sn_number = st.session_state.sn_number
work_order = st.session_state.work_order
test_nature = st.session_state.test_nature
engine_type = st.session_state.selected_engine
num_blades = st.session_state.selected_blades
prop_type = st.session_state.selected_prop
test_sheet_selected_rpm = st.session_state.selected_rpm
cylinderCount = st.session_state.get('cylinders')
turboStatus = st.session_state.get('turbo')
rearOilPressureStatus = st.session_state.get('rearOilPressure')
frontOilPressureStatus = st.session_state.get('frontOilPressure')
meteredFuelPressureStatus = st.session_state.get('meteredFuelPressure')
unmeteredFuelPressureStatus = st.session_state.get('unmeteredFuelPressure')

# Setup MongoDB Client and URI
uri = "mongodb://localhost:27017/"
client = motor.motor_asyncio.AsyncIOMotorClient(uri)
# Initialize db_name in session state if not already done
if 'db_name' not in st.session_state:
    st.session_state.db_name = None

db_str = str(st.session_state.db_name)
db = client[db_str]

# Initialize collection_name in session state if not already done:
if 'collection_name' not in st.session_state:
    st.session_state.collection_name = None
    
collection_str = str(st.session_state.collection_name)
collection = db[collection_str]

def load_config():
    """
    Load JSON configuration file.
    """
    with open(engine_json_path, "r") as f:
        return json.load(f, object_pairs_hook=OrderedDict) # Use OrderedDict to maintain order

engine_config = load_config()

def load_adjustable_config():
    if os.path.exists(adjustable_json_path):
        with open(adjustable_json_path, "r") as f:
            return json.load(f)
        
def save_adjustable_config(adjustable_data):
    """Save updated ajustable configuration to the JSON file"""
    with open(adjustable_json_path, "w") as f:
        json.dump(adjustable_data, f, indent=4)

# @st.cache_data
def test_sheet_df(path):
    """
    Loads the xlxs test sheet and ensures session state is initialized
    """
    # Load xlxs file
    test_sheet_df = pd.read_excel(path)
    # Convert to object type
    test_sheet_df = test_sheet_df.astype(object)
    # Replace Nan with None so they show up empty in the editor
    test_sheet_df = test_sheet_df.where(pd.notnull(test_sheet_df), "")
    # Drop columns that are entirely NaN
    # test_sheet_df = test_sheet_df.dropna(axis=1, how="all")

    return test_sheet_df

def persist_test_sheet():
    """
    Converts the active test sheet dataframe to records and uploads to MongoDb
    """
    db_name = st.session_state.get('db_name')
    coll_name = st.session_state.get('collection_name')

    if not db_name or not coll_name:
        return # No engine running yet
    
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

        # Local client instance tied strictly to this thread's loop execution
        local_client = motor.motor_asyncio.AsyncIOMotorClient(uri)
        active_collection = local_client[str(db_name)][str(coll_name)]

        records = st.session_state.test_sheet.to_dict(orient='records')

        loop.run_until_complete(
            active_collection.update_one(
                {"_id": "testsheet_snapshot"},
                {"$set": {"records": records}},
                upsert=True
            )
        )
        loop.close()
    except Exception as e:
        st.error(f"Auto-save to MongoDB failed: {e}")

def fetch_stored_test_sheet():
    """
    Attempts to recover a saved test sheet from the active database collection
    """
    db_name = st.session_state.get('db_name')
    coll_name = st.session_state.get('collection_name')

    if not db_name or not coll_name:
        return None
    
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

        local_client = motor.motor_asyncio.AsyncIOMotorClient(uri)
        active_collection = local_client[str(db_name)][str(coll_name)]

        async def get_doc():
            return await active_collection.find_one({"_id": "testsheet_snapshot"})
        
        doc = loop.run_until_complete(get_doc())
        loop.close()

        if doc and "records" in doc:
            return pd.DataFrame(doc["records"])
    except Exception as e:
        st.error(f"Failed to recover testsheet from database: {e}")

    return None

# Initialize session state if not already done
if 'test_sheet' not in st.session_state:
    stored_df = fetch_stored_test_sheet()
    
    if stored_df is not None:
        st.session_state.test_sheet = stored_df
        st.toast("Recovered active testsheet from MongoDB")
    else:
    # Load blank testsheet
        st.session_state.test_sheet = test_sheet_df(test_sheet_path)
        st.toast("Cache Empty. Loading test sheet template...")

def get_snapshot_index():
    if 'snapshot_index' not in st.session_state:
        # DYNAMICALLY find the start row
        row_map = get_row_map(st.session_state.test_sheet)
        rpm_sequence = get_rpm_sequence(test_sheet_selected_rpm)
        start_rpm = rpm_sequence[0] # e.g. "2700" or "2200"
        
        if start_rpm in row_map:
            st.session_state.snapshot_index = row_map[start_rpm]
            st.toast(f"Ready. Starting at {start_rpm} RPM (Row {row_map[start_rpm]})")
        else:
            st.error(f"Could not find start RPM '{start_rpm}' in the Excel sheet labels.")
            st.session_state.snapshot_index = 0

    return st.session_state.snapshot_index

def get_row_map(df):
    """
    Creates a mapping of labels to row indices
    Injects indices for 2700 and 750 based on their neighbors
    """
    mapping = {}

    # Standard mapping for existing labels
    # Assuming Column 0 ('Rev. 5') holds the rpm labels
    for idx, value in df.iloc[:, 0].items():
        # Check if value is not None and not a NaN
        if value is not None and pd.notna(value):
            clean_label = str(value).strip()
            mapping[clean_label] = idx

    # Relative mapping for 750 and 2700 rpm rows
    if "1000" in mapping:
        mapping["750"] = mapping["1000"] - 1
    if "2400" in mapping:
        mapping["2700"] = mapping["2400"] + 1

    return mapping
     
def get_rpm_sequence(selected_rpm):
    """
    Returns the list of RPMs to be populated based on the 
    user's selection in session state
    """
    # Base sequence that always exists
    # Note: 750 is included here as the final idle point
    base_sequence = [2400, 2200, 2000, 1800, 1500, 1300, 1000, 750]

    # If 2700 is selected, prepend it to the start of the sequence
    if int(selected_rpm) == 2700:
        full_sequence = [2700] + base_sequence
    else:
        full_sequence = base_sequence

    return [str(x) for x in full_sequence]

def get_dataframe():
    global time_step
    utc_now = datetime.datetime.now(datetime.UTC)
    ast_tz = pytz.timezone('America/Halifax')
    ast_now = utc_now.replace(tzinfo=pytz.utc).astimezone(ast_tz)
    start_time = time.time()  # Start time for performance measurement

    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        data_list = loop.run_until_complete(fetch_data())
        loop.close()
        
        if not data_list:
            st.write("No data fetched from MongoDB.")
            return pd.DataFrame()
        
        df = pd.DataFrame(data_list)
        
        if df.empty:
            st.write("DataFrame is empty.")
            return df
        
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        
        numeric_fields = [
            'volts', 'manifold_pressure', 'rpm', 'hp', 'fuel_flow', 'upper_deck',
            'metered_fuel_pressure', 'unmetered_fuel_pressure', 'front_oil_p', 'rear_oil_p' 'oil_temperature',
            'amps', 'cdt', 'iat_carb', 'oat', 'tit',
            'egt_1', 'egt_2', 'egt_3', 'egt_4', 'egt_5', 'egt_6',
            'cht_1', 'cht_2', 'cht_3', 'cht_4', 'cht_5', 'cht_6' 
        ]

        for field in numeric_fields:
            if field in df.columns:
                df[field] = pd.to_numeric(df[field], errors='coerce') #+ random.uniform(-5,5)
            else:
                df[field] = None  # Handle missing fields


        # df['oil_temperature'] = ((df['oil_temperature']) -32 ) * (5/9)
        # df['oat'] = st.session_state.beaver_oat
        # df['oat'] = ((df['oat']) -32 ) * (5/9)
        df['oat'] = st.session_state.beaver_oat
        df['metered_fuel_pressure'] = df['upper_deck']
        # df['rpm'] = df['rpm'] + random.uniform(700, 1700)
        # df['iat_carb'] = ((df['iat_carb']) -32 ) * (5/9)
    


        df['elapsed_time'] = (df['timestamp'] - df['timestamp'].max()).dt.total_seconds()
        
        fetch_time = time.time() - start_time
        st.session_state['data_fetch_time'] = fetch_time
        
        return df
    
    except Exception as e:
        st.write(f"An error occurred: {e}")
        return pd.DataFrame()
    
async def fetch_data():
    """
    Queries database and fetches nvp-50 data
    """
    db_name = st.session_state.get('db_name')
    coll_name = st.session_state.get('collection_name')

    if not db_name or not coll_name:
        return []
    
    local_client = motor.motor_asyncio.AsyncIOMotorClient(uri)
    active_collection = local_client[str(db_name)][str(coll_name)]

    num_documents = 5
    pipeline = [
        {"$sort": {"timestamp": -1}},  # Sort by timestamp, latest first
        {"$limit": num_documents},  # Limit the number of documents
        {"$project": {
            "_id": 0,  # Exclude the _id field
            "timestamp": 1,
            "volts": "$fields.volts",
            "manifold_pressure": "$fields.manifold_pressure",
            "rpm": "$fields.rpm",
            "hp": "$fields.hp",
            "fuel_flow": "$fields.fuel_flow",
            "upper_deck": "$fields.upper_deck",
            "metered_fuel_pressure": "$fields.metered_fuel_pressure",
            "unmetered_fuel_pressure": "$fields.unmetered_fuel_pressure",
            "front_oil_p": "$fields.front_oil_p",
            "rear_oil_p": "$fields.rear_oil_p", 
            "oil_temperature": "$fields.oil_temperature",
            "amps": "$fields.amps",
            "cdt": "$fields.cdt",
            "tit": "$fields.tit",
            "iat_carb": "$fields.iat_carb",
            "oat": "$fields.oat",
            "mag_drop":"$fields.mag_drop",
            "egt_1": "$fields.egt_1",
            "egt_2": "$fields.egt_2",
            "egt_3": "$fields.egt_3",
            "egt_4": "$fields.egt_4",
            "egt_5": "$fields.egt_5",
            "egt_6": "$fields.egt_6",
            "cht_1": "$fields.cht_1",
            "cht_2": "$fields.cht_2",
            "cht_3": "$fields.cht_3",
            "cht_4": "$fields.cht_4",
            "cht_5": "$fields.cht_5",
            "cht_6": "$fields.cht_6",
        }}
    ]
    cursor = active_collection.aggregate(pipeline)
    data_list = []
    async for document in cursor:
        data_list.append(document)
    
    return data_list

def get_color(field, value):
    """
    Determine the color for the bar based on the value.
    
    Args:
    - value (float): The value to determine the color for.
    
    Returns:
    - str: The color based on the value range.
    """
    if field == 'CHT':
        if value >= st.session_state.get('minCHT') and value <= st.session_state.get('nominalCHT'):
            return 'green'
        elif value >= st.session_state.get('nominalCHT') and value <= st.session_state.get('highCHT'):
            return 'yellow'
        elif value >= st.session_state.get('highCHT') and value <= st.session_state.get('maxCHT'):
            return 'yellow'
        else:
            return 'red'

    if field == 'EGT':
        if value >= st.session_state.get('minEGT') and value <= st.session_state.get('nominalEGT'):
            return 'blue'
        elif value >= st.session_state.get('nominalEGT') and value <= st.session_state.get('highEGT'):
            return 'yellow'
        elif value >= st.session_state.get('highEGT') and value <= st.session_state.get('maxEGT'):
            return 'yellow'
        else:
            return 'red'

def display_metrics(df, engine_type):
    # Determine Turbo status first to decide column layout
    is_turbo = st.session_state.get('turbo')
    
    # Force 3-column layout as requested
    col_oil, col_fuel, col_other = st.columns(3)

    # CSS (UNCHANGED)
    st.markdown("""
    <style>
    .metric-label { font-size: 14px; text-align: left; color: #ffffff; }
    </style>
    """, unsafe_allow_html=True)

    if not df.empty:
        # --- 1. DEFINE DICTIONARIES ---
        if engine_type == dataOnly:
            # Full set for dataOnly
            metrics = {
                "Front Oil P": ("front_oil_p", "PSI"),
                "Rear Oil P": ("rear_oil_p", "PSI"),
                "Oil Temp": ("oil_temperature", "°C"),
                "Fuel Flow": ("fuel_flow", "GPH"),
                "Metered Fuel P": ("metered_fuel_pressure", "PSI"),
                "Unmetered Fuel P": ("unmetered_fuel_pressure", "PSI"),
                "OAT": ("oat", "°C"),
                "IAT Carb": ("iat_carb", "°C"),
                "Mag Drop": ("mag_drop", ""),
                "TIT": ("tit", "°F"),           # Will be routed to Oil col
                "CDT": ("cdt", "°F"),           # Will be routed to Fuel col
                "Boost": ('upper_deck', "PSI")  # Will be routed to Other col
            }
        else:
            # Threshold mode dictionary
            metrics = {
                "Oil Temp": ("oil_temperature", "°C", st.session_state.get('minOilTemp'), st.session_state.get('nominalOilTemp'), st.session_state.get('highOilTemp'), st.session_state.get('maxOilTemp')),
                "Fuel Flow": ("fuel_flow", "GPH", st.session_state.get('minFuelFlow'), st.session_state.get('nominalFuelFlow'), st.session_state.get('highFuelFlow'), st.session_state.get('maxFuelFlow')),
                "OAT": ("oat", "°C", -500.0, 200.0, 200.0, 200.0),
                "Mag Drop": ("mag_drop", "", st.session_state.get('minMagDrop'), st.session_state.get('nominalMagDrop'), st.session_state.get('highMagDrop'), st.session_state.get('maxMagDrop'))
            }
            # Threshold dynamic additions
            if is_turbo:
                metrics["TIT"] = ("tit", "°F", st.session_state.get('minTIT'), st.session_state.get('nominalTIT'), st.session_state.get('highTIT'), st.session_state.get('maxTIT'))
                metrics["CDT"] = ("cdt", "°F", st.session_state.get('minCDT'), st.session_state.get('minCDT'), st.session_state.get('minCDT'), st.session_state.get('minCDT'))
                metrics["Boost"] = ("upper_deck", "PSI", 0, 30, 35, 40)
            
            # Optional threshold metrics
            if st.session_state.get('rearOilPressure'):
                metrics["Rear Oil P"] = ("rear_oil_p", "PSI", st.session_state.get('minRearOilPressure'), st.session_state.get('nominalRearOilPressure'), st.session_state.get('highRearOilPressure'), st.session_state.get('maxRearOilPressure'))
            if st.session_state.get('frontOilPressure'):
                metrics["Front Oil P"] = ("front_oil_p", "PSI", st.session_state.get('minFrontOilPressure'), st.session_state.get('nominalFrontOilPressure'), st.session_state.get('highFrontOilPressure'), st.session_state.get('maxFrontOilPressure'))
            if st.session_state.get('unmeteredFuelPressure'):
                metrics["Unmetered Fuel P"] = ("unmetered_fuel_pressure", "PSI", st.session_state.get('minUnmeteredFuelPressure'), st.session_state.get('nominalUnmeteredFuelPressure'), st.session_state.get('highUnmeteredFuelPressure'), st.session_state.get('maxUnmeteredFuelPressure'))
            if st.session_state.get('meteredFuelPressure'):
                metrics["Metered Fuel P"] = ("upper_deck", "PSI", st.session_state.get('minMeteredFuelPressure'), st.session_state.get('nominalMeteredFuelPressure'), st.session_state.get('highMeteredFuelPressure'), st.session_state.get('maxMeteredFuelPressure'))

        # --- 2. THE NEW ROUTING LOGIC ---
        for label, val_tuple in metrics.items():
            field = val_tuple[0]
            unit = val_tuple[1]
            label_low = label.lower()
            
            # Special logic: Append Turbo fields to specific columns
            if "oil" in label_low or "tit" in label_low:
                target_col = col_oil
            elif "fuel" in label_low or "cdt" in label_low:
                target_col = col_fuel
            else:
                # This catches OAT, Mag Drop, and Boost
                target_col = col_other

            # --- 3. RENDERING (UNCHANGED) ---
            with target_col.container(height=95):
                if field in df.columns and len(df) > 1:
                    latest_value = float(df[field].iloc[0])
                    previous_value = float(df[field].iloc[-1])
                    delta = latest_value - previous_value

                    # Determine Color
                    text_color = "white"
                    if len(val_tuple) > 2:
                        minT, nomT, highT, maxT = val_tuple[2:]
                        if "oil_p" in field.lower():
                            if latest_value < nomT: text_color = "red"
                            elif latest_value < highT: text_color = "yellow"
                            elif latest_value <= maxT: text_color = "green"
                            else: text_color = "red"
                        else:
                            if latest_value > maxT: text_color = "red"
                            elif highT < latest_value <= maxT: text_color = "red"
                            elif nomT < latest_value <= highT: text_color = "yellow"
                            elif minT < latest_value <= nomT: text_color = "green"
                            else: text_color = "red"

                    display_text = f"<span style='color:{text_color}; font-size:24px;'>{latest_value:.1f} {unit}</span>"
                    delta_display = f"<span style='color:{'green' if delta > 0 else 'red'}; font-size:18px;'>{'↑' if delta > 0 else '↓'} {abs(delta):.1f}</span>" if delta != 0 else ""

                    st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
                    st.markdown(f"<div style='text-align:left;'>{display_text}</div>", unsafe_allow_html=True)
                    st.markdown(f"<div style='text-align:left;'>{delta_display}</div>", unsafe_allow_html=True)
                else:
                    st.markdown(f"<div class='metric-label'>{label}</div>", unsafe_allow_html=True)
                    st.markdown("<div style='text-align:left; color:#ffffff;'>N/A</div>", unsafe_allow_html=True)
                    
def plot_egt_cht(df, engine_type):
    """
    Plot EGT and CHT values in a bar graph with Diff metrics.
    """
    temps_start_time = time.time()
    
    # --- 1. Identify EGT and CHT columns based on engine type ---
    if st.session_state.get('cylinders') == 4:
        egt_columns = sorted([col for col in df.columns if col.startswith('egt_')])[:4]
        cht_columns = sorted([col for col in df.columns if col.startswith('cht_')])[:4]
    else:
        # Default to 6 cylinders if 6 or unknown
        egt_columns = sorted([col for col in df.columns if col.startswith('egt_')])[:6]
        cht_columns = sorted([col for col in df.columns if col.startswith('cht_')])[:6]

    # --- 2. Calculate Max, Min, and Diff (The new functionality) ---
    egt_values = [df[col].iloc[0] for col in egt_columns]
    cht_values = [df[col].iloc[0] for col in cht_columns]

    max_egt = max(egt_values)
    min_egt = min(egt_values)
    diff_egt = max_egt - min_egt

    max_cht = max(cht_values)
    min_cht = min(cht_values)
    diff_cht = max_cht - min_cht

    # --- 3. Prepare Data for Plotting ---
    labels = []
    values = []
    text = []
    colors = []

    # Helper function to generate labels/colors
    for egt_col, cht_col in zip(egt_columns, cht_columns):
        # Process EGT
        egt_val = int(df[egt_col].iloc[0])
        labels.append(egt_col)
        values.append(egt_val)
        
        # Add (H) for High and (L) for Low to the bar text
        note = ""
        if egt_val == max_egt: note = " (H)"
        # elif egt_val == min_egt: note = " (L)"
        text.append(f'{egt_val} °F{note}')
        
        # Color logic
        # if engine_type != 'dataOnly':
        if engine_type != dataOnly:
             colors.append(get_color('EGT', egt_val)) # Assuming get_color exists in your scope
        
        # Process CHT
        cht_val = int(df[cht_col].iloc[0])
        labels.append(cht_col)
        values.append(cht_val)

        # Add (H) for High and (L) for Low
        note = ""
        if cht_val == max_cht: note = " (H)"
        # elif cht_val == min_cht: note = " (L)"
        text.append(f'{cht_val} °F{note}')

        # Color logic
        # if engine_type != 'dataOnly':
        if engine_type != dataOnly:
             colors.append(get_color('CHT', cht_val))
        # else:


    # --- 4. Create Traces ---
    # if engine_type == 'dataOnly':
    if engine_type == dataOnly:
        # Simple Green/Blue coloring
        trace = go.Bar(
            x=labels,
            y=values,
            text=text,
            textposition='auto',
            textfont=dict(size=16),
            marker_color=['green' if 'egt_' in label else 'blue' for label in labels]
        )
    else:
        # Custom coloring based on get_color function
        trace = go.Bar(
            x=labels,
            y=values,
            marker_color=colors,
            text=text,
            textfont=dict(size=16),
            textposition='auto'
        )

    fig = go.Figure(data=[trace])

    # --- 5. Update Layout with Diff Metrics in Title ---
    # This formats a title like: "EGT Diff: 100°F  |  CHT Diff: 25°F"
    title_text = (
        f"<b>Temps Diffs</b><br>"
        f"<span style='font-size: 16px; color: gray;'>"
        f"EGT Diff: {diff_egt}°F  |  CHT Diff: {diff_cht}°F"
        f"</span>"
    )

    fig.update_layout(
        title=dict(text=title_text, x=0.5, xanchor='center'), # Center the title
        width=850,
        height=400,
        margin=dict(l=20, r=20, t=60, b=20), # Increased top margin (t) to fit the title
        barmode='group',
        xaxis_tickangle=-45
    )

    # Performance tracking
    temps_plot_time = time.time() - temps_start_time
    st.session_state['temps_plot_render_time'] = temps_plot_time
    
    return fig



def animate_rpm(df, engine_type):
    """
    Create a gauge chart for RPM values.
    
    Args:
    - df (pd.DataFrame): DataFrame containing the data.
    """
    rpm_start_time = time.time()  # Start time for performance measurement
    
    rpm = df['rpm'].iloc[0] if not df.empty else 0

    if engine_type == dataOnly:
        gaugeSetup = {
            'axis': {'range': [None, 3000]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, 750], 'color': "lightblue"},
                {'range': [750, 1500], 'color': "skyblue"},
                {'range': [1500, 2250], 'color': "deepskyblue"},
                {'range': [2250, 3000], 'color': "dodgerblue"}
            ]
        }

    else:
        gaugeSetup = {
            'axis': {'range': [st.session_state.get('minRPM'), int(st.session_state.get('maxRPM'))+500]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, st.session_state.get('maxRPM')/3], 'color': "lightblue"},  
                {'range': [st.session_state.get('maxRPM')/3, 2*(st.session_state.get('maxRPM')/3)], 'color': "skyblue"}, 
                {'range': [2*(st.session_state.get('maxRPM')/3), st.session_state.get('maxRPM')], 'color': "deepskyblue"},
                {'range': [st.session_state.get('maxRPM'), int(st.session_state.get('maxRPM'))+500], 'color': "dodgerblue"}
                ],
            'threshold' : {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': st.session_state.get('maxRPM')}
        }


    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=rpm,
        domain={'x': [0,0.55], 'y': [0,1]},
        title={'text': "RPM"},
        gauge=gaugeSetup
    ))

    # Update layout to adjust size
    fig.update_layout(
        width=300,  # Set the desired width
        height=120,  # Set the desired height
        margin=dict(l=20, r=20, t=45, b=20)  # Adjust margins if necessary
    )

    rpm_plot_time = time.time() - rpm_start_time  # Measure time taken to render plot
    st.session_state['rpm_plot_render_time'] = rpm_plot_time  # Store plot render time in session state

    return fig


def animate_manifold_pressure(df, engine_type):
    """
    Create a gauge chart for Manifold Pressure values.
    
    Args:
    - df (pd.DataFrame): DataFrame containing the data.
    """
    mp_start_time = time.time()  # Start time for performance measurement
    manifold_pressure = df['manifold_pressure'].iloc[0] if not df.empty else 0

    if engine_type == dataOnly:
        gaugeSetup ={
            'axis': {'range': [None, 40]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, 10], 'color': "lightblue"},
                {'range': [10, 20], 'color': "skyblue"},
                {'range': [20, 30], 'color': "deepskyblue"}, 
                {'range': [30, 40], 'color': "dodgerblue"}
            ]
        }
    else:
        gaugeSetup = {
            'axis': {'range': [st.session_state.get('minMP'), int(st.session_state.get('maxMP'))+5]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, st.session_state.get('maxMP')/3], 'color': "lightblue"},
                {'range': [st.session_state.get('maxMP')/3, 2*(st.session_state.get('maxMP')/3)], 'color': "skyblue"},
                {'range': [2*(st.session_state.get('maxMP')/3), st.session_state.get('maxMP')], 'color': "deepskyblue"},
                {'range': [st.session_state.get('maxMP'), int(st.session_state.get('maxMP'))+5], 'color': "dodgerblue"}
            ],
            'threshold' : {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': st.session_state.get('maxMP')}
        }

    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=manifold_pressure,
        number={'suffix': " hg"},  # Add the units here
        domain={'x': [0,0.55], 'y': [0,1]},
        title={'text': "MP"},
        gauge=gaugeSetup
    ))

    # Update layout to adjust size
    fig.update_layout(
        width=300,  # Set the desired width
        height=120,  # Set the desired height
        margin=dict(l=20, r=20, t=45, b=20)  # Adjust margins if necessary
    )

    mp_plot_time = time.time() - mp_start_time  # Measure time taken to render plot
    st.session_state['mp_plot_render_time'] = mp_plot_time  # Store plot render time in session state

    return fig

def animate_hp(df, engine_type):
    """
    Create a gauge chart for horsepower values.
    
    Args:
    - df (pd.DataFrame): DataFrame containing the data.
    """
    hp_start_time = time.time()  # Start time for performance measurement
    horsepower = df['hp'].iloc[0] if not df.empty else 0

    if engine_type == dataOnly:
        gaugeSetup = {
            'axis': {'range': [None, 200]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, 50], 'color': "lightblue"},
                {'range': [50, 100], 'color': "skyblue"},
                {'range': [100, 150], 'color': "deepskyblue"}, 
                {'range': [150, 200], 'color': "dodgerblue"}
            ]
        }

    else:
        gaugeSetup = {
            'axis': {'range': [st.session_state.get('minHP'), int(st.session_state.get('maxHP'))+30]},  # Adjust the range according to your data
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, st.session_state.get('maxHP')/3], 'color': "lightblue"},
                {'range': [st.session_state.get('maxHP')/3, 2*(st.session_state.get('maxHP')/3)], 'color': "skyblue"},
                {'range': [2*(st.session_state.get('maxHP')/3), st.session_state.get('maxHP')], 'color': "deepskyblue"},
                {'range': [st.session_state.get('maxHP'), int(st.session_state.get('maxHP'))+30], 'color': "dodgerblue"}
            ],
            'threshold' : {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': st.session_state.get('maxHP')}
        }


    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=horsepower,
        domain={'x': [0,0.55], 'y': [0,1]},
        title={'text': "HP"},
        gauge=gaugeSetup
    ))

    # Update layout to adjust size
    fig.update_layout(
        width=300,  # Set the desired width
        height=120,  # Set the desired height
        margin=dict(l=20, r=20, t=45, b=20)  # Adjust margins if necessary
    )

    mp_plot_time = time.time() - hp_start_time  # Measure time taken to render plot
    st.session_state['hp_plot_render_time'] = mp_plot_time  # Store plot render time in session state

    return fig


def plot_performance_metrics():
    """
    Create gauge charts for Data Fetch Time and Plot Render Time.
    """

    global refresh_time
    data_fetch_time = st.session_state.get('data_fetch_time', 0)
    temps_plot_render_time = st.session_state.get('temps_plot_render_time', 0)
    rpm_plot_render_time = st.session_state.get('rpm_plot_render_time', 0)
    mp_plot_render_time = st.session_state.get('mp_plot_render_time', 0)
    hp_plot_render_time = st.session_state.get('hp_plot_render_time', 0)
    total_plot_time = temps_plot_render_time + rpm_plot_render_time + mp_plot_render_time + refresh_time + hp_plot_render_time

    # Create figure with two subplots
    #fig = go.Figure()

    # Data Fetch & Plot Render Time Gauge
    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        domain={'x': [0,0.75], 'y': [0,1]},
        value=data_fetch_time + total_plot_time,
        number={'suffix': " s"},  # Add the units here
        title={'text': "Lag"},
        gauge={
            'axis': {'range': [None, 1.5]},  # Adjust the range according to your needs
            'bar': {'color': "blue"},
            'steps': [
                {'range': [0, 0.4], 'color': "lightblue"},
                {'range': [0.4, 0.8], 'color': "skyblue"},
                {'range': [0.8, 1.5], 'color': "deepskyblue"},
                # {'range': [1.4, 2.5], 'color': "dodgerblue"}
            ]
        }
    ))

    # Update layout to adjust size
    fig.update_layout(
        width=300,  # Set the desired width
        height=120,  # Set the desired height
        margin=dict(l=20, r=20, t=45, b=20)  # Adjust margins if necessary
    )
    
    return fig


# Function to convert DataFrame to CSV
# def convert_df_to_csv(test_sheet_df):
#     return test_sheet_df.to_csv(index=False).encode('utf-8')


def start_button_handler(df):
    row_map = get_row_map(st.session_state.test_sheet) # Generate the map
    
    now = datetime.datetime.now()
    test_date = now.strftime("%Y-%m-%d")
    utc_now = datetime.datetime.now(datetime.UTC)
    ast_tz = pytz.timezone('America/Halifax')
    ast_now = utc_now.replace(tzinfo=pytz.utc).astimezone(ast_tz)
    start_time_str = ast_now.strftime("%H:%M:%S")

    # Get values from state
    # ... (Your existing variable retrieval code here) ...
    specOilPress = st.session_state.get('SpecOilPressure')
    specIdlePress = st.session_state.get('SpecIdleOilPressure')
    specFTFuelPressureMetered = st.session_state.get('SpecF/tFuelPressureMetered')
    specFTFuelPressureUnMetered = st.session_state.get('SpecF/tFuelPressureUnmetered')
    specIdleFuelPressureMetered = st.session_state.get('SpecIdleFuelPressureMetered')
    specIdleFuelPressureUnMetered = st.session_state.get('SpecIdleFuelPressureUnmetered')
    maxCHT = st.session_state.get("MAXCHT")
    maxEGTTIT = st.session_state.get("MAXEGT/TIT")
    maxManifold = st.session_state.get("MaxManifold")
    maxTurboBoost = st.session_state.get("MaxTurboBoost")
    
    customer_name = st.session_state.get("customer_name")
    sn_number = st.session_state.get("sn_number")
    work_order = st.session_state.get("work_order")
    test_nature = st.session_state.get("test_nature")
    prop_type = st.session_state.get("selected_prop")
    engine_type = st.session_state.get("selected_engine") 

    split_engine_str = engine_type.split() 
    engineMake = split_engine_str[0] if len(split_engine_str) > 0 else ""
    engineModel = split_engine_str[1] if len(split_engine_str) > 1 else ""

    # Ensure columns exist (Keep your existing loop)
    # for col in ['Unnamed: 1','Unnamed: 4','Unnamed: 7','Unnamed: 8','Unnamed: 13','Unnamed: 15','Unnamed: 17',]:
    for col in ['Unnamed: 2','Unnamed: 3','Unnamed: 7','Unnamed: 10','Unnamed: 11','Unnamed: 17','Unnamed: 18']:
        if col not in st.session_state.test_sheet.columns:
            st.session_state.test_sheet[col] = None

    # --- ANCHOR BASED UPDATES ---
    # We use row_map.get('Label') to find the row index dynamically
    
    # Header Info (Usually fixed at top, but safer to anchor if possible, or keep hardcoded if header never changes)
    st.session_state.test_sheet.loc[2, 'Unnamed: 9'] = work_order
    st.session_state.test_sheet.loc[3, 'Unnamed: 9'] = test_date
    st.session_state.test_sheet.loc[5, 'Unnamed: 9'] = test_nature
    st.session_state.test_sheet.loc[2, 'Unnamed: 17'] = prop_type
    st.session_state.test_sheet.loc[3, 'Unnamed: 17'] = start_time_str
    st.session_state.test_sheet.loc[2, 'Unnamed: 1'] = customer_name
    st.session_state.test_sheet.loc[3, 'Unnamed: 1'] = engineMake
    st.session_state.test_sheet.loc[4, 'Unnamed: 1'] = engineModel
    st.session_state.test_sheet.loc[5, 'Unnamed: 1'] = sn_number
    
    # Specs - Dynamic Rows
    try:
        # Find rows based on labels
        r_spec_oil = row_map['Spec Oil Pressure:']
        r_spec_idle = row_map['Spec Idle Oil Press:'] # Ensure this matches your excel label exactly!
        
        # Populate based on found rows
      
        st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 2'] = specOilPress
        st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 2'] = specIdlePress
        
        
        st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 6'] = specFTFuelPressureMetered
        st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 6'] = specIdleFuelPressureMetered
        
        
        st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 10'] = specFTFuelPressureUnMetered
        st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 10'] = specIdleFuelPressureUnMetered
        
       
        st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 16'] = maxCHT
        st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 16'] = maxEGTTIT
        
       
        st.session_state.test_sheet.loc[r_spec_oil, 'Unnamed: 19'] = maxManifold
        st.session_state.test_sheet.loc[r_spec_idle, 'Unnamed: 19'] = maxTurboBoost
        st.session_state.start_btn_toast = "Date, Time & Specs populated"
        st.toast(st.session_state.start_btn_toast)
        
    except KeyError as e:
        st.error(f"Error: Could not find row label {e} in the Excel sheet.")

    st.session_state.test_date = test_date
    st.session_state.start_time_str = start_time_str

def snap_button_handler(df):
    """
    Writes rpm fields into the testsheet
    """
    if df.empty:
        st.warning("Cannot snap empty dataset")
        return
    
    # Ensure the sheet is in 'object' mode to prevent TypeError
    if not (st.session_state.test_sheet.dtypes == 'object').all():
        st.session_state.test_sheet = st.session_state.test_sheet.astype(object)

    current_index = get_snapshot_index()
    if current_index is None:
        # Default to the first RPM in the sequence if index isn't set
        current_index = row_map.get(rpm_sequence[0])
    
    # Get the mapping of Label -> Row Index
    row_map = get_row_map(st.session_state.test_sheet)
    rpm_sequence = get_rpm_sequence(test_sheet_selected_rpm) # e.g. ['2700', '2400', ... '750']
    
    # Identify the RPM label via the index to address missing label issue
    # Create a reverse map: Index -> Label
    reverse_row_map = {v: k for k, v in row_map.items()}
    

    # 1. Identify where we are in the sequence
    # Get the RPM string of the current row
    # current_row_label = str(st.session_state.test_sheet.loc[current_index, 'Rev. 5']).strip()
    current_row_label = reverse_row_map.get(current_index)

    # Verify we are on a valid RPM row
    if current_row_label not in rpm_sequence:
        st.warning(f"Current row '{current_row_label}' is not a recognized RPM step.")
        # Try to recover? For now, just proceed with data capture
    
    # ... (Your existing Data Capture logic variables here) ...
    # manifold_pressure = ...
    # front_oil_p = ...
    # ...
    manifold_pressure = float(df['manifold_pressure'].iloc[0])
    front_oil_p = int(df['front_oil_p'].iloc[0])
    rear_oil_p = int(df['rear_oil_p'].iloc[0])
    oil_temp = int(df['oil_temperature'].iloc[0])
    metered_fuel_pressure = float(df['metered_fuel_pressure'].iloc[0])
    unmetered_fuel_pressure = float(df['unmetered_fuel_pressure'].iloc[0])
    boost_pressure = float(df['upper_deck'].iloc[0])
    egt_1 = int(df['egt_1'].iloc[0])
    egt_2 = int(df['egt_2'].iloc[0])
    egt_3 = int(df['egt_3'].iloc[0])
    egt_4 = int(df['egt_4'].iloc[0])
    egt_5 = int(df['egt_5'].iloc[0])
    egt_6 = int(df['egt_5'].iloc[0])
    cht_1 = int(df['cht_1'].iloc[0])
    cht_2 = int(df['cht_2'].iloc[0])
    cht_3 = int(df['cht_3'].iloc[0])
    cht_4 = int(df['cht_4'].iloc[0])
    cht_5 = int(df['cht_5'].iloc[0])
    cht_6 = int(df['cht_6'].iloc[0])
    row_rpm = st.session_state.test_sheet.loc[current_index, 'Rev. 5']
    
    # ... (Your existing st.session_state.test_sheet.loc assignments here) ...
    # This part writes to 'current_index' which is correct.
    # Ensure the columns exist
    for col in ['Unnamed: 1', 'Unnamed: 2','Unnamed: 4','Unnamed: 6','Unnamed: 8'
               ,'Unnamed: 9','Unnamed: 10','Unnamed: 11', 'Unnamed: 12', 'Unnamed: 13'
               ,'Unnamed: 14','Unnamed: 15','Unnamed: 16','Unnamed: 17', 'Unnamed: 18', 'Unnamed: 19']:
        if col not in st.session_state.test_sheet.columns:
            st.session_state.test_sheet[col] = None

    # Modify specific cells with current date and time
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 1'] = manifold_pressure
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 3'] = oil_temp
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 4'] = front_oil_p
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 5'] = rear_oil_p
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 6'] = metered_fuel_pressure
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 7'] = unmetered_fuel_pressure
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 8'] = egt_1
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 9'] = egt_2
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 10'] = egt_3
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 11'] = egt_4
    
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 14'] = cht_1
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 15'] = cht_2
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 16'] = cht_3
    st.session_state.test_sheet.loc[current_index, 'Unnamed: 17'] = cht_4
    

    if cylinderCount == 6:
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 12'] = egt_5
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 13'] = egt_6
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 18'] = cht_5
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 19'] = cht_6

    else:
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 12'] = "-"
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 13'] = "-"
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 18'] = "-"
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 19'] = "-"

    if turboStatus:
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 2'] = boost_pressure
    
    else:
        st.session_state.test_sheet.loc[current_index, 'Unnamed: 2'] = "-"

    # Check if cells are populated
    snap_cells = [(current_index, 'Unnamed: 1'), (current_index, 'Unnamed: 3'), (current_index, 'Unnamed: 4')
                  ,(current_index, 'Unnamed: 6'), (current_index, 'Unnamed: 6'), (current_index, 'Unnamed: 8')
                  ,(current_index, 'Unnamed: 9'), (current_index, 'Unnamed: 10'), (current_index, 'Unnamed: 11')
                  , (current_index, 'Unnamed: 12'), (current_index, 'Unnamed: 13'), (current_index, 'Unnamed: 14')
                  , (current_index, 'Unnamed: 15'), (current_index, 'Unnamed: 16'), (current_index, 'Unnamed: 17')
                  , (current_index, 'Unnamed: 18'), (current_index, 'Unnamed: 19')]


    
    # 2. MOVE TO NEXT RPM
    try:

        # Find index in the list ['2700', '2400'...]
        seq_idx = rpm_sequence.index(current_row_label)
        
        # Check if there is a next step
        if seq_idx < len(rpm_sequence) - 1:
            next_rpm = rpm_sequence[seq_idx + 1] # Get next lower RPM, e.g. "2400"
            next_row_index = row_map.get(next_rpm)
            
            if next_row_index:
                st.session_state.snapshot_index = next_row_index
                st.toast(f"Snapshot taken. Next: {next_rpm} RPM")
            else:
                st.error(f"Next RPM {next_rpm} not found in sheet.")
        else:
            # We are at the end (750 RPM)
            st.toast("All RPMs populated. Confirm Idle.")
            
    except ValueError:
        # current_row_label was not in our list (maybe user manually clicked elsewhere?)
        st.error("Lost track of RPM sequence.")

    # Push testsheeet to mongodb
    # persist_test_sheet()

def undo_button_handler(df):
    # Setup Mappping and Sequence
    row_map = get_row_map(st.session_state.test_sheet)
    reverse_map = {v: k for k, v in row_map.items()}
    rpm_sequence = get_rpm_sequence(test_sheet_selected_rpm)
    
    # Get current index
    current_index = get_snapshot_index()  
    current_label = reverse_map.get(current_index)
    
    # Clear the data at the undo_index
    snap_cols = [
        'Unnamed: 1', 'Unnamed: 2', 'Unnamed: 3', 'Unnamed: 4','Unnamed: 5','Unnamed: 6'
        ,'Unnamed: 7','Unnamed: 8','Unnamed: 9','Unnamed: 10','Unnamed: 11'
        ,'Unnamed: 12', 'Unnamed: 13','Unnamed: 14','Unnamed: 15','Unnamed: 16'
        ,'Unnamed: 17','Unnamed: 18','Unnamed: 19'
    ]

    # Check if any of the cells in the current row are not empty
    is_current_row_populated = any(pd.notna(st.session_state.test_sheet.at[current_index, col]) and
        st.session_state.test_sheet.at[current_index, col] != ''
        for col in snap_cols)

    # Find the RPM we want to "Undo"
    
    if is_current_row_populated:
        # If 750 has data, we want to clear 750!
        undo_index = current_index
        undo_rpm_label = current_label
    else:
            try:
                seq_idx = rpm_sequence.index(current_label)

                if seq_idx == 0:
                    st.toast(f"Limits exceeded. Current Label: {current_label}, Current Index: {current_index}")
                    return
            
                # The target  for undo is the previous RPM in the list
                undo_rpm_label = rpm_sequence[seq_idx - 1]
                undo_index = row_map.get(undo_rpm_label)

                # if undo_index is None:
                #     st.toast(f"Could not map row for {undo_rpm_label}")

            except ValueError:
                st.toast("Undo failed: Current position not found in RPM sequence")

    for col in snap_cols:
        st.session_state.test_sheet.at[undo_index, col] = ''

    # Reset the snapshot_index back to the one we just cleared
    st.session_state.snapshot_index =  undo_index
    st.toast(f"RPM: {undo_rpm_label} cleared")


def end_button_handler(df):
    """
    
    """
    # Ensure the sheet is in 'object' mode to prevent TypeError
    if not (st.session_state.test_sheet.dtypes == 'object').all():
        st.session_state.test_sheet = st.session_state.test_sheet.astype(object)
    row_map = get_row_map(st.session_state.test_sheet)
    
    # ... (Time calculation logic remains the same) ...
    # Get current UTC time and convert to AST
    utc_now = datetime.datetime.now(datetime.UTC)
    ast_tz = pytz.timezone('America/Halifax')  # Atlantic Standard Time (AST) timezone
    ast_now = utc_now.replace(tzinfo=pytz.utc).astimezone(ast_tz)
    end_time = ast_now.isoformat()  # Use AST timestamp
    end_time_str = ast_now.strftime("%H:%M:%S")
    # outside_temp = int(get_beaver_bank_weather())
    # now = datetime.now()
    # end_time = now.strftime("%H:%M:%S")
    st.session_state.end_time = end_time

    # Ensure the columns 'Unnamed: 15' exist
    # if 'Unnamed: 15' not in st.session_state.test_sheet.columns:
    #     st.session_state.test_sheet['Unnamed: 16'] = None
    if 'Unnamed: 16' not in st.session_state.test_sheet.columns:
        st.session_state.test_sheet['Unnamed: 16'] = None

    # Calculate the total time of the test
    start_time_hold = st.session_state.start_time_str
    end_time_hold = end_time_str

    # Define a time format
    time_format = "%H:%M:%S"

    # Convert the time strings to datetime objects (using today's date)
    now = datetime.datetime.now()
    start_time_obj = datetime.datetime.strptime(start_time_hold, time_format).replace(year=now.year, month=now.month, day=now.day)
    end_time_obj = datetime.datetime.strptime(end_time_hold, time_format).replace(year=now.year, month=now.month, day=now.day)

    # start_time_obj = datetime.strptime(st.session_state.start_time, '%H:%M:%S')
    # end_time_obj = datetime.strptime(end_time, '%H:%M:%S')
    total_time = end_time_obj - start_time_obj

    # Add outside air temperature
    #outside_temp = int(df['oat'].iloc[0])
    # st.session_state.test_sheet.loc[3, 'Unnamed: 8'] = outside_temp
    st.session_state.test_sheet.loc[4, 'Unnamed: 9'] = st.session_state.beaver_oat

    # Modify specific cells with current date and time
    # st.session_state.test_sheet.loc[3, 'Unnamed: 15'] = end_time_str
    # st.session_state.test_sheet.loc[4, 'Unnamed: 15'] = str(total_time)
    st.session_state.test_sheet.loc[4, 'Unnamed: 17'] = end_time_str
    # st.session_state.test_sheet.loc[5, 'Unnamed: 17'] = str(total_time)
    st.session_state.test_sheet.loc[5, 'Unnamed: 17'] = st.session_state.run_time
    
    # --- ANCHOR BASED FOOTER POPULATION ---
    try:
        # Find Footer Rows
        r_idle_fuel = row_map.get('Idle Fuel Pressure:') # Or whatever the label is in the footer
        r_ft_fuel = row_map.get('Full Throttle Fuel Press:') # Check exact label in Excel
        r_idle_rpm = row_map.get('Idle RPM:')
        r_ft_rpm = row_map.get('Full Throttle RPM:')
        
        # Find Source Data Rows (The actual RPM test rows)
        r_src_idle = 9 # Idle RPM row
        
        # Determine Full Throttle RPM Source 
        if test_sheet_selected_rpm == 2700:
            r_src_ft = 18
        else:
            r_src_ft = 17
        
        # prop_type = st.session_state.get("selected_prop")
        # # ft_rpm_label = "2700" if "2 Blade" in prop_type else "2200"
        # if prop_type == "2 Blade (Yellow Tipped)":
        #     ft_rpm_label = "2700"
        # else:
        #     ft_rpm_label = "2200"
        # r_src_ft = row_map.get(ft_rpm_label)

        if not r_src_idle or not r_src_ft:
            st.error("Could not find source data rows (750 or Max RPM)")
            return

        # 1. Populate Idle Section
        if r_idle_fuel:
            # Idle Fuel Press (Metered/Unmetered)
            st.session_state.test_sheet.loc[r_idle_fuel, 'Unnamed: 2'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 6']
            st.session_state.test_sheet.loc[r_idle_fuel, 'Unnamed: 4'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 7']
            # Idle Rear Oil Press
            st.session_state.test_sheet.loc[r_idle_fuel, 'Unnamed: 8'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 5']
            # Idle Rear Oil Pressure at Oil Temp
            st.session_state.test_sheet.loc[r_idle_fuel, 'Unnamed: 11'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 3']

        # 2. Populate Full Throttle Section
        if r_ft_fuel:
            # F/T Fuel Press (Metered/Unmetered)
            st.session_state.test_sheet.loc[r_ft_fuel, 'Unnamed: 2'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 6']
            st.session_state.test_sheet.loc[r_ft_fuel, 'Unnamed: 4'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 7']
            
        if r_idle_rpm:
            # Idle RPM @ Manifold Pressure
            # st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 1'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 0']
            # st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 4'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 1']
            st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 2'] = st.session_state.test_sheet.loc[r_src_idle, 'Rev. 5']
            st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 4'] = st.session_state.test_sheet.loc[r_src_idle, 'Unnamed: 1']
            # F/T Rear Oil Press
            st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 8'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 5']
            # F/T Rear Oil Pressure at Oil Temp
            st.session_state.test_sheet.loc[r_idle_rpm, 'Unnamed: 11'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 3']

        if r_ft_rpm:
            # F/T RPM @ Manifold Pressure
            # st.session_state.test_sheet.loc[r_ft_rpm, 'Unnamed: 1'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 0']
            # st.session_state.test_sheet.loc[r_ft_rpm, 'Unnamed: 4'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 1']
            st.session_state.test_sheet.loc[r_ft_rpm, 'Unnamed: 2'] = st.session_state.test_sheet.loc[r_src_ft, 'Rev. 5']
            st.session_state.test_sheet.loc[r_ft_rpm, 'Unnamed: 4'] = st.session_state.test_sheet.loc[r_src_ft, 'Unnamed: 1']

        st.toast("Footer populated successfully")
        
    except Exception as e:
        st.error(f"Error populating footer: {str(e)}")

async def calculate_engine_run_time(uri, db_local, collection_local):
    # Create new client inside active event loop
    local_client = motor.motor_asyncio.AsyncIOMotorClient(uri)
    local_db = local_client[db_local]
    local_collection = local_db[collection_local]

    # Fetch only timestamp and rpm from database
    pipeline = [
        {"$project": {
            "_id": 0,
            "timestamp": 1,
            "rpm": "$fields.rpm"
        }},
        {"$sort": {"timestamp": 1}}
    ]

    cursor = local_collection.aggregate(pipeline)
    data = await cursor.to_list(length=None)

    if not data:
        return datetime.timedelta(seconds=0)
    
    df = pd.DataFrame(data)
    df['timestamp'] = pd.to_datetime(df['timestamp'])

    # Ensure rpm is numeric, fill NaNs with 0
    df['rpm'] = pd.to_numeric(df['rpm'], errors='coerce').fillna(0)

    # Calculate the time difference between consecutive rows
    df['time_delta_seconds'] = df['timestamp'].diff().dt.total_seconds().fillna(0)

    # Filter for rows where the engine was actually running
    # Using rpm > 0 or rpm > title
    running_df = df[df['rpm'] > 0]

    # Sum the time deltas
    total_running_seconds = running_df['time_delta_seconds'].sum()

    # Calculate hours and minutes
    hours = total_running_seconds // 3600
    minutes = (total_running_seconds % 3600) // 60

    return running_df, f"{int(hours):02d}:{int(minutes):02d}"

# Function to start the subprocess and store its information in session_state
def start_process():
    if 'subprocess' in st.session_state:
        # poll() return None if the background process is still running
        if st.session_state.subprocess.poll() is None:
            st.warning("A test is already running! Action blocked.")
            return
        else:
            # The process died or finished on another page; clear subprocess
            del st.session_state.subprocess

    # Get the current date and time from session state
    now = datetime.datetime.now()
    # current_date = st.session_state.test_date
    current_date = now.strftime("%Y-%m-%d")
    utc_now = datetime.datetime.now(datetime.UTC)
    ast_tz = pytz.timezone('America/Halifax')
    ast_now = utc_now.replace(tzinfo=pytz.utc).astimezone(ast_tz)
    current_time = ast_now.strftime("%H:%M:%S")
    # current_time = st.session_state.start_time_str

    # Check id db_name and collection_name are in session state
    if 'db_name' not in st.session_state:
        st.session_state.db_name = None
    elif 'collection_name' not in st.session_state:
        st.session_state.collection_name = None
    
    #  Define the database and collection names
    # db_name = current_date
    # collection_name = f"{current_date}_{current_time}"

    split_engine_str = engine_type.split() 
    engineMake = split_engine_str[0]
    engineModel = split_engine_str[1]
    db_name = engineMake + "-" + engineModel + "-" + current_date
    collection_name = current_time


    # Store the db_name & collection name in session state
    st.session_state.db_name = db_name
    st.session_state.collection_name = collection_name

    if 'subprocess' not in st.session_state:
        process = subprocess.Popen(["python", parse_data_path, db_name, collection_name])
        st.session_state.subprocess = process
        st.session_state.process_running = True # Dynamic flag for buttons
        st.toast(f"Process started. Writing to {db_name}_{collection_name}")

    # Load blank/saved test sheet into run
    # persist_test_sheet()

# Function to stop the subprocess if it's running
def stop_process():
    if 'subprocess' in st.session_state:
        process = st.session_state.subprocess
        if process.poll():
            process.terminate()  # This sends SIGTERM to the process
        del st.session_state.subprocess # Remove it from session_state
        st.toast("Process terminated")
    st.session_state.process_running = False # Reset dynamic flag for buttons

# 1. Define your target addresses (Row Index, Column Name)
TARGET_ADDRESSES = [
    # Idle and f/t cells
    (9, "Rev. 5"), (18, "Rev. 5"),
    # Oil Level Cells
    (14, "Unnamed: 18"), (19, "Unnamed: 18"), 
    # Other Required Cells
    (23, "Unnamed: 17"), (24, "Unnamed: 17"), (25, "Unnamed: 16"),
    (25, "Unnamed: 18"), (26, "Unnamed: 18"), (27, "Unnamed: 2"),
    (27, "Unnamed: 4"), (27, "Unnamed: 8"), (27, "Unnamed: 12"),
    (27, "Unnamed: 18"), (30, "Unnamed: 12"), (31, "Unnamed: 12"),
    # Oil Details Cells
    (32, "Unnamed: 2"), (32, "Unnamed: 4"), (32, "Unnamed: 6"), 
    (32, "Unnamed: 8"), (32, "Unnamed: 10"), (32, "Unnamed: 12"), 
    (32, "Unnamed: 15"), (32, "Unnamed: 18"),
    # Cylinder Stats
    (30, "Unnamed: 2"), (30, "Unnamed: 3"), (30, "Unnamed: 4"),
    (30, "Unnamed: 5"), #(30, "Unnamed: 6"), (30, "Unnamed: 7"),
    (31, "Unnamed: 2"), (31, "Unnamed: 3"), (31, "Unnamed: 4"),
    (31, "Unnamed: 5") #, (31, "Unnamed: 6"), (31, "Unnamed: 7")
]

# Add cyl_5 and cyl_6 to target address
if cylinderCount == 6:
    TARGET_ADDRESSES = TARGET_ADDRESSES + [(30, "Unnamed: 6"), (30, "Unnamed: 7"),
                                           (31, "Unnamed: 6"), (31, "Unnamed: 7")]

# Convert tuples to lists for JavaScript compatibility
JS_TARGETS = [list(addr) for addr in TARGET_ADDRESSES]

# 2. Updated JavaScript Logic
# This checks if the current cell matches an address in our list and is empty.
highlight_jscode = JsCode(f"""
function(params) {{
    const targets = {JS_TARGETS};
    const row = params.node.rowIndex;
    const col = params.column.getColId();
    const val = params.value;

    // Check if this specific cell is one of our targets
    const isTarget = targets.some(t => t[0] === row && t[1] === col);
    
    if (isTarget) {{
        // Check if the cell is empty (null, undefined, "", or NaN)
        const isEmpty = (val === null || val === undefined || val === "" || val === "______" || (typeof val === 'number' && isNaN(val)));
                                                                                     
        if (isEmpty) {{
            return {{
                'color': 'white',
                'backgroundColor': '#0000FF', // Orange for "Attention"
                'fontWeight': 'bold'
            }};
        }}
    }}
    
    // If it's NOT a target, OR it's a target that has been filled:
    // Returning null tells AgGrid to use the standard theme styling.
    return null;
}};
""")

cell_dropdown_jscode = JsCode("""
function(params) {
    const row = params.node.rowIndex;
    const col = params.column.getColId();

    // PROP. CIRCUIT TEST (Row 30, Unnamed: 12)
    if (row === 30 && col === "Unnamed: 12") {
        return {
            component: 'agSelectCellEditor',
            params: { values: ['N/A', 'YES', 'NO'] }
        };
    }

    // RUN WITH INHIBITING OIL (Row 31, Unnamed: 12)
    if (row === 31 && col === "Unnamed: 12") {
        return {
            component: 'agSelectCellEditor',
            params: { values: ['YES', 'NO'] }
        };
    }

    // Default to a standard text editor for everything else
    return undefined;
}
""")

# Define a function to display the grid and capture changes
def display_ag_grid(df):
    # Build grid options
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(editable=True,
                                cellStyle=highlight_jscode,
                                cellEditorSelector=cell_dropdown_jscode)
    
    gb.configure_grid_options(singleClickEdit=True)

    # gb.configure_grid_options(domLayout='autoHeight')
    
    #gb.configure_grid_options(domLayout='autoWidth')

    # gb.configure_column("Rev. 5",
    #                     cellRenderer=oil_details_renderer,
    #                     autoheight=True,
    #                     width=600,
    #                     editable=True)
    
    grid_options = gb.build()

    # with st.container(border=True):
    grid_response = AgGrid(
        df, 
        gridOptions=grid_options,
        data_return_mode=DataReturnMode.AS_INPUT,
        update_mode=GridUpdateMode.VALUE_CHANGED,
        # update_mode=GridUpdateMode.MANUAL,
        fit_columns_on_grid_load=True,
        allow_unsafe_jscode=True,
        # height=400
        )

    return grid_response


def copy_to_excel(df, path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active 

    # Define Styles
    red_font = InlineFont(color="FF0000") # Red
    black_font = InlineFont(color="000000") # Black
    red_underscore_font = InlineFont(color="FF0000", u="single") # Red + Underscore
    
    # Alignment for the merged rows
    center_alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in df.iterrows():
        excel_row = r_idx + 2 
        
        # --- 1. OIL LEVEL ROWS (Excel 16 & 21 / df index 14 & 19) ---
        if r_idx in [14, 19]:
            # Static part
            prefix = "All readings above are to be listed when the engine has reached operating limits. Shut down check oil screen. Oil Level: "
            user_val = str(row.get("Unnamed: 18", "")).strip()
            suffix = " qt."
            
            # Build Rich Text
            rich_text = CellRichText([
                TextBlock(black_font, prefix),
                TextBlock(red_underscore_font, user_val if user_val else "____"),
                TextBlock(red_font, suffix)
            ])
            
            target_cell = ws.cell(row=excel_row, column=1)
            target_cell.value = rich_text
            target_cell.alignment = center_alignment
            
            # Clear other cells for merge
            for c_idx in range(2, len(row) + 1):
                ws.cell(row=excel_row, column=c_idx).value = None
            continue

        # --- 2. OIL DETAILS ROW (Excel 34 / df index 32) ---
        elif r_idx == 32:
            # We map out the specific fragments based on your Unnamed columns
            # Using the format: Text Piece -> User Input Column
            fragments = [
                ("Oil Details. Type: ", "Unnamed: 2"),
                ("  Qty in: ", "Unnamed: 4"),
                (" qt  Qty lost: ", "Unnamed: 6"),
                (" qt  Qty added: ", "Unnamed: 8"),
                (" qt  Qty out: ", "Unnamed: 10"),
                ("  Qty total used: ", "Unnamed: 12"),
                (" qt  Consumption Spec: ", "Unnamed: 15"),
                (" qt/hr  Actual: ", "Unnamed: 18"),
                (" qt/hr.", None)
            ]
            
            rich_blocks = []
            for text, col_name in fragments:
                rich_blocks.append(TextBlock(red_font, text))
                if col_name:
                    val = str(row.get(col_name, "")).strip()
                    rich_blocks.append(TextBlock(red_underscore_font, val if val else "____"))
            
            target_cell = ws.cell(row=excel_row, column=1)
            target_cell.value = CellRichText(rich_blocks)
            target_cell.alignment = center_alignment
            
            for c_idx in range(2, len(row) + 1):
                ws.cell(row=excel_row, column=c_idx).value = None
            continue

        # --- 3. DROPDOWN ROWS (32 & 33) ---
        elif excel_row in [32, 33]:
            # Normal logic for standard text placement, then apply centering
            current_label = str(row.get("Unnamed: 8", "")).split(":")[0]
            user_selection = str(row.get("Unnamed: 12", "")).strip()
            new_value = f"{current_label}: {user_selection}"
            
            cell = ws.cell(row=excel_row, column=9)
            cell.value = new_value
            cell.alignment = center_alignment
            
            for c_idx, value in enumerate(row):
                if (c_idx + 1) != 9 and (c_idx + 1) != 13:
                    ws.cell(row=excel_row, column=c_idx + 1).value = value
            continue

        # --- 4. DEFAULT ROWS ---
        for c_idx, value in enumerate(row):
            ws.cell(row=excel_row, column=c_idx + 1).value = value

    # --- MERGING ---
    required_merges = [
    "A1:T2", "A3:T3",
    "B4:G4", "B5:G5", "B6:G6", "B7:G7",
    "H4:I4", "H5:I5", "H6:I6", "H7:I7",
    "J4:N4", "J5:N5", "J6:N6", "J7:N7",
    "R4:T4", "R5:T5", "R6:T6", "R7:T7",
    "O4:Q4", "O5:Q5", "O6:Q6", "O7:Q7",
    "B8:C8", "E8:F8", "G8:H8", "I8:N8", "O8:Q8",
    "I9:N9", "O9:T9",
    "A16:T16","A21:T21",
    "A22:B22", "A23:B23", "E22:F22", "E23:F23", "M23:N23"
    "I22:J22", "I23:J23", "M22:N22", "O22:P22", "O23:P23", 
    "R22:S22", "R23:S23",
    "C22:D22", "C23:D23",
    "G22:H22", "G23:H23",
    "K22:L22", "K23:L23",
    "A25:B25", "A26:B26", "A27:B27", "A28:B28", "A29:B29",
    "G25:H25", "G27:H27", "G29:H29", "K29:L29",
    "O25:Q25", "O26:Q26", "O27:P27", "O28:R28", "O29:R29",
    "P31:R31", "O33:T33",
    "I25:J25", "I27:J27", "I29:J29",
    "I32:N32", "I33:N33",
    "P32:T32",
    "L25:M25", "L27:M27", "M29:N29",
    "R25:T25", "R26:S26", "S28:T28", "S29:T29",
    "A31:B31", "A32:B32", "A33:B33",
    "A34:T34"
    ]

    for merge_range in required_merges:
        try:
            ws.merge_cells(merge_range)
            # After merging, ensure the top-left cell is centered
            first_cell = merge_range.split(":")[0]
            ws[first_cell].alignment = center_alignment
        except Exception:
            pass

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --- 2. CALLBACK FUNCTION ---
def save_excel_callback(buffer, file_name):
    """Callback to save the excel buffer to a specific local directory."""
    try:
        if not file_name or file_name.strip() == ".xlsx":
            st.error("Please enter a valid file name.")
            return

        # Ensure the directory exists
        if not os.path.exists(save_xlxs_path):
            os.makedirs(save_xlxs_path)

        # Construct full path
        full_save_path = os.path.join(save_xlxs_path, file_name)

        # Write buffer to disk
        with open(full_save_path, "wb") as f:
            f.write(buffer.getbuffer())
        
        st.toast(f"File successfully saved to: {full_save_path}")

    except Exception as e:
        st.error(f"Error saving file: {str(e)}")

def trigger_save_flow_callback(excel_buffer, excel_path):

    if num_blades == 3:
        # Persistent switch to open the dialog
        st.session_state.show_prop_dialog = True
        # Initalize the internal dialog step
        if "dialog_step" not in st.session_state:
            st.session_state.dialog_step = "ask_change"
    else:
        # Save the testsheet
        save_excel_callback(excel_buffer, excel_path)
        st.toast("Testsheet saved successfully")

@st.dialog("Propeller Settings Verification", width="large")
def handle_save_with_prop_check(prop_name, save_callback_function, excel_buffer, excel_path):
    """
    This dialog handles the following:
    1. Asks if the operator has changed prop length
    2. If yes, show the data editor, updates JSON, then proceeds to save test sheet
    3. If no, bypass directly to saving test sheet
    """

    # Ask if operator has changed the length
    if st.session_state.dialog_step == "ask_change":
        st.write("### Did you adjust the prop length setting during this run?")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Yes", type="primary", use_container_width=True):
                st.session_state.dialog_step = "show_form"
                st.rerun()
        with col2:
            if st.button("No", type="primary", use_container_width=True):
                # No changes made, proceed to save test sheet
                save_callback_function(excel_buffer, excel_path)
                st.success("Testsheet saved")
                time.sleep(1.5)
                # Clear dialog state before rerun
                st.session_state.show_prop_dialog = False
                # Set tracking flag
                st.session_state.prop_log_finalized = True
                del st.session_state.dialog_step
                st.rerun()
    
    # Display the adjustable prop form and data editor
    elif st.session_state.dialog_step == "show_form":
        st.markdown(f"### {prop_name} Propeller Settings")
        form_info = st.info("Add a new row at the bottom of the table for each new entry")
        adjustable_logs = load_adjustable_config()
        raw_data = adjustable_logs.get(prop_name, [])

        if not raw_data:
            df = pd.DataFrame([{
                "Work Order": "",
                "Engine Model": "",
                "Horsepower": 0,
                "Date": date.today(),
                "Length": 0.0
            }])
        else:
            df = pd.DataFrame(raw_data)
            if "Date" in df.columns:
                df["Date"] = pd.to_datetime(df["Date"]).dt.date

        edited_df = st.data_editor(
            df,
            column_config={
                "Work Order": st.column_config.TextColumn("Work Order", width="medium"),
                "Engine Model": st.column_config.TextColumn("Engine Model"),
                "Horsepower": st.column_config.NumberColumn("Horsepower", min_value=0),
                "Date": st.column_config.DateColumn("Date",
                                                    min_value=date(2001, 1, 1),
                                                    max_value=date(2100, 1, 1),
                                                    format="YYYY-MM-DD",
                                                    required=True),
                "Length": st.column_config.NumberColumn("Length (inches)")                            
            },
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            key=f"setting_editor_{prop_name}"
        )

        if st.button(f"Update Log & Save Testsheet", type="primary", use_container_width=True):
            # Process dataframe for JSON compatibility
            final_df = edited_df.copy()
            if "Date" in final_df.columns:
                final_df["Date"] = final_df["Date"].astype(str)

            # Save back to main dictionary
            adjustable_logs[prop_name] = final_df.to_dict('records')
            save_adjustable_config(adjustable_logs)

            # Set tracking flag
            st.session_state.prop_log_finalized = True

            # Save the testsheet
            save_callback_function(excel_buffer, excel_path)

            st.success("Prop log updated and test sheet saved")
            time.sleep(1.5)

            # Clear session state and close dialog
            del st.session_state.dialog_step
            st.rerun()

def main():
    """
    Main function to run the Streamlit app.
    """
    # st.write(st.session_state.time_step)
    df = get_dataframe()

    st.markdown("""
    <style>
    
           /* Remove blank space at top and bottom */ 
           .block-container {
               padding-top: 15px;
               padding-bottom: 15px;
               padding-left: 25px;
               padding-right: 25px; 
            }
           
           
    </style>
    """, unsafe_allow_html=True)
    
    is_active = st.session_state.process_running

    # st.write(st.session_state.selected_rpm)
    
    # Using session state to track components
    if 'components_created' not in st.session_state:
        st.session_state['components_created'] = False

    if not st.session_state['components_created']:

        col1, col2, col3, col4, col5 = st.columns([1, 1, 1, 1, 1.5], gap="small", vertical_alignment='center')  # Create three columns side by side
        with col1:
            with st.container(height=130):
                cola, colb, colc = st.columns([1.5, 6, 0.2])
                with cola:
                    st.write("")
                with colb:
                    performance_fig = plot_performance_metrics()
                    st.plotly_chart(performance_fig, use_container_width=False)
                with colc:
                    st.write("")
                

        with col2:
            with st.container(height=130):
                cola, colb, colc = st.columns([1.5, 6, 0.2])
                with cola:
                    st.write("")
                with colb:
                    rpm_fig = animate_rpm(df, engine_type)
                    st.plotly_chart(rpm_fig, use_container_width=False)
                with colc:
                    st.write("")

        with col5:
            with st.container(height=130):
                cola, colb, colc, cold, cole, colf = st.columns([0.01, 1.3, 1.3, 1.3, 1.3, 0.01], vertical_alignment='top')
                with colb:
                    st.write("")
                    st.write("")
                    # We define a wrapper if we need to call multiple functions
                    def handle_start():
                        start_process()
                        start_button_handler(df)

                    st.button("START", on_click=handle_start, disabled=is_active)
                with colc:
                    st.write("")
                    st.write("")
                    st.button("SNAP", on_click=snap_button_handler, args=(df,), disabled=not is_active)
                with cold:
                    st.write("")
                    st.write("")
                    st.button("UNDO", on_click=undo_button_handler, args=(df,))
                with cole:
                    st.write("")
                    st.write("")
                    # FIX: End callback
                    def handle_end():
                        
                        try:
                            # Fetch names from session state
                            db_local_str = str(st.session_state.db_name)
                            collection_local_str = str(st.session_state.collection_name)

                            # Run loop
                            local_loop = asyncio.new_event_loop()
                            asyncio.set_event_loop(local_loop)

                            # Pass the connection URI and strings
                            running_df, engine_run_time_hhmm = local_loop.run_until_complete(calculate_engine_run_time(uri, db_local_str, collection_local_str))
                            local_loop.close()
                            # st.write(running_df)
                            st.session_state.run_time = engine_run_time_hhmm
                            st.toast(f"Engine Runtime Calculated: {engine_run_time_hhmm}")
                        except Exception as e:
                            st.error(f"Error calculating actual engine run time: {e}")
                        end_button_handler(df)
                        # time.sleep(1.5)
                        stop_process()
                        
                        
                    st.button("END", on_click=handle_end, disabled=not is_active)

        with col3:
            with st.container(height=130):
                cola, colb, colc = st.columns([1.5, 6, 0.2])
                with cola:
                    st.write("")
                with colb:
                    manifold_pressure_fig = animate_manifold_pressure(df, engine_type)
                    st.plotly_chart(manifold_pressure_fig, use_container_width=False)
                with colc:
                    st.write("")

        with col4:
            with st.container(height=130):
                cola, colb, colc = st.columns([1.5, 6, 0.2])
                with cola:
                    st.write("")
                with colb:
                    horsepower_fig = animate_hp(df, engine_type)
                    st.plotly_chart(horsepower_fig, use_container_width=False)
                with colc:
                    st.write("")

        with st.container(height=450):
            col6, col7 = st.columns(2, vertical_alignment='center')
            with col6:
                if df.empty:
                    st.write("Turn on MASTER, and click the START BUTTON to populate the plot")
                else:
                    temps_bar_plot = plot_egt_cht(df, engine_type)
                    st.plotly_chart(temps_bar_plot, use_container_width=False)

            with col7:
                display_metrics(df, engine_type)


        # Display the engine test sheet dataframe using streamlit aggrid and update session state
        grid_response = display_ag_grid(st.session_state.test_sheet)
        if grid_response['data'] is not None:
            updated_data = pd.DataFrame(grid_response['data'])
            has_changes = False

            for i in range(len(updated_data)):
                for col in updated_data.columns:
                    old_val = st.session_state.test_sheet.iloc[i, st.session_state.test_sheet.columns.get_loc(col)]
                    new_val = updated_data.iloc[i][col]
                    if old_val != new_val:
                        st.session_state.test_sheet.iloc[i, st.session_state.test_sheet.columns.get_loc(col)] = new_val
                        has_changes = True

                # Save tp database if cell modification is complete
                if has_changes:
                    persist_test_sheet()
           

        col8, col9, col10 = st.columns([1.2, 1.2, 1], vertical_alignment='bottom')
        with col8:
            pass
        with col9:

            # Prompt the user for a file name
            file_name_input = st.text_input("Confirm File Name (WO#-Testsheet)", f"{work_order}-Testsheet")
            output_file_name = f"{file_name_input}.xlsx"

            colf, colg, colh = st.columns([1.2, 1.2, 1], vertical_alignment='bottom')
            with colf:
                pass
            with colg:
                buffer = copy_to_excel(st.session_state.test_sheet, xlxs_copy_path)

                if "show_prop_dialog" not in st.session_state:
                    st.session_state.show_prop_dialog = False

                if st.session_state.show_prop_dialog and not st.session_state.prop_log_finalized:
                    handle_save_with_prop_check(prop_type, save_excel_callback, buffer, output_file_name)

                st.button(
                    label="Save to Folder",
                    type="primary",
                    use_container_width=True,
                    on_click=trigger_save_flow_callback,
                    args=(buffer, output_file_name)
                )
                
            with colh:
                pass
        with col10:
            pass

if __name__ == "__main__":
    main()
    # subprocess.Popen(["python", r"C:\Users\theca\Documents\Victor\Aerotec\MVP 50\Code\AerotecTestCell\Parse_data.py"])

